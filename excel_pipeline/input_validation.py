"""Structured validation of user-uploaded input files.

Each Sunrise and Astra input file gets a small validator that returns a
list of ``ValidationIssue`` describing every detected problem. The
``ValidationReport`` aggregates issues across files and tells the UI
whether the pipeline can run (no errors) or must be blocked.

Design rule: validators never crash. If a file cannot be opened or has
an unexpected layout the validator emits an ``error`` issue and returns
— the pipeline is then blocked from running. Warnings are
informational and never block; they're surfaced to the user so they can
decide whether to fix the input or run anyway.

Two entry points are exposed to the pages:

- ``validate_sunrise_inputs(paths, year, semester)``
- ``validate_astra_inputs(paths)``

Both return a ``ValidationReport``. After a successful run callers can
persist the report with ``report.save(run_dir / 'validation.json')``.
"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


# ===========================================================================
# Data structures
# ===========================================================================
@dataclass
class ValidationIssue:
    """A single detected problem on a single input file."""

    file: str
    severity: str  # "error" or "warning"
    location: str  # "file", "sheet", "header col 3", "row 12", ...
    code: str  # short machine-readable code, e.g. "MISSING_SHEET"
    message: str


@dataclass
class ValidationReport:
    """Aggregated result of running all validators for one pipeline phase."""

    errors: list[ValidationIssue] = field(default_factory=list)
    warnings: list[ValidationIssue] = field(default_factory=list)

    @property
    def is_blocking(self) -> bool:
        return bool(self.errors)

    def add(self, issue: ValidationIssue) -> None:
        if issue.severity == "error":
            self.errors.append(issue)
        else:
            self.warnings.append(issue)

    def extend(self, issues: list[ValidationIssue]) -> None:
        for i in issues:
            self.add(i)

    def to_dict(self) -> dict[str, Any]:
        return {
            "errors": [asdict(i) for i in self.errors],
            "warnings": [asdict(i) for i in self.warnings],
        }

    def save(self, path: Path) -> None:
        """Persist the report as JSON next to the run outputs."""
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        Path(path).write_text(
            json.dumps(self.to_dict(), indent=2, ensure_ascii=False),
            encoding="utf-8",
        )


# ===========================================================================
# Suffix constants (kept in sync with pipeline.py)
# ===========================================================================
SUNRISE_CEDED_SUFFIX = "_Ceded"
SUNRISE_ASSUMED_SUFFIX = "_Assumed"
TRANSCODIFICA_SUFFIX = "Transcodifica_aggregazione_GOC_H_NH"
PAYMENT_PATTERNS_SUFFIX = "Payment_Patterns_&_Risk_Adjustments"

PROJECTION_PARAMETERS_SUFFIX = "PROJECTION_PARAMETERS_ENTITY"
MP_GOC_SEG_SUFFIX = "MP_GOC_SEG"
MP_GOC_SUFFIX = "MP_GOC"
ACTUARIAL_AOM_IMPACT_SUFFIX = "ACTUARIAL_AOM_IMPACT"
CURVE_ID_PARAM_SUFFIX = "CURVE_ID_PARAM"


# ===========================================================================
# Shared low-level helpers
# ===========================================================================
_DATE_RE = re.compile(r"(\d{4})\.(\d{2})\.(\d{2})")


def _files_ending_with(paths: list[Path], suffix: str) -> list[Path]:
    return [p for p in paths if p.stem.endswith(suffix)]


def _sniff_csv_delimiter(sample: str, fallback: str = ";") -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t").delimiter
    except csv.Error:
        return fallback


def _read_csv_rows(path: Path) -> list[list[str]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        delim = _sniff_csv_delimiter(sample)
        return [list(r) for r in csv.reader(f, delimiter=delim)]


def _normalize_header(h: Any) -> str:
    """Lower-cased + stripped form for case-insensitive header comparison."""
    if h is None:
        return ""
    return str(h).strip().casefold()


def _header_issue(
    file: str, actual: Any, expected: str, col_idx: int,
) -> ValidationIssue | None:
    """Compare one header cell. Return None if exact match, warning if only
    case differs, error if the value is different (or missing).
    """
    if _normalize_header(actual) == _normalize_header(expected):
        if actual is not None and str(actual).strip() != expected:
            return ValidationIssue(
                file=file, severity="warning",
                location=f"header col {col_idx + 1}",
                code="HEADER_CASE_MISMATCH",
                message=(
                    f'Header {col_idx + 1} is "{actual}", expected '
                    f'"{expected}" (matched case-insensitively).'
                ),
            )
        return None
    return ValidationIssue(
        file=file, severity="error",
        location=f"header col {col_idx + 1}",
        code="HEADER_MISMATCH",
        message=(
            f'Header {col_idx + 1} is "{actual}", expected "{expected}".'
        ),
    )


def _expected_sunrise_dates(year: int, semester: int) -> tuple[str, str]:
    """Return ``(current_date, previous_date)`` for the Sunrise inputs."""
    if semester == 1:
        month_day = "06.30"
    elif semester == 2:
        month_day = "12.31"
    else:
        raise ValueError(f"semester must be 1 or 2, got {semester}")
    return (f"{year}.{month_day}", f"{year - 1}.{month_day}")


# ===========================================================================
# Per-file validators — Sunrise
# ===========================================================================
CEDED_ASSUMED_HEADERS = (
    "GOC", "ANNO", "PERIMETRO", "SINISTRI", "RISERVA_SINISTRI",
)
_SAMPLE_ROW_LIMIT = 200  # how many data rows to type-check per file


def validate_ceded_assumed_file(path: Path) -> list[ValidationIssue]:
    """L0/L1/L2 checks for a single ``_Ceded`` / ``_Assumed`` file."""
    issues: list[ValidationIssue] = []
    fname = path.name
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return [ValidationIssue(
            file=fname, severity="error", location="file",
            code="CANNOT_OPEN", message=f"Cannot open file: {e}",
        )]
    try:
        if "Input_Sunrise" not in wb.sheetnames:
            issues.append(ValidationIssue(
                file=fname, severity="error", location="sheet",
                code="MISSING_SHEET",
                message=(
                    f"Sheet 'Input_Sunrise' not found. Available: "
                    f"{wb.sheetnames}"
                ),
            ))
            return issues
        ws = wb["Input_Sunrise"]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            issues.append(ValidationIssue(
                file=fname, severity="error", location="header",
                code="EMPTY_FILE",
                message="Sheet 'Input_Sunrise' has no rows.",
            ))
            return issues
        # L1: header
        for i, expected in enumerate(CEDED_ASSUMED_HEADERS):
            actual = header[i] if i < len(header) else None
            iss = _header_issue(fname, actual, expected, i)
            if iss:
                issues.append(iss)
        # L2: type sanity on first N data rows
        n_rows = 0
        non_zero_seen = False
        row_idx = 1  # header is row 1
        for row in rows_iter:
            row_idx += 1
            if n_rows >= _SAMPLE_ROW_LIMIT:
                break
            goc = row[0] if len(row) > 0 else None
            anno = row[1] if len(row) > 1 else None
            sin = row[3] if len(row) > 3 else None
            ris = row[4] if len(row) > 4 else None
            if goc is None or str(goc).strip() == "":
                continue
            n_rows += 1
            try:
                int(anno)
            except (TypeError, ValueError):
                issues.append(ValidationIssue(
                    file=fname, severity="error",
                    location=f"row {row_idx}",
                    code="INVALID_ANNO",
                    message=f"ANNO value {anno!r} is not an integer.",
                ))
            for cell, name in [(sin, "SINISTRI"), (ris, "RISERVA_SINISTRI")]:
                if cell is None or cell == "":
                    continue
                try:
                    if float(cell) != 0.0:
                        non_zero_seen = True
                except (TypeError, ValueError):
                    issues.append(ValidationIssue(
                        file=fname, severity="error",
                        location=f"row {row_idx}",
                        code="INVALID_NUMERIC",
                        message=f"{name} value {cell!r} is not numeric.",
                    ))
        if n_rows > 0 and not non_zero_seen:
            issues.append(ValidationIssue(
                file=fname, severity="warning", location="data",
                code="ALL_ZERO_VALUES",
                message=(
                    "Sample rows have SINISTRI and RISERVA_SINISTRI all zero. "
                    "No GoC from this file will appear in MP_ModelPoint."
                ),
            ))
    finally:
        wb.close()
    return issues


TRANSCODIFICA_HEADERS = ("GOC_ID", "Aggregation1", "Aggregation2", "H-NH")


def validate_transcodifica_file(path: Path) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    fname = path.name
    suffix = path.suffix.lower()
    try:
        if suffix in {".xlsx", ".xlsm"}:
            wb = openpyxl.load_workbook(path, data_only=True)
            try:
                ws = wb.active
                rows = [
                    [ws.cell(row=r, column=c).value for c in range(1, 5)]
                    for r in range(1, ws.max_row + 1)
                ]
            finally:
                wb.close()
        else:
            rows = _read_csv_rows(path)
    except Exception as e:
        return [ValidationIssue(
            file=fname, severity="error", location="file",
            code="CANNOT_OPEN", message=f"Cannot open file: {e}",
        )]
    if not rows:
        return [ValidationIssue(
            file=fname, severity="error", location="header",
            code="EMPTY_FILE", message="File is empty.",
        )]
    header = rows[0]
    if len(header) < 4:
        issues.append(ValidationIssue(
            file=fname, severity="error", location="header",
            code="MISSING_HNH_COLUMN",
            message=(
                f"Transcodifica header has {len(header)} columns, expected 4. "
                "The 4th column (H-NH) is required to compute the Health "
                "perimeter."
            ),
        ))
    for i, expected in enumerate(TRANSCODIFICA_HEADERS):
        actual = header[i] if i < len(header) else None
        iss = _header_issue(fname, actual, expected, i)
        if iss:
            issues.append(iss)
    # L2: scan H-NH column and GoC lengths
    h_count = 0
    seen_goc: set[str] = set()
    valid_flags = {"H", "NH"}
    for row_idx, row in enumerate(rows[1:], start=2):
        goc = row[0] if len(row) > 0 else None
        flag = row[3] if len(row) > 3 else None
        if goc is None or str(goc).strip() == "":
            continue
        goc_s = str(goc).strip()
        if len(goc_s) != 11:
            issues.append(ValidationIssue(
                file=fname, severity="warning",
                location=f"row {row_idx}",
                code="GOC_LENGTH",
                message=(
                    f"GOC_ID '{goc_s}' has {len(goc_s)} chars; the pipeline "
                    "matches GoCs on the first 11 chars."
                ),
            ))
        if goc_s in seen_goc:
            issues.append(ValidationIssue(
                file=fname, severity="warning",
                location=f"row {row_idx}",
                code="DUPLICATE_GOC",
                message=f"GOC_ID '{goc_s}' appears more than once; the "
                "duplicate will be ignored.",
            ))
        else:
            seen_goc.add(goc_s)
        flag_s = str(flag).strip().upper() if flag is not None else ""
        if flag_s == "H":
            h_count += 1
        elif flag_s == "":
            pass  # blank tolerated
        elif flag_s not in valid_flags:
            issues.append(ValidationIssue(
                file=fname, severity="error",
                location=f"row {row_idx}",
                code="INVALID_HNH",
                message=(
                    f"H-NH value {flag!r} is not 'H' or 'NH'. The Health "
                    "perimeter will not include this row."
                ),
            ))
    if h_count == 0:
        issues.append(ValidationIssue(
            file=fname, severity="warning", location="data",
            code="NO_H_FLAGGED_GOCS",
            message=(
                "No row is flagged 'H'. The Health perimeter will be empty "
                "and MP_GOC_SEG will not rewrite any 'P&C' to 'HLTH_PC'."
            ),
        ))
    return issues


def validate_payment_patterns_file(
    path: Path, sheet_suffix: str = "AAI",
) -> list[ValidationIssue]:
    """Verify the Payment_Patterns_&_Risk_Adjustments workbook contains
    the two sheets ``ra_<suffix>_REINS`` and ``pp_<suffix>_REINS`` with
    the expected header layout."""
    issues: list[ValidationIssue] = []
    fname = path.name
    ra_sheet = f"ra_{sheet_suffix}_REINS"
    pp_sheet = f"pp_{sheet_suffix}_REINS"
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return [ValidationIssue(
            file=fname, severity="error", location="file",
            code="CANNOT_OPEN", message=f"Cannot open file: {e}",
        )]
    try:
        if ra_sheet not in wb.sheetnames:
            issues.append(ValidationIssue(
                file=fname, severity="error", location="sheet",
                code="MISSING_SHEET",
                message=(
                    f"Sheet '{ra_sheet}' not found. Available: "
                    f"{wb.sheetnames}"
                ),
            ))
        if pp_sheet not in wb.sheetnames:
            issues.append(ValidationIssue(
                file=fname, severity="error", location="sheet",
                code="MISSING_SHEET",
                message=(
                    f"Sheet '{pp_sheet}' not found. Available: "
                    f"{wb.sheetnames}"
                ),
            ))
        if ra_sheet in wb.sheetnames:
            ws = wb[ra_sheet]
            header = next(ws.iter_rows(values_only=True), None) or ()
            if len(header) < 8 or header[6] is None:
                issues.append(ValidationIssue(
                    file=fname, severity="error",
                    location=f"sheet '{ra_sheet}', header",
                    code="MISSING_GOC_COLUMN",
                    message=(
                        f"Column G ('GoC') is missing on sheet '{ra_sheet}'."
                    ),
                ))
            year_pattern = re.compile(r"^(FY|HY)_?\d{4}$")
            year_cols = [
                str(h).strip() for h in header[7:]
                if h is not None and year_pattern.match(str(h).strip())
            ]
            if not year_cols:
                issues.append(ValidationIssue(
                    file=fname, severity="error",
                    location=f"sheet '{ra_sheet}', header",
                    code="MISSING_YEAR_COLUMNS",
                    message=(
                        "No FY_yyyy / HY_yyyy column header found after "
                        "column G."
                    ),
                ))
        if pp_sheet in wb.sheetnames:
            ws = wb[pp_sheet]
            header = next(ws.iter_rows(values_only=True), None) or ()
            if len(header) < 4 or header[2] is None or header[3] is None:
                issues.append(ValidationIssue(
                    file=fname, severity="error",
                    location=f"sheet '{pp_sheet}', header",
                    code="MISSING_KEY_COLUMNS",
                    message=(
                        "Columns C ('GoC') and D ('Year') must be "
                        f"populated on the header of sheet '{pp_sheet}'."
                    ),
                ))
            expected = [str(i) for i in range(23)]
            found = []
            for h in header[4:27]:
                if h is None:
                    continue
                found.append(str(h).strip())
            missing = [e for e in expected if e not in found]
            if missing:
                issues.append(ValidationIssue(
                    file=fname, severity="error",
                    location=f"sheet '{pp_sheet}', header",
                    code="MISSING_PP_COLUMNS",
                    message=(
                        "Some of the 23 data columns ('0'..'22') are missing "
                        f"after column D. Missing: {missing[:5]}"
                        + ("..." if len(missing) > 5 else "")
                    ),
                ))
    finally:
        wb.close()
    return issues


# ===========================================================================
# Per-file validators — Astra
# ===========================================================================
PROJECTION_PARAMETER_TARGETS = (
    "CF_TIMESTEP", "REPORTING_MONTH",
    "FX_OPENING_DATE", "FX_AVERAGE_DATE",
    "FX_CLOSING_DATE", "FX_REPORTING_DATE",
)


def _read_table_for_validation(path: Path) -> list[list[Any]]:
    """Format-agnostic reader: CSV (sniffed) or XLSX (active sheet)."""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            return [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    return _read_csv_rows(path)


def validate_projection_parameters_entity_file(
    path: Path,
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    fname = path.name
    try:
        rows = _read_table_for_validation(path)
    except Exception as e:
        return [ValidationIssue(
            file=fname, severity="error", location="file",
            code="CANNOT_OPEN", message=f"Cannot open file: {e}",
        )]
    if not rows:
        return [ValidationIssue(
            file=fname, severity="error", location="header",
            code="EMPTY_FILE", message="File is empty.",
        )]
    header = rows[0]
    for i, expected in enumerate(["PARAMETER", "VALUE"]):
        actual = header[i] if i < len(header) else None
        iss = _header_issue(fname, actual, expected, i)
        if iss:
            issues.append(iss)
    seen: set[str] = set()
    for row in rows[1:]:
        if not row:
            continue
        param = row[0]
        if param is not None:
            seen.add(str(param).strip().upper())
    for target in PROJECTION_PARAMETER_TARGETS:
        if target not in seen:
            issues.append(ValidationIssue(
                file=fname, severity="warning", location="data",
                code="MISSING_PARAMETER",
                message=(
                    f"PARAMETER '{target}' is missing; the corresponding "
                    "VALUE will not be rewritten."
                ),
            ))
    return issues


def validate_mp_goc_seg_file(path: Path) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    fname = path.name
    try:
        rows = _read_table_for_validation(path)
    except Exception as e:
        return [ValidationIssue(
            file=fname, severity="error", location="file",
            code="CANNOT_OPEN", message=f"Cannot open file: {e}",
        )]
    if not rows:
        return [ValidationIssue(
            file=fname, severity="error", location="header",
            code="EMPTY_FILE", message="File is empty.",
        )]
    header = rows[0]
    expected_headers = (
        "GOC_SEG_ID", "GOC_ID", "SEG_ID", "ALLOCATION_RATIO",
    )
    for i, expected in enumerate(expected_headers):
        actual = header[i] if i < len(header) else None
        iss = _header_issue(fname, actual, expected, i)
        if iss:
            issues.append(iss)
    # L2: spot-check GOC_ID length
    checked = 0
    for row_idx, row in enumerate(rows[1:], start=2):
        if checked >= _SAMPLE_ROW_LIMIT:
            break
        goc = row[1] if len(row) > 1 else None
        if goc is None or str(goc).strip() == "":
            continue
        checked += 1
        if len(str(goc).strip()) < 11:
            issues.append(ValidationIssue(
                file=fname, severity="warning",
                location=f"row {row_idx}",
                code="GOC_ID_TOO_SHORT",
                message=(
                    f"GOC_ID '{goc}' has fewer than 11 chars; the row will "
                    "be skipped by the perimeter check."
                ),
            ))
    return issues


def validate_mp_goc_file(path: Path) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    fname = path.name
    try:
        rows = _read_table_for_validation(path)
    except Exception as e:
        return [ValidationIssue(
            file=fname, severity="error", location="file",
            code="CANNOT_OPEN", message=f"Cannot open file: {e}",
        )]
    if not rows:
        return [ValidationIssue(
            file=fname, severity="error", location="header",
            code="EMPTY_FILE", message="File is empty.",
        )]
    header = rows[0]
    if len(header) < 18:
        issues.append(ValidationIssue(
            file=fname, severity="error", location="header",
            code="NOT_ENOUGH_COLUMNS",
            message=(
                f"MP_GOC header has {len(header)} columns, at least 18 are "
                "required (cols A..R) so the skill can read AGGREG_2_ID."
            ),
        ))
        return issues
    if _normalize_header(header[2]) != _normalize_header("ANNUAL_COHORT"):
        issues.append(ValidationIssue(
            file=fname, severity="warning", location="header col 3",
            code="HEADER_MISMATCH",
            message=(
                f"Column C header is {header[2]!r}, expected 'ANNUAL_COHORT'."
            ),
        ))
    if _normalize_header(header[17]) != _normalize_header("AGGREG_2_ID"):
        issues.append(ValidationIssue(
            file=fname, severity="warning", location="header col 18",
            code="HEADER_MISMATCH",
            message=(
                f"Column R header is {header[17]!r}, expected 'AGGREG_2_ID'."
            ),
        ))
    # L2: spot-check ANNUAL_COHORT type
    checked = 0
    for row_idx, row in enumerate(rows[1:], start=2):
        if checked >= _SAMPLE_ROW_LIMIT:
            break
        cohort = row[2] if len(row) > 2 else None
        if cohort is None or cohort == "":
            continue
        checked += 1
        try:
            int(cohort)
        except (TypeError, ValueError):
            issues.append(ValidationIssue(
                file=fname, severity="warning",
                location=f"row {row_idx}",
                code="INVALID_COHORT_YEAR",
                message=(
                    f"ANNUAL_COHORT value {cohort!r} is not an integer; the "
                    "row will be skipped."
                ),
            ))
    return issues


def validate_actuarial_aom_impact_file(path: Path) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    fname = path.name
    try:
        rows = _read_table_for_validation(path)
    except Exception as e:
        return [ValidationIssue(
            file=fname, severity="error", location="file",
            code="CANNOT_OPEN", message=f"Cannot open file: {e}",
        )]
    if not rows:
        return [ValidationIssue(
            file=fname, severity="error", location="header",
            code="EMPTY_FILE", message="File is empty.",
        )]
    header = rows[0]
    if len(header) < 3:
        issues.append(ValidationIssue(
            file=fname, severity="error", location="header",
            code="NOT_ENOUGH_COLUMNS",
            message=(
                f"ACTUARIAL_AOM_IMPACT header has {len(header)} columns, "
                "at least 3 are required (GOC_ID, STEP_ID, '1')."
            ),
        ))
        return issues
    for i, expected in enumerate(["GOC_ID", "STEP_ID"]):
        actual = header[i] if i < len(header) else None
        iss = _header_issue(fname, actual, expected, i)
        if iss:
            issues.append(iss)
    return issues


def validate_curve_id_param_file(path: Path) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    fname = path.name
    try:
        rows = _read_table_for_validation(path)
    except Exception as e:
        return [ValidationIssue(
            file=fname, severity="error", location="file",
            code="CANNOT_OPEN", message=f"Cannot open file: {e}",
        )]
    if not rows:
        return [ValidationIssue(
            file=fname, severity="error", location="header",
            code="EMPTY_FILE", message="File is empty.",
        )]
    header = rows[0]
    for i, expected in enumerate(["GOC_ID", "VARIABLE_NAME"]):
        actual = header[i] if i < len(header) else None
        iss = _header_issue(fname, actual, expected, i)
        if iss:
            issues.append(iss)
    known_vars = {"CLOSING_CURVE_ID", "OPENING_CURVE_ID", "CREDITED_RATE_CURVE_ID"}
    seen_vars: set[str] = set()
    for row in rows[1:]:
        if len(row) < 2:
            continue
        var = row[1]
        if var is None:
            continue
        seen_vars.add(str(var).strip().upper())
    for missing_var in sorted(known_vars - seen_vars):
        issues.append(ValidationIssue(
            file=fname, severity="warning", location="data",
            code="MISSING_VARIABLE_NAME",
            message=(
                f"VARIABLE_NAME '{missing_var}' does not appear in the "
                "file; the corresponding skill branch will not write any "
                "column C value."
            ),
        ))
    return issues


# ===========================================================================
# Aggregate validators — called by the Streamlit pages
# ===========================================================================
def validate_sunrise_inputs(
    input_paths: list[Path],
    year: int,
    semester: int,
    sheet_suffix: str = "AAI",
) -> ValidationReport:
    """Validate the full set of Sunrise inputs (L0 + L1 + L2).

    ``sheet_suffix`` locates the Risk-Adjustment / Payment-pattern
    sheets inside the ``Payment_Patterns_&_Risk_Adjustments`` workbook
    (defaults to ``"AAI"`` so the historical layout keeps validating
    without a code change).
    """
    report = ValidationReport()

    # L0: required files
    transcodifica_files = _files_ending_with(input_paths, TRANSCODIFICA_SUFFIX)
    if not transcodifica_files:
        report.add(ValidationIssue(
            file="(missing)", severity="error", location="upload",
            code="MISSING_TRANSCODIFICA",
            message=(
                f"No upload ends with '{TRANSCODIFICA_SUFFIX}'. The "
                "master list is required."
            ),
        ))

    payment_files = _files_ending_with(input_paths, PAYMENT_PATTERNS_SUFFIX)
    if not payment_files:
        report.add(ValidationIssue(
            file="(missing)", severity="error", location="upload",
            code="MISSING_PAYMENT_PATTERNS",
            message=(
                f"No upload ends with '{PAYMENT_PATTERNS_SUFFIX}'. The "
                "Risk_Adjustment and Payment_pattern outputs need it."
            ),
        ))

    ceded_files = _files_ending_with(input_paths, SUNRISE_CEDED_SUFFIX)
    assumed_files = _files_ending_with(input_paths, SUNRISE_ASSUMED_SUFFIX)

    current_date, previous_date = _expected_sunrise_dates(year, semester)
    for date_str in (current_date, previous_date):
        ceded_for_date = [p for p in ceded_files if date_str in p.name]
        if not ceded_for_date:
            report.add(ValidationIssue(
                file="(missing)", severity="error", location="upload",
                code="MISSING_CEDED_FOR_DATE",
                message=(
                    f"No '_Ceded' file for date {date_str}. Upload at "
                    f"least one file containing {date_str} in its name."
                ),
            ))

    # L1 + L2: per-file
    for f in ceded_files + assumed_files:
        report.extend(validate_ceded_assumed_file(f))
    for f in transcodifica_files:
        report.extend(validate_transcodifica_file(f))
    for f in payment_files:
        report.extend(validate_payment_patterns_file(f, sheet_suffix))

    return report


def validate_astra_inputs(input_paths: list[Path]) -> ValidationReport:
    """Validate the full set of Astra inputs (L0 + L1 + L2)."""
    report = ValidationReport()

    suffixes_validators = [
        (PROJECTION_PARAMETERS_SUFFIX, validate_projection_parameters_entity_file),
        (MP_GOC_SEG_SUFFIX, validate_mp_goc_seg_file),
        (MP_GOC_SUFFIX, validate_mp_goc_file),
        (ACTUARIAL_AOM_IMPACT_SUFFIX, validate_actuarial_aom_impact_file),
        (CURVE_ID_PARAM_SUFFIX, validate_curve_id_param_file),
    ]

    for suffix, validator in suffixes_validators:
        matches = _files_ending_with(input_paths, suffix)
        if not matches:
            report.add(ValidationIssue(
                file="(missing)", severity="error", location="upload",
                code=f"MISSING_{suffix}",
                message=f"No upload ends with '{suffix}'.",
            ))
            continue
        if len(matches) > 1:
            report.add(ValidationIssue(
                file=", ".join(m.name for m in matches),
                severity="error", location="upload",
                code=f"DUPLICATE_{suffix}",
                message=(
                    f"Multiple uploads end with '{suffix}'. Keep only one."
                ),
            ))
            continue
        report.extend(validator(matches[0]))

    return report

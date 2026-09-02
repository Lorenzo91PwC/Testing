"""Excel skill — deterministic, typed operations on workbooks.

Every Excel transformation of the pipeline lives here as a plain,
typed Python function. The functions are pure with respect to their
inputs and outputs, so they are trivially testable via ``pytest``
against fixture workbooks under ``tests/``.

Adding a new transformation:
  1. Write a plain Python function below and type its arguments.
  2. Add a ``pytest`` in ``tests/`` that proves it works on a fixture.
  3. Call it from ``pipeline.py`` in the appropriate phase.
"""
from __future__ import annotations

import csv
from collections import OrderedDict
from collections.abc import Iterable
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook


# ===========================================================================
# Low-level helpers (not exposed as tools — used by other skill functions)
# ===========================================================================
def load_workbook(path: str) -> Workbook:
    """Open an Excel file for reading or editing."""
    return openpyxl.load_workbook(path)


def save_workbook(wb: Workbook, path: str) -> None:
    """Save a workbook to disk, creating parent dirs if needed."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def list_run_files(run_dir: Path) -> list[Path]:
    """List all .csv files in a run directory, sorted by name."""
    return sorted(run_dir.glob("*.csv"))


def _sniff_csv_delimiter(sample: str, fallback: str = ";") -> str:
    """Return the most likely CSV delimiter for ``sample``.

    Checked candidates: ``;``, ``,`` and ``\\t``. Falls back to
    ``fallback`` (``;`` by default) when ``csv.Sniffer`` cannot decide.
    """
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        return dialect.delimiter
    except csv.Error:
        return fallback


def _read_csv_table(path: str) -> list[list[Any]]:
    """Read a CSV into a list of rows; empty cells are normalised to ``None``.

    The field separator is auto-detected among ``;``, ``,`` and ``\\t``
    by sniffing the first part of the file. Files written by
    ``_write_csv_rows`` (semicolon-separated) are detected as-is; files
    exported from Excel with a different list separator — common for
    user-uploaded inputs like MP_GOC_SEG — are also handled correctly.
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        delim = _sniff_csv_delimiter(sample)
        reader = csv.reader(f, delimiter=delim)
        return [[(cell if cell != "" else None) for cell in row] for row in reader]


def _format_csv_value(value: Any, decimal_sep: str = ",") -> str:
    """Convert a Python value to its CSV string form.

    ``None`` -> empty string. ``float`` -> ``repr(value)`` with the
    default Python decimal point replaced by ``decimal_sep`` when it
    differs (default ``,`` for the Sunrise / Italian-Excel convention;
    Astra passes ``.``). ``bool`` is preserved as ``True/False``.
    Everything else (``int``, ``str``, ...) is passed through ``str``.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, float):
        s = repr(value)
        return s if decimal_sep == "." else s.replace(".", decimal_sep)
    return str(value)


def _write_csv_rows(
    path: str,
    rows: Iterable[Iterable[Any]],
    field_sep: str = ";",
    decimal_sep: str = ",",
) -> None:
    """Write rows to a CSV file (UTF-8 with BOM), creating parent dirs.

    Defaults follow the **Sunrise** output convention: ``;`` between
    fields, ``,`` as decimal separator — opens natively in
    Italian-locale Excel. Astra skills pass ``field_sep=","`` and
    ``decimal_sep="."`` for the US-format CSVs the downstream Astra
    system expects; the module-level ``_ASTRA_CSV_KWARGS`` bundles both.
    """
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=field_sep)
        for row in rows:
            writer.writerow([_format_csv_value(v, decimal_sep) for v in row])


# Astra downstream expects US-format CSVs: `,` between fields, `.` as
# decimal separator. Bundle both flags so every Astra skill can pass
# `**_ASTRA_CSV_KWARGS` without repeating itself.
_ASTRA_CSV_KWARGS = {"field_sep": ",", "decimal_sep": "."}


# ===========================================================================
# Tools exposed to Claude — inspection
# ===========================================================================
def inspect_workbook(path: str) -> dict[str, Any]:
    """Summarise a workbook: sheet names, sizes, and column headers.

    This is the first call the subagent makes on any file — it tells Claude
    what it's working with so it can pick the right transformations.
    """
    wb = load_workbook(path)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [
            ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
        ]
        sheets.append({
            "name": name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "headers": headers,
        })
    return {"path": path, "sheets": sheets}


def preview_rows(path: str, sheet: str, n: int = 5) -> dict[str, Any]:
    """Return the first N rows of a sheet as a list of dicts.

    Useful for the subagent to sanity-check data shape before transforming.
    """
    wb = load_workbook(path)
    ws = wb[sheet]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, min(2 + n, ws.max_row + 1)):
        row = {
            headers[c - 1]: ws.cell(row=r, column=c).value
            for c in range(1, ws.max_column + 1)
        }
        rows.append(row)
    return {"sheet": sheet, "headers": headers, "rows": rows}


def extract_unique_goc_names(
    path: str,
    sheet: str = "AAI_P&C_Ceded_H_NH",
    column: str = "AA",
    start_row: int = 3,
) -> dict[str, Any]:
    """Return the unique non-empty GoC names from a column of a Ceded workbook.

    Intended for the input file whose name ends with the fixed suffix
    ``AAI_P&C_Ceded`` (e.g. ``1.1_2025.12.31_AAI_P&C_Ceded.xlsx``). The
    default sheet, column and ``start_row`` (3, since rows 1-2 are header
    / sub-header) match that file's layout. Read-only and idempotent.
    Order follows first occurrence; whitespace is stripped and empty
    cells are skipped.
    """
    # data_only=True returns the cached computed value for formula cells
    # rather than the formula string — required for the Ceded file whose
    # column AA is typically populated by lookup formulas.
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    col_idx = column_index_from_string(column)
    seen: list[str] = []
    seen_set: set[str] = set()
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        if s not in seen_set:
            seen_set.add(s)
            seen.append(s)
    return {
        "sheet": sheet,
        "column": column,
        "count": len(seen),
        "values": seen,
    }


def extract_input_sunrise_goc_names(
    paths: list[str],
    sheet: str = "Input_Sunrise",
    column: str = "A",
    start_row: int = 2,
) -> dict[str, Any]:
    """Union of unique GoC names from the ``Input_Sunrise`` sheet of many files.

    Designed for the new Sunrise input model: the user uploads multiple
    files whose names carry the suffix ``_Ceded`` or ``_Assumed``, each
    with a sheet ``Input_Sunrise`` whose first column lists the GoCs.
    This function opens every supplied file, reads column ``A`` from
    ``start_row`` onward, and returns the union of unique values
    preserving first-seen order across all files.

    Whitespace is stripped; empty cells and rows whose GoC value is
    blank are skipped. Opens with ``data_only=True`` so formula cells
    return their cached computed value.
    """
    col_idx = column_index_from_string(column)
    seen_set: set[str] = set()
    seen: list[str] = []
    for path in paths:
        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet not in wb.sheetnames:
            raise KeyError(
                f"Sheet '{sheet}' not found in '{path}'. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet]
        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            if s not in seen_set:
                seen_set.add(s)
                seen.append(s)
    return {
        "sheet": sheet,
        "column": column,
        "count": len(seen),
        "values": seen,
        "file_count": len(paths),
    }


def extract_unique_goc_cohort_pairs(
    path: str,
    sheet: str = "AAI_P&C_Ceded_H_NH",
    goc_column: str = "AA",
    year_column: str = "AB",
    start_row: int = 3,
) -> dict[str, Any]:
    """Return unique (GoC, cohort_year) pairs from a Ceded workbook.

    For each row from ``start_row`` onward, reads ``goc_column`` (the
    GoC name, e.g. ``IT05PABPPLE``) and ``year_column`` (the cohort year
    as an integer, e.g. ``2024``), concatenates them into a ``GOC_ID``
    (e.g. ``IT05PABPPLE2024``), and dedupes on the joined identifier
    preserving first occurrence. Whitespace is stripped, empty cells
    and rows where the year cannot be coerced to int are skipped. Read-
    only and idempotent (uses ``data_only=True`` so cached formula
    values are returned instead of formula strings).

    Returns ``{"sheet": ..., "count": ..., "pairs": [{"goc_id", "goc",
    "year"}, ...]}``.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    goc_idx = column_index_from_string(goc_column)
    year_idx = column_index_from_string(year_column)
    seen: set[str] = set()
    pairs: list[dict[str, Any]] = []
    for r in range(start_row, ws.max_row + 1):
        goc_v = ws.cell(row=r, column=goc_idx).value
        year_v = ws.cell(row=r, column=year_idx).value
        if goc_v is None or year_v is None:
            continue
        goc = str(goc_v).strip()
        if not goc:
            continue
        try:
            year = int(year_v)
        except (TypeError, ValueError):
            continue
        goc_id = f"{goc}{year}"
        if goc_id in seen:
            continue
        seen.add(goc_id)
        pairs.append({"goc_id": goc_id, "goc": goc, "year": year})
    return {"sheet": sheet, "count": len(pairs), "pairs": pairs}


def _read_transcodifica_rows(path: str) -> list[list[Any]]:
    """Return the Transcodifica file as a list of raw rows (header included).

    Supports both CSV and XLSX. For CSV the field separator is auto-
    detected among ``;``, ``,`` and ``\\t`` by sniffing the first part
    of the file (falling back to ``;`` if Sniffer can not decide), so a
    Transcodifica exported as comma- or tab-delimited from Excel is
    handled correctly. The expected schema is:

    - column A: GoC code (key)
    - column B: Aggregation1
    - column C: Aggregation2
    - column D: H-NH flag (``"H"`` for Health, ``"NH"`` for non-Health)
    """
    suffix = Path(path).suffix.lower()
    rows: list[list[Any]] = []
    if suffix == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            sample = f.read(4096)
            f.seek(0)
            delim = _sniff_csv_delimiter(sample)
            for raw in csv.reader(f, delimiter=delim):
                rows.append(list(raw))
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        for r in range(1, ws.max_row + 1):
            rows.append([ws.cell(row=r, column=c).value for c in range(1, 5)])
    return rows


def _load_transcodifica_table(path: str) -> dict[str, tuple[Any, Any]]:
    """Load the Transcodifica master list as ``{goc: (agg1, agg2)}``.

    Only Aggregation1 and Aggregation2 are returned; the H-NH flag is
    consumed by ``extract_health_perimeter_gocs``.
    """
    table: dict[str, tuple[Any, Any]] = {}
    for row in _read_transcodifica_rows(path)[1:]:  # skip header
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        if not key:
            continue
        agg1 = row[1] if len(row) > 1 else None
        agg2 = row[2] if len(row) > 2 else None
        table[key] = (agg1, agg2)
    return table


def extract_health_perimeter_gocs(path: str) -> list[str]:
    """Return the GoCs flagged as Health in the Transcodifica master list.

    Reads the Transcodifica file (CSV or XLSX) and returns the GoCs
    whose column D value, normalised to upper case and stripped, equals
    ``"H"``. The list preserves the order in which the GoCs appear in
    the file. Used by ``update_mp_goc_seg`` as the perimeter on which
    ``P&C`` is rewritten to ``HLTH_PC``.
    """
    perimeter: list[str] = []
    seen: set[str] = set()
    for row in _read_transcodifica_rows(path)[1:]:  # skip header
        if not row or row[0] is None:
            continue
        goc = str(row[0]).strip()
        if not goc:
            continue
        flag = str(row[3]).strip().upper() if len(row) > 3 and row[3] is not None else ""
        if flag == "H" and goc not in seen:
            seen.add(goc)
            perimeter.append(goc)
    return perimeter


def _iter_input_sunrise_rows(
    path: str,
) -> Iterable[tuple[str, int, float, float]]:
    """Stream ``(GOC, ANNO, SINISTRI, RISERVA_SINISTRI)`` tuples.

    Uses openpyxl in ``read_only=True, data_only=True`` mode so the
    workbook is streamed instead of fully materialised — significantly
    faster and lighter on memory than ``pd.read_excel`` on large files.
    Rows with blank ``GOC`` or non-integer ``ANNO`` are skipped; missing
    ``SINISTRI`` / ``RISERVA_SINISTRI`` become ``0.0``.

    Column positions: A=GOC, B=ANNO, D=SINISTRI, E=RISERVA_SINISTRI.
    Column C (PERIMETRO) is not consumed by MP_ModelPoint.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Input_Sunrise" not in wb.sheetnames:
            raise KeyError(
                f"Sheet 'Input_Sunrise' not found in '{path}'. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb["Input_Sunrise"]
        rows_iter = ws.iter_rows(values_only=True)
        next(rows_iter, None)  # drop header
        for row in rows_iter:
            if not row:
                continue
            goc_v = row[0] if len(row) > 0 else None
            year_v = row[1] if len(row) > 1 else None
            if goc_v is None or year_v is None:
                continue
            goc = str(goc_v).strip()
            if not goc:
                continue
            try:
                acc_year = int(year_v)
            except (TypeError, ValueError):
                continue
            sin_v = row[3] if len(row) > 3 else None
            ris_v = row[4] if len(row) > 4 else None
            sinistri = float(sin_v) if sin_v is not None else 0.0
            riserva = float(ris_v) if ris_v is not None else 0.0
            yield goc, acc_year, sinistri, riserva
    finally:
        wb.close()


def build_input_sunrise_master_table(
    sources: list[tuple[str, int]],
) -> tuple[list[str], list[dict[str, Any]]]:
    """Build the GoC list and the aggregated master table for Sunrise.

    ``sources`` is a list of ``(path, anno_riferimento)`` tuples — one
    per uploaded ``_Ceded`` / ``_Assumed`` file. The function:

    1. starts with an empty GoC list;
    2. for each source file:
       - streams the ``Input_Sunrise`` sheet (openpyxl read-only mode);
       - appends the file's unique GoCs (column A) to the list,
         preserving first-seen order across files;
       - aggregates ``(GOC, ANNO)`` summing ``SINISTRI`` and
         ``RISERVA_SINISTRI``;
       - tags the rows with ``ANNO_RIFERIMENTO = anno_riferimento`` and
         appends them to a master table.

    Returns ``(goc_list, master_table)`` where ``master_table`` is a list
    of dicts with keys ``GOC, ANNO, SINISTRI, RISERVA_SINISTRI,
    ANNO_RIFERIMENTO``. A given ``(GOC, ANNO)`` appears at most once per
    ``ANNO_RIFERIMENTO``.
    """
    goc_list: list[str] = []
    seen_gocs: set[str] = set()
    master: list[dict[str, Any]] = []

    for path, anno_riferimento in sources:
        # Per-file aggregation, preserving first-seen (GoC, year) order.
        agg: dict[tuple[str, int], list[float]] = {}
        agg_order: list[tuple[str, int]] = []
        for goc, acc_year, sinistri, riserva in _iter_input_sunrise_rows(path):
            if goc not in seen_gocs:
                seen_gocs.add(goc)
                goc_list.append(goc)
            key = (goc, acc_year)
            if key in agg:
                agg[key][0] += sinistri
                agg[key][1] += riserva
            else:
                agg[key] = [sinistri, riserva]
                agg_order.append(key)
        anno_riferimento_int = int(anno_riferimento)
        for key in agg_order:
            sin_total, ris_total = agg[key]
            goc, acc_year = key
            master.append({
                "GOC": goc,
                "ANNO": acc_year,
                "SINISTRI": sin_total,
                "RISERVA_SINISTRI": ris_total,
                "ANNO_RIFERIMENTO": anno_riferimento_int,
            })

    return goc_list, master


def _emit_mp_model_point_rows(
    master: list[dict[str, Any]],
    transcodifica: dict[str, tuple[Any, Any]],
    year: int,
) -> list[list[Any]]:
    """Build the MP_ModelPoint data rows from the master aggregated table.

    Horizon per group:

    - ``@Closing`` (``ANNO_RIFERIMENTO == year``): ``[year - 15, year]``
      (16-year depth) when the GoC has any accident year at or below
      ``year - 15`` anywhere in the master; otherwise
      ``[per_goc_min, year]``.
    - ``@Opening`` (``ANNO_RIFERIMENTO == year - 1``): ``[year - 15,
      year - 1]`` (15-year depth) when the GoC has any accident year
      at or below ``year - 15``; otherwise ``[per_goc_min, year - 1]``.

    Pre-horizon fold: for each ``(ANNO_RIFERIMENTO, GoC)`` group, the
    SINISTRI/RISERVA of every accident year strictly older than
    ``fold_year = ANNO_RIFERIMENTO - 14`` are summed into the row at
    ``fold_year`` (creating it if not already present).

    Only ``@Closing`` also gets a zero padding row at ``year - 15``
    whenever the GoC has any accident year at or below ``year - 15``
    in any group — including cases where the ``@Closing`` group itself
    has no pre-horizon data of its own. This keeps ``@Closing`` and
    ``@Opening`` aligned on the same oldest accident year: whenever
    ``@Opening``'s 15-year horizon reaches ``year - 15``, ``@Closing``
    reaches it too (as a zero padding row) and the fold aggregate
    lives on ``year - 14`` instead. The ``@Opening`` group never
    receives a padding row.

    GoCs whose ``SINISTRI`` and ``RISERVA_SINISTRI`` are zero on every
    row of the master table (across all source files and all years) are
    omitted from the output entirely.
    """
    rows: list[list[Any]] = []
    if not master:
        return rows

    # First pass: which GoCs have at least one non-zero SINISTRI or
    # RISERVA_SINISTRI value somewhere in the master? GoCs that are all
    # zeros are excluded from the output.
    nonzero_gocs: set[str] = set()
    for entry in master:
        if entry["SINISTRI"] != 0.0 or entry["RISERVA_SINISTRI"] != 0.0:
            nonzero_gocs.add(entry["GOC"])

    # Per-GoC minimum accident year across all groups (needed to decide
    # whether @Closing has to emit a zero padding row at year - 15 to
    # align its horizon with @Opening).
    goc_min_present: dict[str, int] = {}
    for entry in master:
        goc = entry["GOC"]
        if goc not in nonzero_gocs:
            continue
        acc_year = int(entry["ANNO"])
        prev = goc_min_present.get(goc)
        if prev is None or acc_year < prev:
            goc_min_present[goc] = acc_year

    # Group by (ANNO_RIFERIMENTO, GOC) preserving insertion order.
    grouped: dict[tuple[int, str], list[dict[str, Any]]] = {}
    for entry in master:
        key = (entry["ANNO_RIFERIMENTO"], entry["GOC"])
        grouped.setdefault(key, []).append(entry)

    for (anno_rif, goc), entries in grouped.items():
        if goc not in nonzero_gocs:
            continue
        if anno_rif == year:
            max_year = year
            observation_suffix = "Closing"
        elif anno_rif == year - 1:
            max_year = year - 1
            observation_suffix = "Opening"
        else:
            continue

        # Fold at anno_rif - 14 for both groups. Only @Closing (anno_rif
        # == year) also gets a zero padding row at year - 15, whenever
        # the GoC has any accident year at or below year - 15 anywhere
        # in the master — so @Closing's horizon aligns with @Opening's
        # oldest year even when the @Closing group has no pre-horizon
        # data of its own. @Opening's horizon stops at fold_year
        # (15-year depth) and never gets a padding row.
        fold_year = anno_rif - 14

        in_horizon: dict[int, list[float]] = {}
        pre_horizon_sin = 0.0
        pre_horizon_ris = 0.0
        for entry in entries:
            acc_year = entry["ANNO"]
            sinistri = entry["SINISTRI"]
            riserva = entry["RISERVA_SINISTRI"]
            if acc_year < fold_year:
                pre_horizon_sin += sinistri
                pre_horizon_ris += riserva
            elif acc_year <= max_year:
                in_horizon[acc_year] = [sinistri, riserva]

        if pre_horizon_sin != 0.0 or pre_horizon_ris != 0.0:
            bucket = in_horizon.setdefault(fold_year, [0.0, 0.0])
            bucket[0] += pre_horizon_sin
            bucket[1] += pre_horizon_ris

        # @Closing padding at year - 15: whenever the GoC has data at
        # or below year - 15 in any group. This handles the case where
        # the current-year file has no pre-horizon values on its own
        # but the previous-year file does — the two groups must still
        # start at the same oldest accident year.
        if (
            anno_rif == year
            and goc_min_present.get(goc, year) <= year - 15
        ):
            in_horizon.setdefault(year - 15, [0.0, 0.0])

        agg1, agg2 = transcodifica.get(goc, (None, None))

        for acc_year in sorted(in_horizon.keys(), reverse=True):
            sinistri, riserva = in_horizon[acc_year]
            goc_id = f"{goc}{acc_year}"
            # Convention: MP_ModelPoint stores claim outflows with the
            # opposite sign of the input (positive input → negative
            # output, and vice versa). Zero stays zero.
            rows.append([
                f"{goc_id}@{anno_rif}", goc_id, f"{goc}@{observation_suffix}",
                acc_year, anno_rif,
                agg1, agg2,
                -riserva, 0,
                -sinistri, 0,
            ])
    return rows


MP_MODEL_POINT_HEADERS = [
    "Primary_Key", "GOC_ID", "ObservationKey",
    "Accident_Year", "Observation_Year",
    "Aggregation1", "Aggregation2",
    "EAXA_Reserve", "ULAE_Reserve",
    "Claims_Paid", "Cost_ULAE",
]


def create_mp_model_point(
    sources: list[tuple[str, int]],
    transcodifica_path: str,
    output_path: str,
    year: int,
    gocs_to_exclude: list[str] | None = None,
    goc_renames: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Create ``MP_ModelPoint.csv`` in two clear stages.

    ``sources`` is the list of ``(path, anno_riferimento)`` tuples — one
    per uploaded ``_Ceded`` / ``_Assumed`` file (the caller is expected
    to parse the date out of each filename and pass the year).

    Stage 1 — ``build_input_sunrise_master_table``:
    - Builds a GoC list (unique, first-seen order across files);
    - Builds a master aggregated table with columns
      ``GOC, ANNO, SINISTRI, RISERVA_SINISTRI, ANNO_RIFERIMENTO``,
      summing duplicates within each source file.

    Stage 2 — ``_emit_mp_model_point_rows``:
    - Groups the master table by ``(ANNO_RIFERIMENTO, GOC)``;
    - Determines the per-row horizon and ObservationKey suffix
      (current year → ``@Closing``, previous year → ``@Opening``);
    - Folds pre-horizon years into the ``min_year`` row;
    - Looks up Aggregation1 / Aggregation2 in the Transcodifica table;
    - Emits one CSV row per ``(GOC, accident_year)`` in the horizon.

    Returns ``{output_path, rows, columns, goc_list}``. ``goc_list`` is
    exposed so callers can reuse the list without re-reading the files.
    """
    transcodifica = _load_transcodifica_table(transcodifica_path)
    goc_list, master = build_input_sunrise_master_table(sources)
    warnings_out: list[str] = []

    # Apply per-(GoC, cohort) renames. Both the "Old" and the "New" value
    # are `GOC+cohort` strings (e.g. `IT05PABPPLE2024`): the last 4 chars
    # are the cohort year, the rest is the GoC name. This lets the user
    # rename a single (GoC, cohort) pair without touching the other
    # cohorts of the same GoC — effectively splitting a GoC row-by-row.
    # The exclusion list still operates on the 11-char GoC name.

    def _parse_goc_cohort_id(s: str) -> tuple[str, int] | None:
        s = (s or "").strip()
        if len(s) < 5:
            return None
        try:
            year = int(s[-4:])
        except ValueError:
            return None
        if not (1900 <= year <= 2100):
            return None
        name = s[:-4]
        if not name:
            return None
        return (name, year)

    renames_by_id: dict[tuple[str, int], tuple[str, int]] = {}
    for old, new in (goc_renames or {}).items():
        old_parsed = _parse_goc_cohort_id(old)
        new_parsed = _parse_goc_cohort_id(new)
        if old_parsed is None or new_parsed is None:
            warnings_out.append(
                f"Rename entry '{old}' -> '{new}' has invalid format "
                "(expected `<goc><YYYY>`, e.g. `IT05PABPPLE2024`). Skipped."
            )
            continue
        if old_parsed in renames_by_id:
            warnings_out.append(
                f"Rename for '{old}' specified more than once; "
                "keeping the first."
            )
            continue
        renames_by_id[old_parsed] = new_parsed

    if renames_by_id:
        existing_pairs = {(e["GOC"], int(e["ANNO"])) for e in master}

        # Warning: an "Old" pair that doesn't exist in the input is a no-op.
        for old_id in renames_by_id:
            if old_id not in existing_pairs:
                warnings_out.append(
                    f"Rename source '{old_id[0]}{old_id[1]}' is not in the "
                    "input list - nothing to rename for this entry."
                )

        # Blocking check: a "New" pair that already exists in the input
        # (and isn't itself being renamed away) would collide.
        old_ids = set(renames_by_id.keys())
        collisions: list[str] = []
        for old_id, new_id in renames_by_id.items():
            if new_id == old_id:
                continue  # identity rename, no-op
            if new_id in existing_pairs and new_id not in old_ids:
                collisions.append(f"{new_id[0]}{new_id[1]}")
        if collisions:
            raise ValueError(
                "Rename target(s) already present in the input GoC+cohort "
                f"list: {', '.join(sorted(set(collisions)))}. "
                "Change the *New GOC+cohort* value or remove the "
                "pre-existing input row before re-running."
            )

        # Apply the renames in place on master.
        for e in master:
            key = (e["GOC"], int(e["ANNO"]))
            if key in renames_by_id:
                new_goc, new_year = renames_by_id[key]
                e["GOC"] = new_goc
                e["ANNO"] = new_year

        # Second collision pass: two distinct renames may target the same
        # New id (e.g. Motor2025 -> X2025 and Property2025 -> X2025). After
        # applying, spot any (GoC, cohort, RIF) triple that now has >1 row.
        counts: dict[tuple[str, int, int], int] = {}
        for e in master:
            key3 = (e["GOC"], int(e["ANNO"]), int(e["ANNO_RIFERIMENTO"]))
            counts[key3] = counts.get(key3, 0) + 1
        dup_pairs = sorted(
            {f"{k[0]}{k[1]}" for k, v in counts.items() if v > 1}
        )
        if dup_pairs:
            raise ValueError(
                "Rename created duplicate GoC+cohort entries: "
                f"{', '.join(dup_pairs)}. Two renames cannot map to the "
                "same *New GOC+cohort*."
            )

        # Rebuild goc_list from the updated master so any brand-new GoC
        # name introduced by a rename shows up in MP_LoB / MP_ObservationYear.
        seen_gocs: set[str] = set()
        new_goc_list: list[str] = []
        for e in master:
            g = e["GOC"]
            if g not in seen_gocs:
                seen_gocs.add(g)
                new_goc_list.append(g)
        goc_list = new_goc_list

    exclude = {g.strip() for g in (gocs_to_exclude or []) if g and g.strip()}
    if exclude:
        master = [e for e in master if e["GOC"] not in exclude]
        goc_list = [g for g in goc_list if g not in exclude]
    nonzero = {
        e["GOC"] for e in master
        if e["SINISTRI"] != 0.0 or e["RISERVA_SINISTRI"] != 0.0
    }
    filtered_goc_list = [g for g in goc_list if g in nonzero]

    # For every ``ANNO_RIFERIMENTO`` actually present in the inputs
    # (typically the analysis year and the previous year — the two
    # Sunrise files), ensure every non-zero GoC has a row with
    # ``ANNO == ANNO_RIFERIMENTO``. A missing row would mean an empty
    # ``@Closing`` (when anno_rif == year) or ``@Opening`` (anno_rif ==
    # year - 1) for that GoC, plus a missing cohort pair handed to
    # Astra. When the row is missing we synthesize a zero one. The
    # all-zero filter above is respected: GoCs without a single
    # non-zero row anywhere stay out.
    existing_pairs = {
        (e["ANNO_RIFERIMENTO"], e["GOC"]) for e in master
        if int(e["ANNO"]) == e["ANNO_RIFERIMENTO"]
    }
    anno_rif_present = {e["ANNO_RIFERIMENTO"] for e in master}
    for anno_rif in anno_rif_present:
        for goc in filtered_goc_list:
            if (anno_rif, goc) in existing_pairs:
                continue
            master.append({
                "GOC": goc,
                "ANNO": anno_rif,
                "SINISTRI": 0.0,
                "RISERVA_SINISTRI": 0.0,
                "ANNO_RIFERIMENTO": anno_rif,
            })

    # Cohort-year gap fill: inside every ``(ANNO_RIFERIMENTO, GoC)`` group
    # the accident-year range must be contiguous, AND the oldest cohort
    # year considered for a given GoC must be the same across all of
    # its groups (``@Closing`` and ``@Opening``). The floor is therefore
    # the per-GoC minimum across every group; the ceiling stays the
    # group's own max (so ``@Opening`` keeps stopping at ``year - 1``
    # and ``@Closing`` keeps reaching ``year``). Years missing in the
    # input are added with zero sinistri/riserva, so MP_ModelPoint
    # emits a row for every cohort in that range and
    # ``goc_cohort_pairs`` carries the full set to Astra. Only
    # non-zero GoCs are filled (the all-zero exclusion rule above
    # still wins). Pre-horizon years produced by the fill are folded
    # into the ``min_year`` row at emit time as before.
    group_years: dict[tuple[int, str], set[int]] = {}
    for entry in master:
        goc = entry["GOC"]
        if goc not in nonzero:
            continue
        key = (entry["ANNO_RIFERIMENTO"], goc)
        group_years.setdefault(key, set()).add(int(entry["ANNO"]))
    goc_min_year: dict[str, int] = {}
    for (_anno_rif, goc), years in group_years.items():
        if not years:
            continue
        candidate = min(years)
        cur = goc_min_year.get(goc)
        if cur is None or candidate < cur:
            goc_min_year[goc] = candidate
    for (anno_rif, goc), years in group_years.items():
        if not years:
            continue
        floor = goc_min_year[goc]
        ceiling = max(years)
        for y in range(floor, ceiling + 1):
            if y in years:
                continue
            master.append({
                "GOC": goc,
                "ANNO": y,
                "SINISTRI": 0.0,
                "RISERVA_SINISTRI": 0.0,
                "ANNO_RIFERIMENTO": anno_rif,
            })

    # Unique (GoC, ANNO) pairs across all source files, restricted to the
    # non-zero GoC set and preserving first-seen order. Same shape as
    # the dicts produced by ``extract_unique_goc_cohort_pairs``, so that
    # downstream consumers (the Astra page) can reuse them as-is.
    pairs: list[dict[str, Any]] = []
    seen_pairs: set[tuple[str, int]] = set()
    for entry in master:
        goc = entry["GOC"]
        if goc not in nonzero:
            continue
        acc_year = int(entry["ANNO"])
        key = (goc, acc_year)
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        pairs.append({"goc_id": f"{goc}{acc_year}", "goc": goc, "year": acc_year})

    out_rows = _emit_mp_model_point_rows(master, transcodifica, year)
    _write_csv_rows(output_path, [MP_MODEL_POINT_HEADERS, *out_rows])
    return {
        "output_path": output_path,
        "rows": len(out_rows),
        "columns": MP_MODEL_POINT_HEADERS,
        "goc_list": filtered_goc_list,
        "goc_cohort_pairs": pairs,
        "warnings": warnings_out,
    }


def create_mp_lob(
    goc_names: list[str],
    entities: list[tuple[int, str]],
    output_path: str,
) -> dict[str, Any]:
    """Create an ``MP_LoB`` CSV with two columns: ``GoC_ID`` and ``Entity_ID``.

    ``entities`` is a list of ``(entity_id, entity_name)`` tuples — the user
    may select multiple entities. The output has one row per
    ``(goc, entity)`` pair, iterating GoCs outer and entities inner:

    ``GoC_1, Entity_A`` / ``GoC_1, Entity_B`` / ``GoC_2, Entity_A`` ...

    ``GoC_ID`` carries the GoC name, ``Entity_ID`` carries the entity
    integer code. Overwrites the output file if it already exists.
    """
    headers = ["GoC_ID", "Entity_ID"]
    rows: list[list[Any]] = [headers]
    for name in goc_names:
        for entity_id, _entity_name in entities:
            rows.append([name, entity_id])
    _write_csv_rows(output_path, rows)
    return {
        "output_path": output_path,
        "rows": len(goc_names) * len(entities),
        "columns": headers,
    }


def create_mp_observation_year(
    goc_names: list[str],
    year: int,
    output_path: str,
) -> dict[str, Any]:
    """Create an ``MP_ObservationYear`` CSV.

    Two rows are written per GoC — an ``Opening`` row (year - 1) and a
    ``Closing`` row (year). Columns: ``ObservationID`` (``{goc}@Opening``
    or ``{goc}@Closing``), ``ObservationYear``, ``LoB_ID`` (the GoC),
    ``AdjULAEPagate`` (always ``0``), ``CY`` (always ``Yes``). Overwrites
    the output file if it already exists.
    """
    headers = ["ObservationID", "ObservationYear", "LoB_ID", "AdjULAEPagate", "CY"]
    rows: list[list[Any]] = [headers]
    for name in goc_names:
        rows.append([f"{name}@Opening", year - 1, name, 0, "Yes"])
        rows.append([f"{name}@Closing", year, name, 0, "Yes"])
    _write_csv_rows(output_path, rows)
    return {
        "output_path": output_path,
        "rows": 2 * len(goc_names),
        "columns": headers,
    }


def lookup_risk_adjustment_values(
    path: str,
    goc_names: list[str],
    year: int,
    semester: int,
    sheet: str = "ra_AAI_REINS",
    goc_column: str = "G",
    header_row: int = 1,
) -> dict[str, dict[str, Any]]:
    """Look up Opening/Closing Risk Adjustment values for each GoC.

    Opens a Payment_Patterns_&_Risk_Adjustments workbook with
    ``data_only=True``, locates the table that starts in ``goc_column``
    on ``sheet``, and picks the two year columns matching the selected
    period: ``{prefix}_{year}`` for Closing and ``{prefix}_{year-1}``
    for Opening, where ``prefix = 'HY' if semester == 1 else 'FY'``.

    Returns ``{goc: {"opening": value, "closing": value}}``. GoCs missing
    from the sheet map to ``{"opening": None, "closing": None}``. Raises
    ``KeyError`` if either year column is absent from the header row.
    """
    if semester not in (1, 2):
        raise ValueError(f"semester must be 1 or 2, got {semester}")
    prefix = "HY" if semester == 1 else "FY"
    closing_col_name = f"{prefix}_{year}"
    opening_col_name = f"{prefix}_{year - 1}"

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    goc_col_idx = column_index_from_string(goc_column)

    header_to_idx: dict[str, int] = {}
    for c in range(goc_col_idx + 1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        if h is not None:
            header_to_idx[str(h).strip()] = c

    for needed in (opening_col_name, closing_col_name):
        if needed not in header_to_idx:
            raise KeyError(
                f"Column '{needed}' not found in sheet '{sheet}'. "
                f"Available year columns: {sorted(header_to_idx.keys())}"
            )

    goc_to_row: dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=goc_col_idx).value
        if v is None:
            continue
        key = str(v).strip()
        if key and key not in goc_to_row:
            goc_to_row[key] = r

    result: dict[str, dict[str, Any]] = {}
    for goc in goc_names:
        row = goc_to_row.get(goc)
        if row is None:
            result[goc] = {"opening": None, "closing": None}
            continue
        result[goc] = {
            "opening": ws.cell(row=row, column=header_to_idx[opening_col_name]).value,
            "closing": ws.cell(row=row, column=header_to_idx[closing_col_name]).value,
        }
    return result


def create_risk_adjustment(
    goc_names: list[str],
    values: dict[str, dict[str, Any]],
    output_path: str,
) -> dict[str, Any]:
    """Create a ``Risk_Adjustment`` CSV with two columns.

    ``ObservationID`` follows the ``{goc}@Opening`` / ``{goc}@Closing``
    pattern; ``Risk_Adjustment`` pulls from ``values`` (the dict returned
    by ``lookup_risk_adjustment_values``). Missing values are written as
    empty cells. Overwrites the output file if it already exists.
    """
    headers = ["ObservationID", "Risk_Adjustment"]
    rows: list[list[Any]] = [headers]
    for name in goc_names:
        vals = values.get(name, {"opening": None, "closing": None})
        rows.append([f"{name}@Opening", vals.get("opening")])
        rows.append([f"{name}@Closing", vals.get("closing")])
    _write_csv_rows(output_path, rows)
    return {
        "output_path": output_path,
        "rows": 2 * len(goc_names),
        "columns": headers,
    }


PAYMENT_PATTERN_COLUMN_COUNT = 23  # data columns '0' through '22'


def lookup_payment_pattern_values(
    path: str,
    goc_names: list[str],
    year: int,
    semester: int,
    sheet: str = "pp_AAI_REINS",
    goc_column: str = "C",
    year_column: str = "D",
    header_row: int = 1,
) -> list[dict[str, Any]]:
    """Look up Payment Pattern rows from a Payment_Patterns workbook.

    For each GoC the function emits two rows — one with the reference
    ``year`` and one with ``year - 1``. The source sheet layout is:

    - ``goc_column`` (default C): the GoC name.
    - ``year_column`` (default D): the period label in the format
      ``{prefix}{year}`` (e.g. ``FY2025``, ``HY2024``) — **no underscore**.
    - 23 data columns after ``year_column`` whose header-row values are
      ``'0'`` .. ``'22'`` (string or integer; leading/trailing
      whitespace is tolerated).

    The prefix follows the semester: H1 -> ``HY``, H2 -> ``FY``. Opens
    with ``data_only=True``. Missing (GoC, year) combinations produce a
    row of 23 ``None`` values. Raises ``KeyError`` if fewer than 23 data
    columns can be matched in the header row.

    Returns an ordered list of ``{"goc": str, "year": int, "values":
    list}`` dicts.
    """
    if semester not in (1, 2):
        raise ValueError(f"semester must be 1 or 2, got {semester}")
    prefix = "HY" if semester == 1 else "FY"

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    goc_col_idx = column_index_from_string(goc_column)
    year_col_idx = column_index_from_string(year_column)

    expected_headers = [str(i) for i in range(PAYMENT_PATTERN_COLUMN_COUNT)]
    data_cols: list[int] = []
    for c in range(year_col_idx + 1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        if h is None:
            continue
        if str(h).strip() in expected_headers:
            data_cols.append(c)
            if len(data_cols) == PAYMENT_PATTERN_COLUMN_COUNT:
                break
    if len(data_cols) != PAYMENT_PATTERN_COLUMN_COUNT:
        raise KeyError(
            f"Expected {PAYMENT_PATTERN_COLUMN_COUNT} data columns named "
            f"'0'..'22' after column '{year_column}' on sheet '{sheet}'; "
            f"found {len(data_cols)}."
        )

    key_to_row: dict[tuple[str, str], int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        goc_v = ws.cell(row=r, column=goc_col_idx).value
        year_v = ws.cell(row=r, column=year_col_idx).value
        if goc_v is None or year_v is None:
            continue
        goc_key = str(goc_v).strip()
        year_key = str(year_v).strip()
        if goc_key and year_key:
            key_to_row.setdefault((goc_key, year_key), r)

    result: list[dict[str, Any]] = []
    for goc in goc_names:
        for y in (year, year - 1):
            label = f"{prefix}{y}"
            row_num = key_to_row.get((goc, label))
            if row_num is None:
                values: list[Any] = [None] * PAYMENT_PATTERN_COLUMN_COUNT
            else:
                values = [
                    ws.cell(row=row_num, column=c).value for c in data_cols
                ]
            result.append({"goc": goc, "year": y, "values": values})
    return result


def create_payment_pattern(
    rows: list[dict[str, Any]],
    output_path: str,
) -> dict[str, Any]:
    """Create a ``Payment_pattern`` CSV with 25 columns.

    Columns: ``GoC``, ``Year`` and then ``'0'`` through ``'22'``. ``rows``
    is typically the list returned by ``lookup_payment_pattern_values``.
    Missing values produce empty cells. Overwrites the output file.
    """
    headers = ["GoC", "Year"] + [str(i) for i in range(PAYMENT_PATTERN_COLUMN_COUNT)]
    out_rows: list[list[Any]] = [headers]
    for row in rows:
        values = list(row.get("values", []))
        out_rows.append([row["goc"], row["year"], *values])
    _write_csv_rows(output_path, out_rows)
    return {
        "output_path": output_path,
        "rows": len(rows),
        "columns": headers,
    }


GOC_NAME_LENGTH = 11  # the GoC name is the first 11 chars of GOC_ID


def update_curve_id_param(
    input_path: str,
    output_path: str,
    closing_curve_name: str,
    opening_curve_name: str,
) -> dict[str, Any]:
    """Fill column C of CURVE_ID_PARAM.csv based on the VARIABLE_NAME.

    The input has three columns: ``GOC_ID`` (A), ``VARIABLE_NAME`` (B),
    a value column (C) — typically empty in the source. For each data
    row the function writes column C as follows:

    - ``VARIABLE_NAME == 'CLOSING_CURVE_ID'`` -> ``closing_curve_name``
    - ``VARIABLE_NAME == 'OPENING_CURVE_ID'`` -> ``opening_curve_name``
    - ``VARIABLE_NAME == 'CREDITED_RATE_CURVE_ID'`` -> ``'CR_' + GOC_ID`` (col A)

    Rows whose VARIABLE_NAME is anything else are left untouched, so any
    other historical content in the file passes through unchanged.
    """
    table = _read_csv_table(input_path)
    if not table:
        _write_csv_rows(output_path, [], **_ASTRA_CSV_KWARGS)
        return {"output_path": output_path, "rows_updated": 0}

    rows_updated = 0
    out_rows: list[list[Any]] = [table[0]]
    for row in table[1:]:
        row = list(row)
        while len(row) < 3:
            row.append(None)
        variable_name = row[1]
        if variable_name is not None:
            key = str(variable_name).strip()
            if key == "CLOSING_CURVE_ID":
                row[2] = closing_curve_name
                rows_updated += 1
            elif key == "OPENING_CURVE_ID":
                row[2] = opening_curve_name
                rows_updated += 1
            elif key == "CREDITED_RATE_CURVE_ID":
                col_a = row[0]
                row[2] = f"CR_{col_a}" if col_a is not None else None
                rows_updated += 1
        out_rows.append(row)

    _write_csv_rows(output_path, out_rows, **_ASTRA_CSV_KWARGS)
    return {
        "output_path": output_path,
        "rows_updated": rows_updated,
    }


def append_actuarial_aom_impact(
    input_path: str,
    output_path: str,
    pairs: list[dict[str, Any]],
    step_id_value_pairs: list[tuple[str, Any]],
) -> dict[str, Any]:
    """Append new rows to ACTUARIAL_AOM_IMPACT.csv and sort by A then B.

    The file is treated as an append-only history. Existing rows are
    preserved verbatim. For each ``(goc, cohort_year)`` pair (typically
    the filtered list from ``extract_unique_goc_cohort_pairs``) and each
    ``(step_id, value)`` pair from the user form, a new row is appended:
    ``[goc_id, step_id, value]``. After append, all data rows (header
    excluded) are sorted alphabetically by column A (``GOC_ID``) and
    then column B (``STEP_ID``).

    Columns beyond column C, if any, are preserved in the existing rows
    but left empty for newly appended rows. ``step_id_value_pairs`` with
    blank STEP_ID strings are skipped.
    """
    table = _read_csv_table(input_path)
    if not table:
        header: list[Any] = ["GOC_ID", "STEP_ID", 1]
        existing: list[list[Any]] = []
        n_cols = 3
    else:
        header = list(table[0])
        n_cols = max(len(header), 3)
        existing = []
        for raw in table[1:]:
            row = list(raw)
            while len(row) < n_cols:
                row.append(None)
            if any(v is not None for v in row):
                existing.append(row)

    cleaned_pairs = [
        (str(step).strip(), value)
        for step, value in step_id_value_pairs
        if step is not None and str(step).strip()
    ]
    new_rows: list[list[Any]] = []
    for pair in pairs:
        for step_id, value in cleaned_pairs:
            new_rows.append([pair["goc_id"], step_id, value] + [None] * (n_cols - 3))

    combined = existing + new_rows
    combined.sort(key=lambda r: (str(r[0] or ""), str(r[1] or "")))

    _write_csv_rows(output_path, [header, *combined], **_ASTRA_CSV_KWARGS)
    return {
        "output_path": output_path,
        "rows_appended": len(new_rows),
        "rows_total": len(combined),
    }


def update_mp_goc_seg(
    input_path: str,
    output_path: str,
    health_perimeter_gocs: list[str],
) -> dict[str, Any]:
    """Apply the Health-perimeter rewrite to MP_GOC_SEG.csv.

    The input has four columns: ``GOC_SEG_ID`` (A), ``GOC_ID`` (B),
    ``SEG_ID`` (C), ``ALLOCATION_RATIO`` (D), plus a header row. For
    each data row the GoC name is read as the first 11 characters of
    column B; if it appears in ``health_perimeter_gocs`` the substring
    ``'P&C'`` is replaced with ``'HLTH_PC'`` in columns A and C.
    Columns B and D, plus rows whose GoC is outside the perimeter, are
    left untouched. The file is saved to ``output_path``.
    """
    perimeter = {g.strip() for g in health_perimeter_gocs if g and g.strip()}

    table = _read_csv_table(input_path)
    if not table:
        _write_csv_rows(output_path, [], **_ASTRA_CSV_KWARGS)
        return {"output_path": output_path, "rows_in_perimeter": 0}

    rows_changed = 0
    out_rows: list[list[Any]] = [list(table[0])[:4]]
    # MP_GOC_SEG is contractually four columns. Pad short rows with None
    # and truncate any trailing cells (which would otherwise surface in
    # the output as empty trailing fields like ``1;;;``).
    while len(out_rows[0]) < 4:
        out_rows[0].append(None)
    for raw in table[1:]:
        row = list(raw)[:4]
        while len(row) < 4:
            row.append(None)

        goc_id_v = row[1]
        if goc_id_v is None:
            out_rows.append(row)
            continue
        normalized = str(goc_id_v).strip()
        if len(normalized) < GOC_NAME_LENGTH:
            out_rows.append(row)
            continue
        goc_name = normalized[:GOC_NAME_LENGTH]
        if goc_name not in perimeter:
            out_rows.append(row)
            continue

        col_a = row[0]
        if isinstance(col_a, str) and "P&C" in col_a:
            row[0] = col_a.replace("P&C", "HLTH_PC")
        col_c = row[2]
        if isinstance(col_c, str) and "P&C" in col_c:
            row[2] = col_c.replace("P&C", "HLTH_PC")
        rows_changed += 1
        out_rows.append(row)

    _write_csv_rows(output_path, out_rows, **_ASTRA_CSV_KWARGS)
    return {
        "output_path": output_path,
        "rows_in_perimeter": rows_changed,
    }


def update_projection_parameters_entity(
    input_path: str,
    output_path: str,
    year: int,
    semester: int,
) -> dict[str, Any]:
    """Apply rule-based edits to PROJECTION_PARAMETERS_ENTITY.csv.

    The input has two columns (``PARAMETER``, ``VALUE``) plus a header
    row. The function rewrites a fixed set of VALUE cells based on the
    analysis ``year`` and ``semester`` (1 = H1 = June, 2 = H2 = December),
    then saves to ``output_path``. Other rows are preserved. The output
    is contractually two columns; any extra columns in the input are
    dropped.

    Edits applied (parameter -> new value):

    - ``CF_TIMESTEP`` -> ``'SEMESTRIAL'``
    - ``REPORTING_MONTH`` -> ``'{month_num}_{MONTH_NAME}'``
      (``6_JUNE`` for H1, ``12_DECEMBER`` for H2)
    - ``FX_OPENING_DATE`` -> ``'1M{year_2d}'`` (e.g. ``1M25``)
    - ``FX_AVERAGE_DATE`` -> ``'Q1{year_2d}'`` (H1) or ``'HY{year_2d}'`` (H2)
    - ``FX_CLOSING_DATE`` -> ``'HY{year_2d}'`` (H1) or ``'FY{year_2d}'`` (H2)
    - ``FX_REPORTING_DATE`` -> ``'{year}0630'`` (H1) or ``'{year}1231'`` (H2)

    Parameters not present in the file are silently skipped.
    """
    if semester == 1:
        month_num = 6
        month_name = "JUNE"
        avg_prefix = "Q1"
        closing_prefix = "HY"
        report_date = f"{year}0630"
    elif semester == 2:
        month_num = 12
        month_name = "DECEMBER"
        avg_prefix = "HY"
        closing_prefix = "FY"
        report_date = f"{year}1231"
    else:
        raise ValueError(f"semester must be 1 or 2, got {semester}")

    year_2d = f"{year % 100:02d}"

    updates: dict[str, Any] = {
        "CF_TIMESTEP": "SEMESTRIAL",
        "REPORTING_MONTH": f"{month_num}_{month_name}",
        "FX_OPENING_DATE": f"1M{year_2d}",
        "FX_AVERAGE_DATE": f"{avg_prefix}{year_2d}",
        "FX_CLOSING_DATE": f"{closing_prefix}{year_2d}",
        "FX_REPORTING_DATE": report_date,
    }

    table = _read_csv_table(input_path)
    if not table:
        _write_csv_rows(output_path, [], **_ASTRA_CSV_KWARGS)
        return {"output_path": output_path, "parameters_updated": [], "rows_updated": 0}

    header = list(table[0])[:2]
    while len(header) < 2:
        header.append(None)

    applied: list[str] = []
    out_rows: list[list[Any]] = [header]
    for raw in table[1:]:
        row = list(raw)[:2]
        while len(row) < 2:
            row.append(None)
        param = row[0]
        if param is not None:
            key = str(param).strip()
            if key in updates:
                row[1] = updates[key]
                applied.append(key)
        out_rows.append(row)

    _write_csv_rows(output_path, out_rows, **_ASTRA_CSV_KWARGS)
    return {
        "output_path": output_path,
        "parameters_updated": applied,
        "rows_updated": len(applied),
    }


def create_empty_csv(output_path: str) -> dict[str, Any]:
    """Create an empty CSV file (zero rows, zero columns).

    Used as a placeholder until population rules for an output file are
    defined. Overwrites the output file if it already exists.
    """
    _write_csv_rows(output_path, [], **_ASTRA_CSV_KWARGS)
    return {"output_path": output_path, "rows": 0, "columns": []}


MP_GOC_BUSINESS_TYPE_VALUES = {
    "Diretto": "2_RE_ASSUMED",
    "Ceduto": "3_RE_CEDED_NON_RETRO",
}


def _mp_goc_inception_curve_id(cohort_year: int, mmdd: str) -> str:
    if cohort_year <= 2015:
        # Fixed value, independent of cohort year and semester.
        return "20211231_ITA_LP100"
    if cohort_year <= 2021:
        return f"{cohort_year}{mmdd}_ITA_LP100_AVG"
    if cohort_year == 2022:
        return f"2022{mmdd}_ITA_LP100_FY22_AVG"
    yy = f"{cohort_year % 100:02d}"
    return f"{cohort_year}{mmdd}_EUR_LP100_FY{yy}_AVG"


def update_mp_goc(
    input_path: str,
    output_path: str,
    year: int,
    semester: int,
    business_type: str,
) -> dict[str, Any]:
    """Apply the MP_GOC column rules and save to CSV.

    The input has 21 columns (header in row 1, data from row 2). The cohort
    year is column C (index 2). The function rewrites four columns per
    data row:

    - **E (INCEPTION_CURVE_ID, idx 4)** — year-bucketed string, with
      MMDD = ``0630`` for ``semester == 1`` and ``1231`` for ``semester == 2``:
        - cohort <= 2015: ``"20211231_ITA_LP100"`` (fixed)
        - 2016 <= cohort <= 2021: ``{cohort}{mmdd}_ITA_LP100_AVG``
        - cohort == 2022: ``2022{mmdd}_ITA_LP100_FY22_AVG``
        - cohort >= 2023: ``{cohort}{mmdd}_EUR_LP100_FY{yy}_AVG``
          where ``yy`` is the cohort year mod 100, zero-padded to 2.
    - **F (TIMING_INCEPTION_CURVE, idx 5)** — only when cohort_year == 2025:
      ``"7_JULY"`` for H1, ``"13_YEAR_END"`` for H2. Other rows keep the
      existing F value.
    - **L (GOC_DURATION, idx 11)** — ``max(0, year - 1 - curve_year) * 12``
      where ``curve_year`` is the integer parsed from the first 4 chars
      of the column E value just written. Equals the cohort year for
      cohorts >= 2016 and is pinned to 2021 for cohorts <= 2015.
    - **P (GOC_TYPE_REINSURANCE, idx 15)** — derived from column R
      (AGGREG_2_ID) on the same row: ``"3_RE_CEDED_NON_RETRO"`` when
      column R equals ``"PAA_Ceded"`` (whitespace-trimmed),
      ``"2_RE_ASSUMED"`` otherwise (including missing / blank cells).
    - **T (AGGREG_4_ID, idx 19) and U (AGGREG_5_ID, idx 20)** — both
      filled with the GoC name read from column A by stripping the
      trailing cohort-year suffix (e.g. ``IT05PABPPLE2024`` ->
      ``IT05PABPPLE``). Applied to every data row, including the
      ones where the cohort year is missing or unparseable (those
      rows fall back to the unmodified column A value).

    ``business_type`` is validated for backward compatibility with the
    Astra UI form but no longer drives any output — column P now
    depends only on column R.

    ``semester`` follows the pipeline convention (1 = H1 / June, 2 = H2 /
    December). Idempotent: rerunning with the same inputs produces the
    same output.
    """
    # ``business_type`` is reserved for future use (the UI no longer
    # collects it); column P is driven by column R alone. We still
    # accept the old values for backward compatibility but no longer
    # require them — anything else (including ``""``) is a no-op.
    if business_type and business_type not in MP_GOC_BUSINESS_TYPE_VALUES:
        raise ValueError(
            "business_type must be one of "
            f"{sorted(MP_GOC_BUSINESS_TYPE_VALUES)} or '', "
            f"got {business_type!r}"
        )
    if semester == 1:
        mmdd = "0630"
        f_value_for_2025 = "7_JULY"
    elif semester == 2:
        mmdd = "1231"
        f_value_for_2025 = "13_YEAR_END"
    else:
        raise ValueError(f"semester must be 1 or 2, got {semester}")
    # business_type is validated for backward-compat with the existing
    # UI form but no longer drives column P; the value is now derived
    # from column R (AGGREG_2_ID) of each row.

    table = _read_csv_table(input_path)
    if not table:
        _write_csv_rows(output_path, [], **_ASTRA_CSV_KWARGS)
        return {"output_path": output_path, "rows_changed": 0}

    out_rows: list[list[Any]] = [list(table[0])]
    rows_changed = 0
    for raw in table[1:]:
        row = list(raw)
        # Need to access columns up to U (index 20). Pad missing
        # trailing cells.
        while len(row) < 21:
            row.append(None)

        # Parse the cohort year (col C) up front — we use it both to
        # drive the E/F/L/P rules and to strip the suffix off column A
        # for the T/U population.
        cohort_raw = row[2]
        cohort_year: int | None
        if cohort_raw is None or cohort_raw == "":
            cohort_year = None
        else:
            try:
                cohort_year = int(cohort_raw)
            except (TypeError, ValueError):
                cohort_year = None

        # Columns T and U carry the GoC name (col A stripped of the
        # trailing cohort-year suffix). Applied to every data row,
        # falling back to the raw col A value when the cohort year
        # can't be parsed.
        goc_id_str = str(row[0]) if row[0] is not None else ""
        if cohort_year is not None:
            suffix = str(cohort_year)
            if goc_id_str.endswith(suffix):
                goc_name = goc_id_str[:-len(suffix)]
            else:
                goc_name = goc_id_str
        else:
            goc_name = goc_id_str
        row[19] = goc_name
        row[20] = goc_name

        if cohort_year is None:
            out_rows.append(row)
            continue

        row[4] = _mp_goc_inception_curve_id(cohort_year, mmdd)
        if cohort_year == 2025:
            row[5] = f_value_for_2025
        # GOC_DURATION (col L) uses the year embedded at the head of the
        # column E value, not the cohort_year directly. This makes the
        # duration follow the curve's effective year — relevant for the
        # cohort <= 2015 case where column E is pinned to "2021...".
        try:
            curve_year = int(str(row[4])[:4])
        except (TypeError, ValueError):
            curve_year = cohort_year
        row[11] = max(0, year - 1 - curve_year) * 12
        # GOC_TYPE_REINSURANCE (col P) is driven by the value of column R
        # (AGGREG_2_ID) on the same row: "PAA_Ceded" -> ceded, anything
        # else (including missing / different cells) -> assumed.
        col_r = row[17]
        col_r_str = col_r.strip() if isinstance(col_r, str) else ""
        if col_r_str == "PAA_Ceded":
            row[15] = "3_RE_CEDED_NON_RETRO"
        else:
            row[15] = "2_RE_ASSUMED"
        rows_changed += 1
        out_rows.append(row)

    _write_csv_rows(output_path, out_rows, **_ASTRA_CSV_KWARGS)
    return {
        "output_path": output_path,
        "rows_changed": rows_changed,
        "year": year,
        "semester": semester,
        "business_type": business_type,
    }


ASTRA_COHORT_YEAR_SPAN = 16  # analysis year + 15 prior cohorts; shared by all per-GoC cohort outputs
COVERAGE_UNIT_PROJECTION_COLUMN_COUNT = 100


def create_new_business_ppos(
    pairs: list[dict[str, Any]],
    output_path: str,
) -> dict[str, Any]:
    """Create a ``NEW_BUSINESS_PPOS`` CSV with three columns.

    One row per ``(goc, cohort_year)`` pair, in pair order. Pairs come
    from ``extract_unique_goc_cohort_pairs`` and are typically already
    filtered to the analysis window by the pipeline.

    Columns:
    - ``GOC_ID``: ``pair['goc_id']`` (e.g. ``IT05PABPPLE2024``).
    - ``VARIABLE_NAME``: always ``'CROSS_SUB_FASSCHNG'``.
    - ``1`` (literal integer header): always ``0``.

    Overwrites the output file if it already exists.
    """
    headers: list[Any] = ["GOC_ID", "VARIABLE_NAME", 1]
    rows: list[list[Any]] = [headers]
    for pair in pairs:
        rows.append([pair["goc_id"], "CROSS_SUB_FASSCHNG", 0])
    _write_csv_rows(output_path, rows, **_ASTRA_CSV_KWARGS)
    return {
        "output_path": output_path,
        "rows": len(pairs),
        "columns": ["GOC_ID", "VARIABLE_NAME", "1"],
    }


def create_coverage_unit(
    pairs: list[dict[str, Any]],
    output_path: str,
) -> dict[str, Any]:
    """Create a ``COVERAGE_UNIT`` CSV with 102 columns.

    One row per ``(goc, cohort_year)`` pair, in pair order.

    Columns:
    - ``GOC_ID``: ``pair['goc_id']``.
    - ``PROJECTION_PERIOD``: always ``1``.
    - ``1``, ``2``, ..., ``100`` (integer headers): always ``0``.

    Overwrites the output file if it already exists.
    """
    period_range = list(range(1, COVERAGE_UNIT_PROJECTION_COLUMN_COUNT + 1))
    header: list[Any] = ["GOC_ID", "PROJECTION_PERIOD", *period_range]
    rows: list[list[Any]] = [header]
    for pair in pairs:
        rows.append([pair["goc_id"], 1, *([0] * COVERAGE_UNIT_PROJECTION_COLUMN_COUNT)])
    _write_csv_rows(output_path, rows, **_ASTRA_CSV_KWARGS)
    return {
        "output_path": output_path,
        "rows": len(pairs),
        "columns": ["GOC_ID", "PROJECTION_PERIOD"]
        + [str(i) for i in period_range],
    }


REINSURANCE_VARIABLE_NAMES = ["LOSSRECO_IFE_ALLOCATION", "LOSSRECO_CLOSING"]

MANDATORY_ACTUALS_VARIABLE_NAMES = [
    "ACTUAL_PREMIUM_CF_PAST_SERVICE",
    "ACTUAL_PREMIUM_CF_FUTURE_SERVICE",
    "ACTUAL_ACQUISITION_CF_PAST_SERVICE",
    "ACTUAL_ACQUISITION_CF_FUTURE_SERVICE",
    "ACTUAL_INV_COMP_PAYABLE",
    "ACTUAL_CHANGE_FV_FINANCIAL_UNDERLYING_ITEMS",
    "FIN_RISK_MITIGATION",
    "RE_RISK_MITIGATION",
    "RE_RISK_MITIGATION_PNL",
    "ACTUAL_COVERAGE_UNIT",
    "THEORETICAL_PREMIUM_DERECOGNITION_LRC",
    "THEORETICAL_PREMIUM_RECOGNITION",
    "DERECOGNITION_ASSET_ACQ_CF_BF_INIT_RECOG",
    "ACTUAL_COMMISSION_CF_PAST_SERVICE",
    "ACTUAL_COMMISSION_CF_FUTURE_SERVICE",
    "THEORETICAL_PREMIUM_DERECOGNITION_LIC",
]


def create_reinsurance(
    pairs: list[dict[str, Any]],
    output_path: str,
) -> dict[str, Any]:
    """Create a ``REINSURANCE`` CSV with four columns.

    Pairs are grouped by GoC (preserving first-seen order). For each
    GoC, all ``LOSSRECO_IFE_ALLOCATION`` rows are emitted first (one
    per pair of that GoC), followed by all ``LOSSRECO_CLOSING`` rows.

    Columns:
    - ``GOC_ID``: ``pair['goc_id']``.
    - ``VARIABLE_NAME``: ``'LOSSRECO_IFE_ALLOCATION'`` or ``'LOSSRECO_CLOSING'``.
    - ``1`` (literal integer header): always ``0``.
    - ``T``: the cohort year from the pair.

    Overwrites the output file if it already exists.
    """
    grouped: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for p in pairs:
        grouped.setdefault(p["goc"], []).append(p)

    rows: list[list[Any]] = [["GOC_ID", "VARIABLE_NAME", 1, "T"]]
    for goc_pairs in grouped.values():
        for variable_name in REINSURANCE_VARIABLE_NAMES:
            for pair in goc_pairs:
                rows.append([pair["goc_id"], variable_name, 0, pair["year"]])
    _write_csv_rows(output_path, rows, **_ASTRA_CSV_KWARGS)
    return {
        "output_path": output_path,
        "rows": len(pairs) * len(REINSURANCE_VARIABLE_NAMES),
        "columns": ["GOC_ID", "VARIABLE_NAME", "1", "T"],
    }


def create_mandatory_actuals(
    pairs: list[dict[str, Any]],
    output_path: str,
) -> dict[str, Any]:
    """Create a ``MANDATORY_ACTUALS`` CSV with three columns.

    Pairs are grouped by GoC (preserving first-seen order). For each
    GoC, for each pair of that GoC, the 16 fixed VARIABLE_NAMEs are
    emitted in order. Total rows: ``len(pairs) * 16``.

    Columns:
    - ``GOC_ID``: ``pair['goc_id']``.
    - ``VARIABLE_NAME``: one of the 16 fixed values, in order.
    - ``1`` (literal integer header): always ``0``.

    Overwrites the output file if it already exists.
    """
    grouped: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for p in pairs:
        grouped.setdefault(p["goc"], []).append(p)

    rows: list[list[Any]] = [["GOC_ID", "VARIABLE_NAME", 1]]
    for goc_pairs in grouped.values():
        for pair in goc_pairs:
            for variable_name in MANDATORY_ACTUALS_VARIABLE_NAMES:
                rows.append([pair["goc_id"], variable_name, 0])
    _write_csv_rows(output_path, rows, **_ASTRA_CSV_KWARGS)
    return {
        "output_path": output_path,
        "rows": len(pairs) * len(MANDATORY_ACTUALS_VARIABLE_NAMES),
        "columns": ["GOC_ID", "VARIABLE_NAME", "1"],
    }

"""Tests for excel_pipeline.skill."""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest

from excel_pipeline.skill import (
    MANDATORY_ACTUALS_VARIABLE_NAMES,
    create_coverage_unit,
    create_empty_csv,
    create_mandatory_actuals,
    create_mp_lob,
    create_mp_model_point,
    create_mp_observation_year,
    create_new_business_ppos,
    create_payment_pattern,
    create_reinsurance,
    create_risk_adjustment,
    append_actuarial_aom_impact,
    extract_health_perimeter_gocs,
    extract_input_sunrise_goc_names,
    extract_unique_goc_cohort_pairs,
    extract_unique_goc_names,
    lookup_payment_pattern_values,
    lookup_risk_adjustment_values,
    update_curve_id_param,
    update_mp_goc,
    update_mp_goc_seg,
    update_projection_parameters_entity,
)


def _pair(goc: str, year: int) -> dict:
    """Helper for tests — mirrors extract_unique_goc_cohort_pairs output."""
    return {"goc_id": f"{goc}{year}", "goc": goc, "year": year}


def _read_csv(path: Path) -> list[tuple]:
    """Read CSV rows; coerce numeric cells to int/float, '' to None.

    Matches the European convention used by ``_write_csv_rows`` —
    ``;`` field separator, ``,`` decimal — and converts ``"12,5"`` back
    to ``12.5`` (float).
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=";")
        rows: list[tuple] = []
        for raw in reader:
            cells = []
            for cell in raw:
                if cell == "":
                    cells.append(None)
                    continue
                try:
                    cells.append(int(cell))
                    continue
                except ValueError:
                    pass
                try:
                    cells.append(float(cell.replace(",", ".")))
                    continue
                except ValueError:
                    pass
                cells.append(cell)
            rows.append(tuple(cells))
        return rows


def _write_csv(path: Path, rows: list[tuple]) -> None:
    """Write rows to a CSV fixture (UTF-8 with BOM, ``;`` separator)."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        for row in rows:
            writer.writerow(["" if v is None else v for v in row])


def _build_ceded_fixture(path: Path) -> None:
    """Create a minimal workbook shaped like the AAI_P&C_Ceded input file.

    Matches the real file: rows 1-2 are header / sub-header, data from
    row 3 onwards.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AAI_P&C_Ceded_H_NH"
    ws.cell(row=1, column=27, value="GoC")  # column AA = 27
    ws.cell(row=2, column=27, value="Line of Business")  # sub-header
    values = [
        "Motor",
        "Property",
        "Motor",
        None,
        "  Property  ",
        "Liability",
        "",
        "Motor",
    ]
    for i, v in enumerate(values, start=3):
        ws.cell(row=i, column=27, value=v)
    wb.save(path)


def test_extract_unique_goc_names(tmp_path: Path) -> None:
    fixture = tmp_path / "1.1_2025.12.31_AAI_P&C_Ceded.xlsx"
    _build_ceded_fixture(fixture)

    result = extract_unique_goc_names(str(fixture))

    assert result["sheet"] == "AAI_P&C_Ceded_H_NH"
    assert result["column"] == "AA"
    assert result["values"] == ["Motor", "Property", "Liability"]
    assert result["count"] == 3


def _build_input_sunrise_fixture(path: Path, gocs: list[str | None]) -> None:
    """Workbook with an Input_Sunrise sheet — col A from row 2 contains gocs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input_Sunrise"
    # Header row matches the spec: 5 columns
    ws.cell(row=1, column=1, value="GoC")
    ws.cell(row=1, column=2, value="Year")
    ws.cell(row=1, column=3, value="Perimetro")
    ws.cell(row=1, column=4, value="Sinistri")
    ws.cell(row=1, column=5, value="Riserva")
    for i, g in enumerate(gocs, start=2):
        ws.cell(row=i, column=1, value=g)
    wb.save(path)


def test_extract_input_sunrise_goc_names_union_across_files(tmp_path: Path) -> None:
    f1 = tmp_path / "1.1_2025.12.31_AAI_Ceded.xlsx"
    f2 = tmp_path / "1.2_2025.12.31_AAI_Assumed.xlsx"
    f3 = tmp_path / "1.3_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_fixture(f1, ["Motor", "Property", "  Motor  ", None, ""])
    _build_input_sunrise_fixture(f2, ["Property", "Liability"])  # Property dup
    _build_input_sunrise_fixture(f3, ["Marine", "Motor"])  # Motor dup, Marine new

    result = extract_input_sunrise_goc_names(
        [str(f1), str(f2), str(f3)]
    )

    # Union preserving first-seen order; whitespace stripped; blanks skipped
    assert result["values"] == ["Motor", "Property", "Liability", "Marine"]
    assert result["count"] == 4
    assert result["file_count"] == 3


def test_extract_input_sunrise_goc_names_missing_sheet(tmp_path: Path) -> None:
    bad = tmp_path / "wrong.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "OtherSheet"
    wb.save(bad)

    with pytest.raises(KeyError, match="Input_Sunrise"):
        extract_input_sunrise_goc_names([str(bad)])


def test_extract_unique_goc_names_custom_column_and_start_row(tmp_path: Path) -> None:
    fixture = tmp_path / "custom.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AAI_P&C_Ceded_H_NH"
    ws.cell(row=1, column=2, value="GoC")
    for i, v in enumerate(["A", "B", "A"], start=2):
        ws.cell(row=i, column=2, value=v)
    wb.save(fixture)

    result = extract_unique_goc_names(str(fixture), column="B", start_row=2)

    assert result["values"] == ["A", "B"]


def test_create_mp_lob_single_entity(tmp_path: Path) -> None:
    output = tmp_path / "MP_LoB.csv"
    goc_names = ["Motor", "Property", "Liability"]

    result = create_mp_lob(
        goc_names=goc_names,
        entities=[(6, "AAI")],
        output_path=str(output),
    )

    assert result == {
        "output_path": str(output),
        "rows": 3,
        "columns": ["GoC_ID", "Entity_ID"],
    }
    assert output.exists()

    rows = _read_csv(output)
    assert rows == [
        ("GoC_ID", "Entity_ID"),
        ("Motor", 6),
        ("Property", 6),
        ("Liability", 6),
    ]


def test_create_mp_lob_multi_entity_cartesian(tmp_path: Path) -> None:
    output = tmp_path / "MP_LoB.csv"
    goc_names = ["Motor", "Property"]

    result = create_mp_lob(
        goc_names=goc_names,
        entities=[(6, "AAI"), (14, "MPS")],
        output_path=str(output),
    )

    assert result["rows"] == 4  # 2 GoCs * 2 entities

    rows = _read_csv(output)
    # Order: GoC outer, entity inner
    assert rows == [
        ("GoC_ID", "Entity_ID"),
        ("Motor", 6),
        ("Motor", 14),
        ("Property", 6),
        ("Property", 14),
    ]


def test_create_mp_lob_end_to_end(tmp_path: Path) -> None:
    ceded = tmp_path / "1.1_2025.12.31_AAI_P&C_Ceded.xlsx"
    _build_ceded_fixture(ceded)
    output = tmp_path / "MP_LoB.csv"

    extracted = extract_unique_goc_names(str(ceded))
    create_mp_lob(
        goc_names=extracted["values"],
        entities=[(14, "MPS")],
        output_path=str(output),
    )

    rows = _read_csv(output)
    assert rows == [
        ("GoC_ID", "Entity_ID"),
        ("Motor", 14),
        ("Property", 14),
        ("Liability", 14),
    ]


def test_create_mp_observation_year(tmp_path: Path) -> None:
    output = tmp_path / "MP_ObservationYear.csv"
    goc_names = ["Motor", "Property"]

    result = create_mp_observation_year(
        goc_names=goc_names, year=2025, output_path=str(output),
    )

    assert result == {
        "output_path": str(output),
        "rows": 4,
        "columns": [
            "ObservationID",
            "ObservationYear",
            "LoB_ID",
            "AdjULAEPagate",
            "CY",
        ],
    }

    rows = _read_csv(output)
    assert rows == [
        ("ObservationID", "ObservationYear", "LoB_ID", "AdjULAEPagate", "CY"),
        ("Motor@Opening", 2024, "Motor", 0, "Yes"),
        ("Motor@Closing", 2025, "Motor", 0, "Yes"),
        ("Property@Opening", 2024, "Property", 0, "Yes"),
        ("Property@Closing", 2025, "Property", 0, "Yes"),
    ]


def test_create_mp_observation_year_empty(tmp_path: Path) -> None:
    output = tmp_path / "MP_ObservationYear.csv"

    result = create_mp_observation_year(
        goc_names=[], year=2025, output_path=str(output),
    )

    assert result["rows"] == 0
    rows = _read_csv(output)
    assert rows == [
        ("ObservationID", "ObservationYear", "LoB_ID", "AdjULAEPagate", "CY"),
    ]


def _build_payment_patterns_fixture(path: Path) -> None:
    """Fixture shaped like Payment_Patterns_&_Risk_Adjustments."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ra_AAI_REINS"
    # Column G = 7; header row 1.
    ws.cell(row=1, column=7, value="GoC")
    headers = ["HY_2024", "FY_2024", "HY_2025", "FY_2025"]
    for i, h in enumerate(headers, start=8):  # H, I, J, K
        ws.cell(row=1, column=i, value=h)
    data = [
        ("Motor", 100, 110, 120, 130),
        ("Property", 200, 210, 220, 230),
        ("Liability", 300, 310, 320, 330),
    ]
    for r, (goc, hy24, fy24, hy25, fy25) in enumerate(data, start=2):
        ws.cell(row=r, column=7, value=goc)
        ws.cell(row=r, column=8, value=hy24)
        ws.cell(row=r, column=9, value=fy24)
        ws.cell(row=r, column=10, value=hy25)
        ws.cell(row=r, column=11, value=fy25)
    wb.save(path)


def test_lookup_risk_adjustment_values_h2(tmp_path: Path) -> None:
    fixture = tmp_path / "pp.xlsx"
    _build_payment_patterns_fixture(fixture)

    values = lookup_risk_adjustment_values(
        path=str(fixture),
        goc_names=["Motor", "Property", "Liability"],
        year=2025,
        semester=2,
    )

    # H2 2025: Closing = FY_2025, Opening = FY_2024
    assert values == {
        "Motor": {"opening": 110, "closing": 130},
        "Property": {"opening": 210, "closing": 230},
        "Liability": {"opening": 310, "closing": 330},
    }


def test_lookup_risk_adjustment_values_h1(tmp_path: Path) -> None:
    fixture = tmp_path / "pp.xlsx"
    _build_payment_patterns_fixture(fixture)

    values = lookup_risk_adjustment_values(
        path=str(fixture),
        goc_names=["Motor"],
        year=2025,
        semester=1,
    )

    # H1 2025: Closing = HY_2025, Opening = HY_2024
    assert values == {"Motor": {"opening": 100, "closing": 120}}


def test_lookup_risk_adjustment_values_missing_goc(tmp_path: Path) -> None:
    fixture = tmp_path / "pp.xlsx"
    _build_payment_patterns_fixture(fixture)

    values = lookup_risk_adjustment_values(
        path=str(fixture),
        goc_names=["Motor", "UnknownGoC"],
        year=2025,
        semester=2,
    )

    assert values["UnknownGoC"] == {"opening": None, "closing": None}


def test_lookup_risk_adjustment_values_missing_year_column(tmp_path: Path) -> None:
    fixture = tmp_path / "pp.xlsx"
    _build_payment_patterns_fixture(fixture)

    with pytest.raises(KeyError, match=r"FY_20(29|30).*ra_AAI_REINS"):
        lookup_risk_adjustment_values(
            path=str(fixture),
            goc_names=["Motor"],
            year=2030,
            semester=2,
        )


def test_create_risk_adjustment(tmp_path: Path) -> None:
    output = tmp_path / "Risk_Adjustment.csv"
    goc_names = ["Motor", "Property"]
    values = {
        "Motor": {"opening": 110, "closing": 130},
        "Property": {"opening": 210, "closing": 230},
    }

    result = create_risk_adjustment(
        goc_names=goc_names, values=values, output_path=str(output),
    )

    assert result == {
        "output_path": str(output),
        "rows": 4,
        "columns": ["ObservationID", "Risk_Adjustment"],
    }

    rows = _read_csv(output)
    assert rows == [
        ("ObservationID", "Risk_Adjustment"),
        ("Motor@Opening", 110),
        ("Motor@Closing", 130),
        ("Property@Opening", 210),
        ("Property@Closing", 230),
    ]


def test_create_risk_adjustment_missing_values_become_empty(tmp_path: Path) -> None:
    output = tmp_path / "Risk_Adjustment.csv"
    values = {"Motor": {"opening": None, "closing": None}}

    create_risk_adjustment(
        goc_names=["Motor"], values=values, output_path=str(output),
    )

    rows = _read_csv(output)
    assert rows == [
        ("ObservationID", "Risk_Adjustment"),
        ("Motor@Opening", None),
        ("Motor@Closing", None),
    ]


def _build_pp_sheet_fixture(path: Path) -> None:
    """Fixture of the pp_AAI_REINS sheet for Payment Pattern lookups."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "pp_AAI_REINS"
    # Header row: A and B empty, C=GoC, D=Year, E..AA = '0'..'22'
    ws.cell(row=1, column=3, value="GoC")
    ws.cell(row=1, column=4, value="Year")
    for i in range(23):
        ws.cell(row=1, column=5 + i, value=str(i))

    # Data rows — (goc, year_label, seed) -> values[i] = seed + i
    data = [
        ("Motor", "FY2024", 1000),
        ("Motor", "FY2025", 2000),
        ("Motor", "HY2024", 3000),
        ("Motor", "HY2025", 4000),
        ("Property", "FY2024", 5000),
        ("Property", "FY2025", 6000),
    ]
    for r, (goc, year_label, seed) in enumerate(data, start=2):
        ws.cell(row=r, column=3, value=goc)
        ws.cell(row=r, column=4, value=year_label)
        for i in range(23):
            ws.cell(row=r, column=5 + i, value=seed + i)
    wb.save(path)


def test_lookup_payment_pattern_values_h2(tmp_path: Path) -> None:
    fixture = tmp_path / "pp.xlsx"
    _build_pp_sheet_fixture(fixture)

    rows = lookup_payment_pattern_values(
        path=str(fixture),
        goc_names=["Motor"],
        year=2025,
        semester=2,
    )

    # H2 -> FY prefix. Order: reference year, then year-1.
    assert rows == [
        {"goc": "Motor", "year": 2025, "values": [2000 + i for i in range(23)]},
        {"goc": "Motor", "year": 2024, "values": [1000 + i for i in range(23)]},
    ]


def test_lookup_payment_pattern_values_h1(tmp_path: Path) -> None:
    fixture = tmp_path / "pp.xlsx"
    _build_pp_sheet_fixture(fixture)

    rows = lookup_payment_pattern_values(
        path=str(fixture),
        goc_names=["Motor"],
        year=2025,
        semester=1,
    )

    # H1 -> HY prefix
    assert rows[0]["values"][0] == 4000  # HY2025
    assert rows[1]["values"][0] == 3000  # HY2024


def test_lookup_payment_pattern_values_missing_row(tmp_path: Path) -> None:
    fixture = tmp_path / "pp.xlsx"
    _build_pp_sheet_fixture(fixture)

    rows = lookup_payment_pattern_values(
        path=str(fixture),
        goc_names=["Property"],
        year=2025,
        semester=1,  # HY2025 / HY2024 — not in fixture for Property
    )

    assert rows[0]["values"] == [None] * 23
    assert rows[1]["values"] == [None] * 23


def test_lookup_payment_pattern_values_fewer_than_23_columns(tmp_path: Path) -> None:
    fixture = tmp_path / "pp.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "pp_AAI_REINS"
    ws.cell(row=1, column=3, value="GoC")
    ws.cell(row=1, column=4, value="Year")
    # Only 5 data columns
    for i in range(5):
        ws.cell(row=1, column=5 + i, value=str(i))
    wb.save(fixture)

    with pytest.raises(KeyError, match="23 data columns"):
        lookup_payment_pattern_values(
            path=str(fixture),
            goc_names=["Motor"],
            year=2025,
            semester=2,
        )


def test_create_payment_pattern(tmp_path: Path) -> None:
    output = tmp_path / "Payment_pattern.csv"
    rows = [
        {"goc": "Motor", "year": 2025, "values": [i for i in range(23)]},
        {"goc": "Motor", "year": 2024, "values": [i * 10 for i in range(23)]},
    ]

    result = create_payment_pattern(rows=rows, output_path=str(output))

    expected_headers = ["GoC", "Year"] + [str(i) for i in range(23)]
    assert result == {
        "output_path": str(output),
        "rows": 2,
        "columns": expected_headers,
    }

    written = _read_csv(output)
    # CSV reader returns header column "0".."22" as ints via the int-coercion;
    # compare on string form to be unambiguous.
    assert written[0] == ("GoC", "Year") + tuple(range(23))
    assert written[1] == ("Motor", 2025) + tuple(range(23))
    assert written[2] == ("Motor", 2024) + tuple(i * 10 for i in range(23))


def test_create_empty_csv(tmp_path: Path) -> None:
    output = tmp_path / "Empty.csv"

    result = create_empty_csv(str(output))

    assert result == {
        "output_path": str(output),
        "rows": 0,
        "columns": [],
    }
    assert output.exists()
    assert _read_csv(output) == []


def _build_ceded_with_year_fixture(path: Path, rows_data: list[tuple]) -> None:
    """Ceded fixture with GoC name in column AA and cohort year in column AB.

    `rows_data` is a list of (goc, year) tuples written from row 3 onward.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AAI_P&C_Ceded_H_NH"
    ws.cell(row=1, column=27, value="GoC")
    ws.cell(row=1, column=28, value="Cohort Year")
    ws.cell(row=2, column=27, value="Code")
    ws.cell(row=2, column=28, value="Year")
    for i, (goc, year) in enumerate(rows_data, start=3):
        ws.cell(row=i, column=27, value=goc)
        ws.cell(row=i, column=28, value=year)
    wb.save(path)


def test_extract_unique_goc_cohort_pairs(tmp_path: Path) -> None:
    fixture = tmp_path / "ceded.xlsx"
    _build_ceded_with_year_fixture(
        fixture,
        [
            ("IT05PABPPLE", 2024),
            ("IT05PABPPLE", 2023),
            ("IT05PABPPLE", 2024),  # duplicate -> dropped
            ("  IT05PABPPLE  ", 2022),  # stripped
            ("IT06ABCDE", 2024),
            (None, 2020),  # skipped (None GoC)
            ("IT06ABCDE", None),  # skipped (None year)
            ("", 2020),  # skipped (empty GoC)
        ],
    )

    result = extract_unique_goc_cohort_pairs(str(fixture))

    assert result["sheet"] == "AAI_P&C_Ceded_H_NH"
    assert result["count"] == 4
    assert result["pairs"] == [
        {"goc_id": "IT05PABPPLE2024", "goc": "IT05PABPPLE", "year": 2024},
        {"goc_id": "IT05PABPPLE2023", "goc": "IT05PABPPLE", "year": 2023},
        {"goc_id": "IT05PABPPLE2022", "goc": "IT05PABPPLE", "year": 2022},
        {"goc_id": "IT06ABCDE2024", "goc": "IT06ABCDE", "year": 2024},
    ]


def test_create_new_business_ppos(tmp_path: Path) -> None:
    output = tmp_path / "NEW_BUSINESS_PPOS.csv"
    pairs = [_pair("IT05PABPPLE", 2024), _pair("IT05PABPPLE", 2023), _pair("IT06ABCDE", 2024)]

    result = create_new_business_ppos(pairs=pairs, output_path=str(output))

    assert result == {
        "output_path": str(output),
        "rows": 3,
        "columns": ["GOC_ID", "VARIABLE_NAME", "1"],
    }

    rows = _read_csv(output)
    assert rows == [
        ("GOC_ID", "VARIABLE_NAME", 1),
        ("IT05PABPPLE2024", "CROSS_SUB_FASSCHNG", 0),
        ("IT05PABPPLE2023", "CROSS_SUB_FASSCHNG", 0),
        ("IT06ABCDE2024", "CROSS_SUB_FASSCHNG", 0),
    ]


def test_create_new_business_ppos_empty_pairs(tmp_path: Path) -> None:
    output = tmp_path / "NEW_BUSINESS_PPOS.csv"

    result = create_new_business_ppos(pairs=[], output_path=str(output))

    assert result["rows"] == 0
    rows = _read_csv(output)
    assert rows == [("GOC_ID", "VARIABLE_NAME", 1)]


def test_create_coverage_unit(tmp_path: Path) -> None:
    output = tmp_path / "COVERAGE_UNIT.csv"
    pairs = [_pair("IT05PABPPLE", 2024), _pair("IT06ABCDE", 2024)]

    result = create_coverage_unit(pairs=pairs, output_path=str(output))

    expected_columns = ["GOC_ID", "PROJECTION_PERIOD"] + [str(i) for i in range(1, 101)]
    assert result == {
        "output_path": str(output),
        "rows": 2,
        "columns": expected_columns,
    }

    rows = _read_csv(output)
    assert rows[0] == ("GOC_ID", "PROJECTION_PERIOD") + tuple(range(1, 101))
    assert rows[1] == ("IT05PABPPLE2024", 1) + (0,) * 100
    assert rows[2] == ("IT06ABCDE2024", 1) + (0,) * 100
    assert len(rows) == 3


def test_create_reinsurance(tmp_path: Path) -> None:
    output = tmp_path / "REINSURANCE.csv"
    pairs = [
        _pair("IT05PABPPLE", 2024),
        _pair("IT05PABPPLE", 2023),
        _pair("IT06ABCDE", 2024),
    ]

    result = create_reinsurance(pairs=pairs, output_path=str(output))

    assert result == {
        "output_path": str(output),
        "rows": 3 * 2,
        "columns": ["GOC_ID", "VARIABLE_NAME", "1", "T"],
    }

    rows = _read_csv(output)
    assert rows == [
        ("GOC_ID", "VARIABLE_NAME", 1, "T"),
        ("IT05PABPPLE2024", "LOSSRECO_IFE_ALLOCATION", 0, 2024),
        ("IT05PABPPLE2023", "LOSSRECO_IFE_ALLOCATION", 0, 2023),
        ("IT05PABPPLE2024", "LOSSRECO_CLOSING", 0, 2024),
        ("IT05PABPPLE2023", "LOSSRECO_CLOSING", 0, 2023),
        ("IT06ABCDE2024", "LOSSRECO_IFE_ALLOCATION", 0, 2024),
        ("IT06ABCDE2024", "LOSSRECO_CLOSING", 0, 2024),
    ]


def test_create_reinsurance_empty_pairs(tmp_path: Path) -> None:
    output = tmp_path / "REINSURANCE.csv"

    result = create_reinsurance(pairs=[], output_path=str(output))

    assert result["rows"] == 0
    rows = _read_csv(output)
    assert rows == [("GOC_ID", "VARIABLE_NAME", 1, "T")]


def test_create_mandatory_actuals(tmp_path: Path) -> None:
    output = tmp_path / "MANDATORY_ACTUALS.csv"
    assert len(MANDATORY_ACTUALS_VARIABLE_NAMES) == 16  # guard the spec
    pairs = [
        _pair("IT05PABPPLE", 2024),
        _pair("IT05PABPPLE", 2023),
        _pair("IT06ABCDE", 2024),
    ]

    result = create_mandatory_actuals(pairs=pairs, output_path=str(output))

    assert result == {
        "output_path": str(output),
        "rows": 3 * 16,
        "columns": ["GOC_ID", "VARIABLE_NAME", "1"],
    }

    rows = _read_csv(output)
    assert rows[0] == ("GOC_ID", "VARIABLE_NAME", 1)
    assert len(rows) == 1 + 3 * 16

    expected_first = [
        ("IT05PABPPLE2024", v, 0) for v in MANDATORY_ACTUALS_VARIABLE_NAMES
    ]
    assert rows[1:17] == expected_first

    expected_second = [
        ("IT05PABPPLE2023", v, 0) for v in MANDATORY_ACTUALS_VARIABLE_NAMES
    ]
    assert rows[17:33] == expected_second

    expected_third = [
        ("IT06ABCDE2024", v, 0) for v in MANDATORY_ACTUALS_VARIABLE_NAMES
    ]
    assert rows[33:49] == expected_third


def test_create_mandatory_actuals_empty_pairs(tmp_path: Path) -> None:
    output = tmp_path / "MANDATORY_ACTUALS.csv"

    result = create_mandatory_actuals(pairs=[], output_path=str(output))

    assert result["rows"] == 0
    rows = _read_csv(output)
    assert rows == [("GOC_ID", "VARIABLE_NAME", 1)]


PROJECTION_PARAMS_INPUT_ROWS = [
    ("PROJECTED_PERIODS", 110),
    ("SCENARIO_TIMESTEP", "YEARLY"),
    ("CF_TIMESTEP", "SEMESTRIAL"),
    ("REPORTING_MONTH", "12_DECEMBER"),
    ("PROJ_PERIOD_TIMESTEP", "YEARLY"),
    ("COVERAGE_UNIT_FORMAT", "COVERAGE_UNIT"),
    ("OCI_OPTION_APPROACH", "INDIRECT"),
    ("OCI_OPTION_RA_PROXY", "1_PROP_PVFC"),
    ("SCENARIO_MANAGEMENT", "CENTRAL"),
    ("FX_OPENING_DATE", "1M25"),
    ("FX_OPENING_RATE_TYPE", "YTD_VALUE"),
    ("FX_AVERAGE_DATE", "HY25"),
    ("FX_AVERAGE_RATE_TYPE", "YTD_VALUE"),
    ("FX_CLOSING_DATE", "FY25"),
    ("FX_CLOSING_RATE_TYPE", "YTD_VALUE"),
    ("FX_REPORTING_DATE", 20251231),
    ("FX_MANAGEMENT", "CENTRAL"),
    ("ACTUARIAL_AOM_ACQ_CF", "SEPARATE_INPUT"),
    ("OCI_OPTION_CF", "USE_EXISTING_INPUT"),
]


def _build_projection_parameters_fixture(path: Path) -> None:
    rows: list[tuple] = [("PARAMETER", "VALUE")]
    rows.extend(PROJECTION_PARAMS_INPUT_ROWS)
    _write_csv(path, rows)


def test_update_projection_parameters_entity_h2_2025(tmp_path: Path) -> None:
    fixture = tmp_path / "PROJECTION_PARAMETERS_ENTITY.csv"
    output = tmp_path / "out.csv"
    _build_projection_parameters_fixture(fixture)

    result = update_projection_parameters_entity(
        input_path=str(fixture),
        output_path=str(output),
        year=2025,
        semester=2,
    )

    assert result["rows_updated"] == 6
    assert set(result["parameters_updated"]) == {
        "CF_TIMESTEP",
        "REPORTING_MONTH",
        "FX_OPENING_DATE",
        "FX_AVERAGE_DATE",
        "FX_CLOSING_DATE",
        "FX_REPORTING_DATE",
    }

    rows = _read_csv(output)
    by_param = {row[0]: row[1] for row in rows[1:]}
    # Edits applied
    assert by_param["CF_TIMESTEP"] == "SEMESTRIAL"
    assert by_param["REPORTING_MONTH"] == "12_DECEMBER"
    assert by_param["FX_OPENING_DATE"] == "1M25"
    assert by_param["FX_AVERAGE_DATE"] == "HY25"
    assert by_param["FX_CLOSING_DATE"] == "FY25"
    # CSV doesn't preserve the str/int distinction; the int-coercion in _read_csv
    # turns "20251231" into int 20251231.
    assert by_param["FX_REPORTING_DATE"] == 20251231
    # Unchanged values
    assert by_param["PROJECTED_PERIODS"] == 110
    assert by_param["SCENARIO_TIMESTEP"] == "YEARLY"
    assert by_param["FX_MANAGEMENT"] == "CENTRAL"
    assert by_param["OCI_OPTION_CF"] == "USE_EXISTING_INPUT"
    # Header preserved
    assert rows[0] == ("PARAMETER", "VALUE")


def test_update_projection_parameters_entity_h1_2025(tmp_path: Path) -> None:
    fixture = tmp_path / "PROJECTION_PARAMETERS_ENTITY.csv"
    output = tmp_path / "out.csv"
    _build_projection_parameters_fixture(fixture)

    update_projection_parameters_entity(
        input_path=str(fixture),
        output_path=str(output),
        year=2025,
        semester=1,
    )

    rows = _read_csv(output)
    by_param = {row[0]: row[1] for row in rows[1:]}
    assert by_param["REPORTING_MONTH"] == "6_JUNE"
    assert by_param["FX_AVERAGE_DATE"] == "Q125"
    assert by_param["FX_CLOSING_DATE"] == "HY25"
    assert by_param["FX_REPORTING_DATE"] == 20250630
    assert by_param["FX_OPENING_DATE"] == "1M25"
    assert by_param["CF_TIMESTEP"] == "SEMESTRIAL"


def test_update_projection_parameters_entity_drops_extra_columns(tmp_path: Path) -> None:
    """Even if the input has a stray third column, the output is 2-column."""
    fixture = tmp_path / "input.csv"
    output = tmp_path / "out.csv"
    _write_csv(
        fixture,
        [
            ("PARAMETER", "VALUE", "Notes"),
            ("CF_TIMESTEP", "SEMESTRIAL", "should disappear"),
        ],
    )

    update_projection_parameters_entity(
        input_path=str(fixture),
        output_path=str(output),
        year=2025,
        semester=2,
    )

    rows = _read_csv(output)
    assert all(len(r) == 2 for r in rows)
    assert rows[0] == ("PARAMETER", "VALUE")
    assert rows[1] == ("CF_TIMESTEP", "SEMESTRIAL")


def test_update_projection_parameters_entity_invalid_semester(tmp_path: Path) -> None:
    fixture = tmp_path / "PROJECTION_PARAMETERS_ENTITY.csv"
    _build_projection_parameters_fixture(fixture)

    with pytest.raises(ValueError, match="semester must be 1 or 2"):
        update_projection_parameters_entity(
            input_path=str(fixture),
            output_path=str(tmp_path / "out.csv"),
            year=2025,
            semester=3,
        )


def _build_mp_goc_seg_fixture(path: Path, rows: list[tuple]) -> None:
    """rows: list of (col_a, col_b, col_c, col_d). Header row is added."""
    out: list[tuple] = [("GOC_SEG_ID", "GOC_ID", "SEG_ID", "ALLOCATION_RATIO")]
    out.extend(rows)
    _write_csv(path, out)


def test_update_mp_goc_seg_rewrites_only_perimeter_rows(tmp_path: Path) -> None:
    fixture = tmp_path / "in.csv"
    output = tmp_path / "out.csv"
    _build_mp_goc_seg_fixture(
        fixture,
        [
            ("IT05PABPPLE2025_02_P&C", "IT05PABPPLE2025", "02_P&C", 1),
            ("IT05PABPPLE2024_02_P&C", "IT05PABPPLE2024", "02_P&C", 1),
            ("IT05RRIEEBB2025_02_P&C", "IT05RRIEEBB2025", "02_P&C", 1),
            ("IT05RRIEEBB2024_02_P&C", "IT05RRIEEBB2024", "02_P&C", 1),
            ("IT06ABCDE2024_02_P&C", "IT06ABCDE2024", "02_P&C", 1),
        ],
    )

    result = update_mp_goc_seg(
        input_path=str(fixture),
        output_path=str(output),
        health_perimeter_gocs=["IT05RRIEEBB"],
    )

    assert result == {"output_path": str(output), "rows_in_perimeter": 2}

    rows = _read_csv(output)
    assert rows == [
        ("GOC_SEG_ID", "GOC_ID", "SEG_ID", "ALLOCATION_RATIO"),
        ("IT05PABPPLE2025_02_P&C", "IT05PABPPLE2025", "02_P&C", 1),
        ("IT05PABPPLE2024_02_P&C", "IT05PABPPLE2024", "02_P&C", 1),
        ("IT05RRIEEBB2025_02_HLTH_PC", "IT05RRIEEBB2025", "02_HLTH_PC", 1),
        ("IT05RRIEEBB2024_02_HLTH_PC", "IT05RRIEEBB2024", "02_HLTH_PC", 1),
        ("IT06ABCDE2024_02_P&C", "IT06ABCDE2024", "02_P&C", 1),
    ]


def test_update_mp_goc_seg_empty_perimeter_changes_nothing(tmp_path: Path) -> None:
    fixture = tmp_path / "in.csv"
    output = tmp_path / "out.csv"
    _build_mp_goc_seg_fixture(
        fixture,
        [("IT05RRIEEBB2025_02_P&C", "IT05RRIEEBB2025", "02_P&C", 1)],
    )

    result = update_mp_goc_seg(
        input_path=str(fixture),
        output_path=str(output),
        health_perimeter_gocs=[],
    )

    assert result["rows_in_perimeter"] == 0
    rows = _read_csv(output)
    assert rows[1] == ("IT05RRIEEBB2025_02_P&C", "IT05RRIEEBB2025", "02_P&C", 1)


def test_update_mp_goc_seg_skips_short_or_missing_goc(tmp_path: Path) -> None:
    fixture = tmp_path / "in.csv"
    output = tmp_path / "out.csv"
    _build_mp_goc_seg_fixture(
        fixture,
        [
            ("IT05RRIEEBB2025_02_P&C", "IT05RRIEEBB2025", "02_P&C", 1),  # ok
            ("X_02_P&C", "SHORT", "02_P&C", 1),  # B too short -> skipped
            ("Y_02_P&C", None, "02_P&C", 1),  # B None -> skipped
        ],
    )

    update_mp_goc_seg(
        input_path=str(fixture),
        output_path=str(output),
        health_perimeter_gocs=["IT05RRIEEBB", "SHORT", ""],
    )

    rows = _read_csv(output)
    assert rows[1] == ("IT05RRIEEBB2025_02_HLTH_PC", "IT05RRIEEBB2025", "02_HLTH_PC", 1)
    assert rows[2] == ("X_02_P&C", "SHORT", "02_P&C", 1)
    assert rows[3] == ("Y_02_P&C", None, "02_P&C", 1)


def test_update_mp_goc_seg_accepts_tab_separated_input(tmp_path: Path) -> None:
    """A tab-separated input must be parsed correctly (delimiter auto-detect)."""
    fixture = tmp_path / "in.csv"
    fixture.write_text(
        "\n".join([
            "GOC_SEG_ID\tGOC_ID\tSEG_ID\tALLOCATION_RATIO",
            "IT05RRIEEBB2025_02_P&C\tIT05RRIEEBB2025\t02_P&C\t1",
            "IT05RRIEEBB2024_02_P&C\tIT05RRIEEBB2024\t02_P&C\t1",
            "IT05PABPPLE2024_02_P&C\tIT05PABPPLE2024\t02_P&C\t1",
        ]),
        encoding="utf-8-sig",
    )

    output = tmp_path / "out.csv"
    result = update_mp_goc_seg(
        input_path=str(fixture),
        output_path=str(output),
        health_perimeter_gocs=["IT05RRIEEBB"],
    )

    assert result["rows_in_perimeter"] == 2
    rows = _read_csv(output)
    assert rows[1] == ("IT05RRIEEBB2025_02_HLTH_PC", "IT05RRIEEBB2025", "02_HLTH_PC", 1)
    assert rows[2] == ("IT05RRIEEBB2024_02_HLTH_PC", "IT05RRIEEBB2024", "02_HLTH_PC", 1)
    assert rows[3] == ("IT05PABPPLE2024_02_P&C", "IT05PABPPLE2024", "02_P&C", 1)


def test_update_mp_goc_seg_truncates_trailing_empty_columns(tmp_path: Path) -> None:
    """Input rows with trailing empty cells must not surface as ``1;;;`` in output."""
    fixture = tmp_path / "in.csv"
    # Write the file manually so we can control extra trailing ';' fields.
    with open(fixture, "w", newline="", encoding="utf-8-sig") as f:
        f.write("GOC_SEG_ID;GOC_ID;SEG_ID;ALLOCATION_RATIO;;;\n")
        f.write("IT05RRIEEBB2024_02_P&C;IT05RRIEEBB2024;02_P&C;1;;;\n")
        f.write("IT05PABPPLE2024_02_P&C;IT05PABPPLE2024;02_P&C;1;;;\n")

    output = tmp_path / "out.csv"
    update_mp_goc_seg(
        input_path=str(fixture),
        output_path=str(output),
        health_perimeter_gocs=["IT05RRIEEBB"],
    )

    # The raw bytes must end each line after the 4th field (no trailing ';').
    text = output.read_text(encoding="utf-8-sig")
    for line in text.splitlines():
        assert line.count(";") == 3, f"expected 3 separators, got {line!r}"

    # And the parsed rows must still have exactly 4 cells.
    rows = _read_csv(output)
    assert all(len(r) == 4 for r in rows)
    assert rows[1] == ("IT05RRIEEBB2024_02_HLTH_PC", "IT05RRIEEBB2024", "02_HLTH_PC", 1)
    assert rows[2] == ("IT05PABPPLE2024_02_P&C", "IT05PABPPLE2024", "02_P&C", 1)



def _build_aom_impact_fixture(path: Path, rows: list[tuple]) -> None:
    """rows: list of (goc_id, step_id, value). Header is added."""
    out: list[tuple] = [("GOC_ID", "STEP_ID", 1)]
    out.extend(rows)
    _write_csv(path, out)


def test_append_actuarial_aom_impact_appends_and_sorts(tmp_path: Path) -> None:
    fixture = tmp_path / "in.csv"
    output = tmp_path / "out.csv"
    _build_aom_impact_fixture(
        fixture,
        [
            # Pre-existing historical rows, intentionally out of order
            ("IT05PABPPLE2024", "PVFC_LIC_UNWIND", 0),
            ("IT05PABPPLE2024", "RA_LIC_OP", 0),
            ("IT05PABPPLE2025", "PVFC_LIC_UNWIND", 0),
        ],
    )

    pairs = [
        {"goc_id": "IT05PABPPLE2025", "goc": "IT05PABPPLE", "year": 2025},
        {"goc_id": "IT06ABCDE2025", "goc": "IT06ABCDE", "year": 2025},
    ]
    step_pairs = [
        ("DA_LIC_OP", 0),
        ("DA_LIC_CLO", 0),
    ]

    result = append_actuarial_aom_impact(
        input_path=str(fixture),
        output_path=str(output),
        pairs=pairs,
        step_id_value_pairs=step_pairs,
    )

    assert result == {
        "output_path": str(output),
        "rows_appended": 4,  # 2 pairs * 2 steps
        "rows_total": 7,  # 3 existing + 4 new
    }

    rows = _read_csv(output)
    assert rows[0] == ("GOC_ID", "STEP_ID", 1)
    assert rows[1:] == [
        ("IT05PABPPLE2024", "PVFC_LIC_UNWIND", 0),
        ("IT05PABPPLE2024", "RA_LIC_OP", 0),
        ("IT05PABPPLE2025", "DA_LIC_CLO", 0),
        ("IT05PABPPLE2025", "DA_LIC_OP", 0),
        ("IT05PABPPLE2025", "PVFC_LIC_UNWIND", 0),
        ("IT06ABCDE2025", "DA_LIC_CLO", 0),
        ("IT06ABCDE2025", "DA_LIC_OP", 0),
    ]


def test_append_actuarial_aom_impact_empty_inputs(tmp_path: Path) -> None:
    fixture = tmp_path / "in.csv"
    output = tmp_path / "out.csv"
    _build_aom_impact_fixture(
        fixture,
        [("IT05PABPPLE2024", "PVFC_LIC_UNWIND", 0)],
    )

    result = append_actuarial_aom_impact(
        input_path=str(fixture),
        output_path=str(output),
        pairs=[],
        step_id_value_pairs=[],
    )
    assert result["rows_appended"] == 0
    assert result["rows_total"] == 1

    rows = _read_csv(output)
    assert rows == [
        ("GOC_ID", "STEP_ID", 1),
        ("IT05PABPPLE2024", "PVFC_LIC_UNWIND", 0),
    ]


def test_append_actuarial_aom_impact_skips_blank_step_ids(tmp_path: Path) -> None:
    fixture = tmp_path / "in.csv"
    output = tmp_path / "out.csv"
    _build_aom_impact_fixture(fixture, [])

    pairs = [{"goc_id": "G2025", "goc": "G", "year": 2025}]
    step_pairs = [
        ("STEP_A", 1),
        ("", 2),
        ("  ", 3),
        (None, 4),
        ("STEP_B", 5),
    ]

    result = append_actuarial_aom_impact(
        input_path=str(fixture),
        output_path=str(output),
        pairs=pairs,
        step_id_value_pairs=step_pairs,
    )

    assert result["rows_appended"] == 2

    rows = _read_csv(output)
    assert rows[1:] == [
        ("G2025", "STEP_A", 1),
        ("G2025", "STEP_B", 5),
    ]


def _build_curve_id_param_fixture(path: Path, rows: list[tuple]) -> None:
    """rows: list of (goc_id, variable_name, col_c). Header is added."""
    out: list[tuple] = [("GOC_ID", "VARIABLE_NAME", 1)]
    out.extend(rows)
    _write_csv(path, out)


def test_update_curve_id_param_fills_known_variables(tmp_path: Path) -> None:
    fixture = tmp_path / "in.csv"
    output = tmp_path / "out.csv"
    _build_curve_id_param_fixture(
        fixture,
        [
            ("IT05PABPPLE2024", "CLOSING_CURVE_ID", None),
            ("IT05PABPPLE2024", "OPENING_CURVE_ID", None),
            ("IT05PABPPLE2024", "CREDITED_RATE_CURVE_ID", None),
            ("IT06ABCDE2024", "CLOSING_CURVE_ID", None),
            ("IT06ABCDE2024", "CREDITED_RATE_CURVE_ID", None),
        ],
    )

    result = update_curve_id_param(
        input_path=str(fixture),
        output_path=str(output),
        closing_curve_name="pippo",
        opening_curve_name="carlo",
    )

    assert result == {"output_path": str(output), "rows_updated": 5}

    rows = _read_csv(output)
    assert rows == [
        ("GOC_ID", "VARIABLE_NAME", 1),
        ("IT05PABPPLE2024", "CLOSING_CURVE_ID", "pippo"),
        ("IT05PABPPLE2024", "OPENING_CURVE_ID", "carlo"),
        ("IT05PABPPLE2024", "CREDITED_RATE_CURVE_ID", "CR_IT05PABPPLE2024"),
        ("IT06ABCDE2024", "CLOSING_CURVE_ID", "pippo"),
        ("IT06ABCDE2024", "CREDITED_RATE_CURVE_ID", "CR_IT06ABCDE2024"),
    ]


def test_update_curve_id_param_leaves_other_variables_untouched(tmp_path: Path) -> None:
    fixture = tmp_path / "in.csv"
    output = tmp_path / "out.csv"
    _build_curve_id_param_fixture(
        fixture,
        [
            ("IT05PABPPLE2024", "CLOSING_CURVE_ID", None),
            ("IT05PABPPLE2024", "SOMETHING_ELSE", "preserved"),
            ("IT05PABPPLE2024", None, "also preserved"),
        ],
    )

    result = update_curve_id_param(
        input_path=str(fixture),
        output_path=str(output),
        closing_curve_name="pippo",
        opening_curve_name="carlo",
    )

    assert result["rows_updated"] == 1

    rows = _read_csv(output)
    assert rows == [
        ("GOC_ID", "VARIABLE_NAME", 1),
        ("IT05PABPPLE2024", "CLOSING_CURVE_ID", "pippo"),
        ("IT05PABPPLE2024", "SOMETHING_ELSE", "preserved"),
        ("IT05PABPPLE2024", None, "also preserved"),
    ]


# ---------------------------------------------------------------------------
# update_mp_goc
# ---------------------------------------------------------------------------
_MP_GOC_HEADERS = (
    "GOC_ID", "MEASUREMENT_MODEL", "ANNUAL_COHORT", "AOM_ID",
    "INCEPTION_CURVE_ID", "TIMING_INCEPTION_CURVE",
    "CSM_RELEASE_RATIO_CURVE_ID", "SHARE_TECH_EXP_ENTITY_SHARE",
    "OCI_OPTION", "OCI_OPTION_LRC", "OCI_OPTION_LIC", "GOC_DURATION",
    "GOC_TYPE_IF_NB", "GOC_CURRENCY", "REPORTING_CURRENCY",
    "GOC_TYPE_REINSURANCE", "AGGREG_1_ID", "AGGREG_2_ID", "AGGREG_3_ID",
    "AGGREG_4_ID", "AGGREG_5_ID",
)
_MP_GOC_COHORT_YEARS = (2014, 2018, 2021, 2022, 2024, 2025)


def _build_mp_goc_fixture(
    path: Path,
    original_f: str = "ORIGINAL_F",
    col_r_by_cohort: dict[int, str] | None = None,
) -> None:
    """Build an MP_GOC fixture. ``col_r_by_cohort`` lets a test set the
    AGGREG_2_ID (column R, index 17) per cohort row so the column-P logic
    can be exercised.
    """
    col_r_map = col_r_by_cohort or {}
    rows: list[tuple] = [_MP_GOC_HEADERS]
    for y in _MP_GOC_COHORT_YEARS:
        row = [f"IT{y}", "PAA", y, 20] + [None] * (len(_MP_GOC_HEADERS) - 4)
        row[5] = original_f
        if y in col_r_map:
            row[17] = col_r_map[y]
        rows.append(tuple(row))
    _write_csv(path, rows)


def _read_mp_goc(path: Path) -> dict[int, dict]:
    rows = _read_csv(path)
    out: dict[int, dict] = {}
    for r in rows[1:]:
        out[r[2]] = {"E": r[4], "F": r[5], "L": r[11], "P": r[15], "R": r[17]}
    return out


def test_update_mp_goc_h2_diretto(tmp_path: Path) -> None:
    fixture = tmp_path / "MP_GOC.csv"
    output = tmp_path / "out.csv"
    _build_mp_goc_fixture(fixture)

    result = update_mp_goc(
        input_path=str(fixture),
        output_path=str(output),
        year=2025,
        semester=2,
        business_type="Diretto",
    )
    assert result["rows_changed"] == len(_MP_GOC_COHORT_YEARS)

    rows = _read_mp_goc(output)
    # cohort <= 2015 -> fixed value, independent of semester
    assert rows[2014]["E"] == "20211231_ITA_LP100_AVG"
    assert rows[2018]["E"] == "20181231_ITA_LP100_AVG"
    assert rows[2021]["E"] == "20211231_ITA_LP100_AVG"
    assert rows[2022]["E"] == "20221231_ITA_LP100_FY22_AVG"
    assert rows[2024]["E"] == "20241231_EUR_LP100_FY24_AVG"
    assert rows[2025]["E"] == "20251231_EUR_LP100_FY25_AVG"

    assert rows[2025]["F"] == "13_YEAR_END"
    for y in (2014, 2018, 2021, 2022, 2024):
        assert rows[y]["F"] == "ORIGINAL_F"

    # L = max(0, year - 1 - curve_year) * 12 where curve_year is parsed
    # from the first 4 chars of column E. cohort 2014 -> col E starts
    # with "2021" -> curve_year = 2021, so L = (2025-1-2021)*12 = 36.
    assert rows[2014]["L"] == 36
    assert rows[2024]["L"] == 0  # 2025 - 1 - 2024 = 0
    assert rows[2025]["L"] == 0  # max(0, -1) * 12 = 0

    # Col R is empty in this fixture -> col P falls back to 2_RE_ASSUMED
    for y in _MP_GOC_COHORT_YEARS:
        assert rows[y]["P"] == "2_RE_ASSUMED"


def test_update_mp_goc_col_p_driven_by_col_r(tmp_path: Path) -> None:
    """Col P now reads col R (AGGREG_2_ID), not business_type."""
    fixture = tmp_path / "MP_GOC.csv"
    output = tmp_path / "out.csv"
    _build_mp_goc_fixture(
        fixture,
        col_r_by_cohort={
            2014: "PAA_Ceded",          # ceded
            2018: "PAA_Direct",         # other -> assumed
            2021: "PAA_Ceded",          # ceded
            2022: "",                   # blank -> assumed
            2024: "  PAA_Ceded  ",      # whitespace-trimmed match -> ceded
            2025: "SomethingElse",      # other -> assumed
        },
    )

    # business_type is irrelevant for col P now — pass any valid value.
    update_mp_goc(
        input_path=str(fixture),
        output_path=str(output),
        year=2025,
        semester=2,
        business_type="Diretto",
    )

    rows = _read_mp_goc(output)
    assert rows[2014]["P"] == "3_RE_CEDED_NON_RETRO"
    assert rows[2018]["P"] == "2_RE_ASSUMED"
    assert rows[2021]["P"] == "3_RE_CEDED_NON_RETRO"
    assert rows[2022]["P"] == "2_RE_ASSUMED"
    assert rows[2024]["P"] == "3_RE_CEDED_NON_RETRO"
    assert rows[2025]["P"] == "2_RE_ASSUMED"


def test_update_mp_goc_h1_other_curve_dates(tmp_path: Path) -> None:
    """Sanity check that semester=1 dates still flow through correctly."""
    fixture = tmp_path / "MP_GOC.csv"
    output = tmp_path / "out.csv"
    _build_mp_goc_fixture(fixture)

    update_mp_goc(
        input_path=str(fixture),
        output_path=str(output),
        year=2025,
        semester=1,
        business_type="Ceduto",
    )
    rows = _read_mp_goc(output)
    # cohort <= 2015 -> fixed value, semester does not affect it
    assert rows[2014]["E"] == "20211231_ITA_LP100_AVG"
    assert rows[2022]["E"] == "20220630_ITA_LP100_FY22_AVG"
    assert rows[2025]["E"] == "20250630_EUR_LP100_FY25_AVG"
    assert rows[2025]["F"] == "7_JULY"
    assert rows[2024]["F"] == "ORIGINAL_F"
    # With no col R set, every row still defaults to 2_RE_ASSUMED
    # regardless of business_type.
    for y in _MP_GOC_COHORT_YEARS:
        assert rows[y]["P"] == "2_RE_ASSUMED"


def test_update_mp_goc_idempotent(tmp_path: Path) -> None:
    fixture = tmp_path / "MP_GOC.csv"
    out1 = tmp_path / "out1.csv"
    out2 = tmp_path / "out2.csv"
    _build_mp_goc_fixture(fixture)

    update_mp_goc(
        input_path=str(fixture), output_path=str(out1),
        year=2025, semester=2, business_type="Diretto",
    )
    update_mp_goc(
        input_path=str(out1), output_path=str(out2),
        year=2025, semester=2, business_type="Diretto",
    )
    assert _read_csv(out1) == _read_csv(out2)


def test_update_mp_goc_invalid_business_type(tmp_path: Path) -> None:
    fixture = tmp_path / "MP_GOC.csv"
    output = tmp_path / "out.csv"
    _build_mp_goc_fixture(fixture)
    with pytest.raises(ValueError):
        update_mp_goc(
            input_path=str(fixture), output_path=str(output),
            year=2025, semester=2, business_type="Direct",
        )


def test_update_mp_goc_invalid_semester(tmp_path: Path) -> None:
    fixture = tmp_path / "MP_GOC.csv"
    output = tmp_path / "out.csv"
    _build_mp_goc_fixture(fixture)
    with pytest.raises(ValueError):
        update_mp_goc(
            input_path=str(fixture), output_path=str(output),
            year=2025, semester=3, business_type="Diretto",
        )


def _build_input_sunrise_rows_fixture(
    path: Path, rows: list[tuple[str, int, str, float, float]]
) -> None:
    """Build an Input_Sunrise workbook with 5 columns. ``rows`` is data only."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input_Sunrise"
    ws.cell(row=1, column=1, value="GOC")
    ws.cell(row=1, column=2, value="ANNO")
    ws.cell(row=1, column=3, value="PERIMETRO")
    ws.cell(row=1, column=4, value="SINISTRI")
    ws.cell(row=1, column=5, value="RISERVA_SINISTRI")
    for i, (goc, yr, peri, sin, ris) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=goc)
        ws.cell(row=i, column=2, value=yr)
        ws.cell(row=i, column=3, value=peri)
        ws.cell(row=i, column=4, value=sin)
        ws.cell(row=i, column=5, value=ris)
    wb.save(path)


def _build_transcodifica_csv(path: Path, rows: list[tuple]) -> None:
    """Build a Transcodifica CSV (``;`` separator, header + data rows).

    Each row should be (GOC, Aggregation1, Aggregation2, H_NH).
    Shorter tuples are padded with empty cells.
    """
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["GOC", "Aggregation1", "Aggregation2", "H-NH"])
        for r in rows:
            padded = list(r) + [""] * (4 - len(r))
            w.writerow(padded[:4])


def test_extract_health_perimeter_gocs(tmp_path: Path) -> None:
    fixture = tmp_path / "transcodifica.csv"
    _build_transcodifica_csv(
        fixture,
        [
            ("IT05PABPPLE", "Comm P&C", "PAA_Direct", "NH"),
            ("IT05RRIEEBB", "Health line A", "PAA_Direct", "H"),
            ("IT05RRIHEBB", "Health line B", "PAA_Direct", "h"),  # lower-case still counts
            ("IT05PABLMIN", "Comm P&C", "PAA_Direct", "NH"),
            ("IT05RRIHMAF", "Health line C", "PAA_Direct", "H"),
            # blanks and duplicates
            ("IT05RRIEEBB", "Health line A dup", "PAA_Direct", "H"),
            ("", "ignored row", "", "H"),
        ],
    )

    result = extract_health_perimeter_gocs(str(fixture))

    # Order preserved (first occurrence), duplicates dropped, NH excluded,
    # blank GOC ignored, lower-case "h" matched.
    assert result == ["IT05RRIEEBB", "IT05RRIHEBB", "IT05RRIHMAF"]


def test_extract_health_perimeter_gocs_no_h(tmp_path: Path) -> None:
    fixture = tmp_path / "transcodifica.csv"
    _build_transcodifica_csv(
        fixture,
        [
            ("IT05PABPPLE", "Comm P&C", "PAA_Direct", "NH"),
            ("IT05PABLMIN", "Comm P&C", "PAA_Direct", "NH"),
        ],
    )
    assert extract_health_perimeter_gocs(str(fixture)) == []


def test_extract_health_perimeter_gocs_auto_detects_delimiter(tmp_path: Path) -> None:
    """CSV files exported with ',' or tab as separator must also work."""
    for label, delimiter in [("comma", ","), ("tab", "\t")]:
        fixture = tmp_path / f"transcodifica_{label}.csv"
        lines = [
            ("GOC_ID", "Aggregation1", "Aggregation2", "H-NH"),
            ("IT05RRIBOND", "Commercial P&C", "PAA_Ceded", "NH"),
            ("IT05RRIEEBB", "P&C - Health", "PAA_Ceded", "H"),
        ]
        fixture.write_text(
            "\n".join(delimiter.join(row) for row in lines),
            encoding="utf-8-sig",
        )
        result = extract_health_perimeter_gocs(str(fixture))
        assert result == ["IT05RRIEEBB"], f"failed for {label}-separated file: {result}"


def test_create_mp_model_point_happy_path(tmp_path: Path) -> None:
    """Two files (current + previous), sums per (GoC, year), aggregations."""
    curr = tmp_path / "1.1_2025.12.31_AAI_Ceded.xlsx"
    prev = tmp_path / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_rows_fixture(
        curr,
        [
            ("IT05PABPPLE", 2025, "INTERNA", 100.0, 200.0),
            ("IT05PABPPLE", 2024, "INTERNA", 50.0, 75.0),
            # Same (GoC, year) — should be summed
            ("IT05PABPPLE", 2024, "INTERNA", 25.0, 25.0),
        ],
    )
    _build_input_sunrise_rows_fixture(
        prev,
        [
            ("IT05PABPPLE", 2024, "INTERNA", 80.0, 90.0),
        ],
    )
    transcodifica = tmp_path / "3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(
        transcodifica,
        [("IT05PABPPLE", "Commercial P&C", "PAA_Direct")],
    )
    output = tmp_path / "MP_ModelPoint.csv"

    result = create_mp_model_point(
        sources=[(str(curr), 2025), (str(prev), 2024)],
        transcodifica_path=str(transcodifica),
        output_path=str(output),
        year=2025,
    )

    assert result["rows"] == 3  # 2 from current, 1 from previous
    assert result["columns"][0] == "Primary_Key"

    rows = _read_csv(output)
    # Header
    assert rows[0] == tuple(result["columns"])
    # Current-year rows: 2025 then 2024 (descending), @Closing
    assert rows[1] == (
        "IT05PABPPLE2025@2025", "IT05PABPPLE2025", "IT05PABPPLE@Closing",
        2025, 2025, "Commercial P&C", "PAA_Direct",
        200.0, 0, 100.0, 0,
    )
    # 2024 is summed: 50+25=75 sinistri, 75+25=100 riserva
    assert rows[2] == (
        "IT05PABPPLE2024@2025", "IT05PABPPLE2024", "IT05PABPPLE@Closing",
        2024, 2025, "Commercial P&C", "PAA_Direct",
        100.0, 0, 75.0, 0,
    )
    # Previous-year file row: 2024 only, @Opening
    assert rows[3] == (
        "IT05PABPPLE2024@2024", "IT05PABPPLE2024", "IT05PABPPLE@Opening",
        2024, 2024, "Commercial P&C", "PAA_Direct",
        90.0, 0, 80.0, 0,
    )


def test_create_mp_model_point_folds_pre_horizon_years(tmp_path: Path) -> None:
    """Years older than year-15 are summed into the oldest year of the horizon."""
    curr = tmp_path / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_rows_fixture(
        curr,
        [
            # min_year = 2025 - 15 = 2010 — these three are pre-horizon
            ("IT05PABPPLE", 2007, "INTERNA", 10.0, 1.0),
            ("IT05PABPPLE", 2008, "INTERNA", 20.0, 2.0),
            ("IT05PABPPLE", 2009, "INTERNA", 30.0, 3.0),
            # In-horizon
            ("IT05PABPPLE", 2010, "INTERNA", 5.0, 0.5),
            ("IT05PABPPLE", 2025, "INTERNA", 100.0, 200.0),
        ],
    )
    transcodifica = tmp_path / "3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(transcodifica, [])
    output = tmp_path / "MP_ModelPoint.csv"

    create_mp_model_point(
        sources=[(str(curr), 2025)],
        transcodifica_path=str(transcodifica),
        output_path=str(output),
        year=2025,
    )

    rows = _read_csv(output)
    # Two rows: 2025 (latest) and 2010 (folded pre-horizon)
    assert len(rows) == 3  # header + 2 data rows
    assert rows[1][3] == 2025  # Accident_Year of newest row
    assert rows[2][3] == 2010  # Accident_Year of oldest row
    # 2010 row: 5 + (10+20+30) = 65 sinistri, 0.5 + (1+2+3) = 6.5 riserva
    assert rows[2][9] == 65.0  # Claims_Paid
    assert rows[2][7] == 6.5  # EAXA_Reserve


def test_create_mp_model_point_folds_pre_horizon_creates_min_year_row(
    tmp_path: Path,
) -> None:
    """If min_year has no native data, the folded row is still emitted.

    Since the analysis-year auto-fill rule was added, this GoC also
    gets a synthetic ``Accident_Year=year`` row with zeros (rows are
    emitted newest-first within each ``(anno_rif, goc)`` group).
    """
    curr = tmp_path / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_rows_fixture(
        curr,
        [
            # Only pre-horizon data
            ("IT05PABPPLE", 2007, "INTERNA", 10.0, 1.0),
            ("IT05PABPPLE", 2008, "INTERNA", 20.0, 2.0),
        ],
    )
    transcodifica = tmp_path / "3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(transcodifica, [])
    output = tmp_path / "MP_ModelPoint.csv"

    create_mp_model_point(
        sources=[(str(curr), 2025)],
        transcodifica_path=str(transcodifica),
        output_path=str(output),
        year=2025,
    )

    rows = _read_csv(output)
    assert len(rows) == 3  # header + auto-filled 2025 row + folded 2010 row
    # Newest first inside the group
    assert rows[1][3] == 2025
    assert rows[1][9] == 0  # Claims_Paid = 0 on the synthetic row
    assert rows[1][7] == 0  # EAXA_Reserve = 0 on the synthetic row
    assert rows[2][3] == 2010
    assert rows[2][9] == 30.0  # 10 + 20


def test_create_mp_model_point_missing_transcodifica_entry_is_empty(
    tmp_path: Path,
) -> None:
    curr = tmp_path / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_rows_fixture(
        curr,
        [("ITUNKNOWN", 2025, "INTERNA", 1.0, 2.0)],
    )
    transcodifica = tmp_path / "3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(transcodifica, [])  # empty
    output = tmp_path / "MP_ModelPoint.csv"

    create_mp_model_point(
        sources=[(str(curr), 2025)],
        transcodifica_path=str(transcodifica),
        output_path=str(output),
        year=2025,
    )

    rows = _read_csv(output)
    assert rows[1][0] == "ITUNKNOWN2025@2025"
    assert rows[1][5] is None  # Aggregation1 missing
    assert rows[1][6] is None  # Aggregation2 missing


def test_create_mp_model_point_drops_all_zero_gocs(tmp_path: Path) -> None:
    """A GoC whose SINISTRI and RISERVA_SINISTRI are 0 everywhere is excluded."""
    curr = tmp_path / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_rows_fixture(
        curr,
        [
            # GoC with mixed values — kept
            ("IT_KEEP", 2025, "INTERNA", 100.0, 200.0),
            ("IT_KEEP", 2024, "INTERNA", 0.0, 0.0),
            # GoC with all-zero values across every row — dropped
            ("IT_DROP_ALLZERO", 2025, "INTERNA", 0.0, 0.0),
            ("IT_DROP_ALLZERO", 2024, "INTERNA", 0.0, 0.0),
            ("IT_DROP_ALLZERO", 2023, "INTERNA", 0.0, 0.0),
            # Edge case: a single non-zero row keeps the GoC
            ("IT_KEEP_ONE", 2024, "INTERNA", 0.0, 0.0),
            ("IT_KEEP_ONE", 2023, "INTERNA", 0.0, 0.01),
        ],
    )
    transcodifica = tmp_path / "3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(transcodifica, [])
    output = tmp_path / "MP_ModelPoint.csv"

    create_mp_model_point(
        sources=[(str(curr), 2025)],
        transcodifica_path=str(transcodifica),
        output_path=str(output),
        year=2025,
    )

    rows = _read_csv(output)
    goc_ids_in_output = {r[1] for r in rows[1:]}
    # IT_KEEP and IT_KEEP_ONE survive (at least one non-zero row each)
    assert any(g.startswith("IT_KEEP") for g in goc_ids_in_output)
    assert any(g.startswith("IT_KEEP_ONE") for g in goc_ids_in_output)
    # IT_DROP_ALLZERO is excluded entirely
    assert not any(g.startswith("IT_DROP_ALLZERO") for g in goc_ids_in_output)


def test_create_mp_model_point_keeps_goc_with_value_in_other_file(
    tmp_path: Path,
) -> None:
    """All-zero in current-year file but non-zero in previous-year file -> kept."""
    curr = tmp_path / "1.1_2025.12.31_AAI_Ceded.xlsx"
    prev = tmp_path / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_rows_fixture(
        curr,
        [("IT_CROSS", 2025, "INTERNA", 0.0, 0.0)],
    )
    _build_input_sunrise_rows_fixture(
        prev,
        [("IT_CROSS", 2024, "INTERNA", 5.0, 0.0)],
    )
    transcodifica = tmp_path / "3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(transcodifica, [])
    output = tmp_path / "MP_ModelPoint.csv"

    create_mp_model_point(
        sources=[(str(curr), 2025), (str(prev), 2024)],
        transcodifica_path=str(transcodifica),
        output_path=str(output),
        year=2025,
    )

    rows = _read_csv(output)
    # Both the @Closing (2025-file) and @Opening (2024-file) rows are emitted
    # because the GoC has at least one non-zero value somewhere.
    obs_keys = {r[2] for r in rows[1:]}
    assert "IT_CROSS@Closing" in obs_keys
    assert "IT_CROSS@Opening" in obs_keys

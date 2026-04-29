"""Tests for the Astra MP_GOC transformation."""
from __future__ import annotations

import openpyxl
import pytest

from excel_pipeline.skill import astra_transform_mp_goc

HEADERS = [
    "GOC_ID", "MEASUREMENT_MODEL", "ANNUAL_COHORT", "AOM_ID",
    "INCEPTION_CURVE_ID", "TIMING_INCEPTION_CURVE",
    "CSM_RELEASE_RATIO_CURVE_ID", "SHARE_TECH_EXP_ENTITY_SHARE",
    "OCI_OPTION", "OCI_OPTION_LRC", "OCI_OPTION_LIC", "GOC_DURATION",
    "GOC_TYPE_IF_NB", "GOC_CURRENCY", "REPORTING_CURRENCY",
    "GOC_TYPE_REINSURANCE", "AGGREG_1_ID", "AGGREG_2_ID", "AGGREG_3_ID",
    "AGGREG_4_ID", "AGGREG_5_ID",
]
COHORT_YEARS = [2014, 2018, 2021, 2022, 2024, 2025]


def _build_fixture(path, original_f="ORIGINAL_F"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for y in COHORT_YEARS:
        row = [f"IT{y}"] + [None] * (len(HEADERS) - 1)
        row[2] = y
        row[5] = original_f
        ws.append(row)
    wb.save(path)


def _read(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    return {
        ws.cell(row=r, column=3).value: {
            "E": ws.cell(row=r, column=5).value,
            "F": ws.cell(row=r, column=6).value,
            "L": ws.cell(row=r, column=12).value,
            "P": ws.cell(row=r, column=16).value,
        }
        for r in range(2, ws.max_row + 1)
    }


def test_year_end_direct(tmp_path):
    src = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    _build_fixture(src)

    result = astra_transform_mp_goc(
        input_path=str(src),
        output_path=str(out),
        valuation_date="2025-12-31",
        business_type="Direct",
    )
    assert result["rows_changed"] == len(COHORT_YEARS)

    rows = _read(out)
    assert rows[2014]["E"] == "20141231_ITA_LP100"
    assert rows[2018]["E"] == "20181231_ITA_LP100_AVG"
    assert rows[2021]["E"] == "20211231_ITA_LP100_AVG"
    assert rows[2022]["E"] == "20221231_ITA_LP100_FY22_AVG"
    assert rows[2024]["E"] == "20241231_EUR_LP100_FY24_AVG"
    assert rows[2025]["E"] == "20251231_EUR_LP100_FY25_AVG"

    # F changes only for cohort 2025.
    assert rows[2025]["F"] == "13_YEAR_END"
    for y in (2014, 2018, 2021, 2022, 2024):
        assert rows[y]["F"] == "ORIGINAL_F"

    # L = max(0, val_year - cohort_year) * 12.
    assert rows[2014]["L"] == (2025 - 2014) * 12
    assert rows[2024]["L"] == 12
    assert rows[2025]["L"] == 0

    # P from business_type.
    for y in COHORT_YEARS:
        assert rows[y]["P"] == "2_RE_ASSUMED"


def test_first_semester_ceduto(tmp_path):
    src = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    _build_fixture(src)

    astra_transform_mp_goc(
        input_path=str(src),
        output_path=str(out),
        valuation_date="2025-06-30",
        business_type="Ceduto",
    )
    rows = _read(out)
    assert rows[2014]["E"].endswith("0630_ITA_LP100")
    assert rows[2022]["E"] == "20220630_ITA_LP100_FY22_AVG"
    assert rows[2025]["E"] == "20250630_EUR_LP100_FY25_AVG"
    assert rows[2025]["F"] == "7_JULY"
    assert rows[2024]["F"] == "ORIGINAL_F"
    for y in COHORT_YEARS:
        assert rows[y]["P"] == "3_RE_CEDED_NON_RETRO"


def test_idempotent(tmp_path):
    src = tmp_path / "in.xlsx"
    out1 = tmp_path / "out1.xlsx"
    out2 = tmp_path / "out2.xlsx"
    _build_fixture(src)

    astra_transform_mp_goc(
        input_path=str(src),
        output_path=str(out1),
        valuation_date="2025-12-31",
        business_type="Direct",
    )
    astra_transform_mp_goc(
        input_path=str(out1),
        output_path=str(out2),
        valuation_date="2025-12-31",
        business_type="Direct",
    )
    assert _read(out1) == _read(out2)


def test_invalid_business_type(tmp_path):
    src = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    _build_fixture(src)
    with pytest.raises(ValueError):
        astra_transform_mp_goc(
            input_path=str(src),
            output_path=str(out),
            valuation_date="2025-12-31",
            business_type="Other",
        )

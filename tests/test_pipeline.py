"""Tests for excel_pipeline.pipeline (deterministic, no API calls)."""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest

from excel_pipeline.pipeline import (
    run_astra_phase1,
    run_phase1,
)


def _read_csv(
    path: Path,
    field_sep: str = ";",
    decimal_sep: str = ",",
) -> list[tuple]:
    """Read CSV rows; coerce numeric cells to int/float, '' to None.

    Defaults match Sunrise output (``;`` field separator, ``,`` decimal).
    Astra output uses ``,`` field separator with ``.`` decimal — pass
    ``field_sep=","`` and ``decimal_sep="."`` when reading Astra CSVs.
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=field_sep)
        rows: list[tuple] = []
        for raw in reader:
            cells = []
            for cell in raw:
                if cell == "":
                    cells.append(None)
                    continue
                try:
                    cells.append(int(cell))
                    continue
                except ValueError:
                    pass
                normalized = (
                    cell.replace(decimal_sep, ".")
                    if decimal_sep != "."
                    else cell
                )
                try:
                    cells.append(float(normalized))
                    continue
                except ValueError:
                    pass
                cells.append(cell)
            rows.append(tuple(cells))
        return rows


def _read_astra_csv(path: Path) -> list[tuple]:
    """Shortcut for reading an Astra-format CSV (``,`` sep, ``.`` decimal)."""
    return _read_csv(path, field_sep=",", decimal_sep=".")


def _write_csv(path: Path, rows: list[tuple]) -> None:
    """Write a CSV fixture using the project convention (``;`` separator)."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        for row in rows:
            writer.writerow(["" if v is None else v for v in row])


def _build_ceded_fixture(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AAI_P&C_Ceded_H_NH"
    ws.cell(row=1, column=27, value="GoC")
    ws.cell(row=2, column=27, value="Line of Business")
    for i, v in enumerate(
        ["Motor", "Property", "Motor", "Liability"], start=3,
    ):
        ws.cell(row=i, column=27, value=v)
    wb.save(path)


def _build_payment_patterns_fixture(path: Path) -> None:
    """Fixture with both the ra_AAI_REINS and pp_AAI_REINS sheets."""
    wb = openpyxl.Workbook()

    # --- ra_AAI_REINS ---
    ra = wb.active
    ra.title = "ra_AAI_REINS"
    ra.cell(row=1, column=7, value="GoC")
    for i, h in enumerate(["HY_2024", "FY_2024", "HY_2025", "FY_2025"], start=8):
        ra.cell(row=1, column=i, value=h)
    ra_data = [
        ("Motor", 100, 110, 120, 130),
        ("Property", 200, 210, 220, 230),
        ("Liability", 300, 310, 320, 330),
    ]
    for r, (goc, *vals) in enumerate(ra_data, start=2):
        ra.cell(row=r, column=7, value=goc)
        for j, v in enumerate(vals):
            ra.cell(row=r, column=8 + j, value=v)

    # --- pp_AAI_REINS ---
    pp = wb.create_sheet("pp_AAI_REINS")
    pp.cell(row=1, column=3, value="GoC")
    pp.cell(row=1, column=4, value="Year")
    for i in range(23):
        pp.cell(row=1, column=5 + i, value=str(i))
    pp_data = [
        ("Motor", "FY2024", 1000),
        ("Motor", "FY2025", 2000),
        ("Property", "FY2024", 5000),
        ("Property", "FY2025", 6000),
        ("Liability", "FY2024", 7000),
        ("Liability", "FY2025", 8000),
    ]
    for r, (goc, yr, seed) in enumerate(pp_data, start=2):
        pp.cell(row=r, column=3, value=goc)
        pp.cell(row=r, column=4, value=yr)
        for i in range(23):
            pp.cell(row=r, column=5 + i, value=seed + i)

    wb.save(path)


def _build_input_sunrise_workbook(
    path: Path,
    gocs: list[str | None],
    *,
    year: int = 2025,
    sinistri: float = 0.0,
    riserva: float = 0.0,
) -> None:
    """Workbook with the new Input_Sunrise sheet.

    By default each GoC is paired with the same ``year`` and zero
    SINISTRI / RISERVA — enough to exercise the row-count path.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input_Sunrise"
    ws.cell(row=1, column=1, value="GoC")
    ws.cell(row=1, column=2, value="Year")
    ws.cell(row=1, column=3, value="Perimetro")
    ws.cell(row=1, column=4, value="Sinistri")
    ws.cell(row=1, column=5, value="Riserva")
    for i, g in enumerate(gocs, start=2):
        ws.cell(row=i, column=1, value=g)
        ws.cell(row=i, column=2, value=year)
        ws.cell(row=i, column=4, value=sinistri)
        ws.cell(row=i, column=5, value=riserva)
    wb.save(path)


def test_run_phase1_emits_only_mp_model_point(tmp_path: Path) -> None:
    """Phase 1 currently runs only the MP_ModelPoint step."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["Motor", "Property"], year=2025, sinistri=100.0, riserva=50.0,
    )
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Property", "Liability"], year=2024, sinistri=80.0, riserva=40.0,
    )
    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [
            ("GOC", "Aggregation1", "Aggregation2"),
            ("Motor", "Agg1_Motor", "Agg2_Motor"),
            ("Property", "Agg1_Prop", "Agg2_Prop"),
        ],
    )
    # Payment_Patterns may still be uploaded but is not consumed here.
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI"), (14, "MPS")],
        year=2025,
        semester=2,
    )

    assert result["entities"] == [(6, "AAI"), (14, "MPS")]
    assert result["year"] == 2025
    assert result["semester"] == 2

    outputs = result["outputs"]
    assert outputs == [
        tmp_path / "MP_ModelPoint.csv",
        tmp_path / "MP_LoB.csv",
        tmp_path / "MP_ObservationYear.csv",
        tmp_path / "Risk_Adjustment.csv",
        tmp_path / "Payment_pattern.csv",
    ]
    for p in outputs:
        assert p.exists()

    # MP_ModelPoint has data rows from both source files.
    rows = _read_csv(outputs[0])
    assert rows[0][0] == "Primary_Key"
    assert len(rows) > 1

    # MP_LoB: one row per (GoC, entity). Filtered to non-zero GoCs
    # (Motor, Property and Liability all have non-zero sinistri/riserva
    # in the fixture). 3 GoCs * 2 entities = 6 rows.
    mp_lob_rows = _read_csv(outputs[1])
    assert mp_lob_rows[0] == ("GoC_ID", "Entity_ID")
    assert len(mp_lob_rows) == 1 + 3 * 2

    # MP_ObservationYear: two rows per GoC (@Opening + @Closing).
    mp_obs_rows = _read_csv(outputs[2])
    assert mp_obs_rows[0] == (
        "ObservationID", "ObservationYear", "LoB_ID", "AdjULAEPagate", "CY",
    )
    assert len(mp_obs_rows) == 1 + 3 * 2

    # Risk_Adjustment: two rows per GoC (@Opening + @Closing).
    ra_rows = _read_csv(outputs[3])
    assert ra_rows[0] == ("ObservationID", "Risk_Adjustment")
    assert len(ra_rows) == 1 + 3 * 2

    # Payment_pattern: 25 columns (GoC, Year, 0..22); 2 rows per GoC.
    pp_rows = _read_csv(outputs[4])
    assert pp_rows[0][:2] == ("GoC", "Year")
    assert pp_rows[0][2:] == tuple(range(23))
    assert len(pp_rows) == 1 + 3 * 2


def test_run_phase1_raises_without_any_ceded_or_assumed(tmp_path: Path) -> None:
    payments = tmp_path / "1.2_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    with pytest.raises(FileNotFoundError, match="_Ceded|_Assumed"):
        run_phase1(
            input_paths=[payments],
            run_dir=tmp_path,
            entities=[(6, "AAI")],
            year=2025,
            semester=2,
        )


def _build_payment_patterns_fixture_with_suffix(path: Path, suffix: str) -> None:
    """Same shape as ``_build_payment_patterns_fixture`` but the sheets are
    named ``ra_<suffix>_REINS`` and ``pp_<suffix>_REINS`` — used to prove
    that ``sheet_suffix`` is threaded end-to-end.
    """
    wb = openpyxl.Workbook()
    ra = wb.active
    ra.title = f"ra_{suffix}_REINS"
    ra.cell(row=1, column=7, value="GoC")
    for i, h in enumerate(["HY_2024", "FY_2024", "HY_2025", "FY_2025"], start=8):
        ra.cell(row=1, column=i, value=h)
    ra.cell(row=2, column=7, value="Motor")
    for j, v in enumerate([777, 778, 779, 780]):
        ra.cell(row=2, column=8 + j, value=v)
    pp = wb.create_sheet(f"pp_{suffix}_REINS")
    pp.cell(row=1, column=3, value="GoC")
    pp.cell(row=1, column=4, value="Year")
    for i in range(23):
        pp.cell(row=1, column=5 + i, value=str(i))
    for r, (goc, yr) in enumerate(
        [("Motor", "FY2025"), ("Motor", "FY2024")], start=2,
    ):
        pp.cell(row=r, column=3, value=goc)
        pp.cell(row=r, column=4, value=yr)
        for i in range(23):
            pp.cell(row=r, column=5 + i, value=9000 + i)
    wb.save(path)


def test_run_phase1_uses_sheet_suffix(tmp_path: Path) -> None:
    """Selecting a different sheet_suffix reads the RA and PP data from
    ``ra_<suffix>_REINS`` / ``pp_<suffix>_REINS`` instead of the default
    ``ra_AAI_REINS`` / ``pp_AAI_REINS`` sheets.
    """
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    ceded = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded, ["Motor"], year=2025, sinistri=100.0, riserva=50.0,
    )
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Motor"], year=2024, sinistri=80.0, riserva=40.0,
    )
    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [("GOC", "Aggregation1", "Aggregation2"), ("Motor", "A", "B")],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture_with_suffix(payments, "AMAD")

    result = run_phase1(
        input_paths=[ceded, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(14, "MPS")],
        year=2025,
        semester=2,
        sheet_suffix="AMAD",
    )

    # RA closing = FY_2025 = 780, opening = FY_2024 = 778
    ra_rows = _read_csv(result["outputs"][3])
    assert ra_rows[0] == ("ObservationID", "Risk_Adjustment")
    by_key = {r[0]: r[1] for r in ra_rows[1:]}
    assert by_key["Motor@Closing"] == 780
    assert by_key["Motor@Opening"] == 778

    # PP row for Motor / FY2025 comes from the same custom sheet
    pp_rows = _read_csv(result["outputs"][4])
    motor_2025 = next(r for r in pp_rows[1:] if r[0] == "Motor" and r[1] == 2025)
    assert motor_2025[2] == 9000  # column "0" of the AMAD sheet


def test_run_phase1_default_sheet_suffix_is_aai(tmp_path: Path) -> None:
    """Not passing sheet_suffix keeps the historical `ra_AAI_REINS` /
    `pp_AAI_REINS` behaviour — proves backward compat for existing runs.
    """
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    ceded = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded, ["Motor"], year=2025, sinistri=100.0, riserva=50.0,
    )
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Motor"], year=2024, sinistri=80.0, riserva=40.0,
    )
    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [("GOC", "Aggregation1", "Aggregation2"), ("Motor", "A", "B")],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)  # default: sheet names use AAI

    result = run_phase1(
        input_paths=[ceded, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
    )
    # Motor @Closing FY_2025 = 130 in the AAI fixture
    ra_rows = _read_csv(result["outputs"][3])
    by_key = {r[0]: r[1] for r in ra_rows[1:]}
    assert by_key["Motor@Closing"] == 130


def _build_ceded_with_pairs_fixture(
    path: Path, rows_data: list[tuple[str | None, int | None]]
) -> None:
    """Ceded fixture with GoC names in column AA and cohort years in column AB."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AAI_P&C_Ceded_H_NH"
    ws.cell(row=1, column=27, value="GoC")
    ws.cell(row=1, column=28, value="Cohort Year")
    ws.cell(row=2, column=27, value="Code")
    ws.cell(row=2, column=28, value="Year")
    for i, (goc, year) in enumerate(rows_data, start=3):
        ws.cell(row=i, column=27, value=goc)
        ws.cell(row=i, column=28, value=year)
    wb.save(path)


def _build_projection_parameters_fixture(path: Path) -> None:
    rows: list[tuple] = [
        ("PARAMETER", "VALUE"),
        ("PROJECTED_PERIODS", 110),
        ("CF_TIMESTEP", "SEMESTRIAL"),
        ("REPORTING_MONTH", "12_DECEMBER"),
        ("FX_OPENING_DATE", "1M25"),
        ("FX_AVERAGE_DATE", "HY25"),
        ("FX_CLOSING_DATE", "FY25"),
        ("FX_REPORTING_DATE", 20251231),
    ]
    _write_csv(path, rows)


def _build_mp_goc_seg_fixture(path: Path, rows: list[tuple]) -> None:
    out: list[tuple] = [("GOC_SEG_ID", "GOC_ID", "SEG_ID", "ALLOCATION_RATIO")]
    out.extend(rows)
    _write_csv(path, out)


def _build_aom_impact_fixture(path: Path, rows: list[tuple]) -> None:
    out: list[tuple] = [("GOC_ID", "STEP_ID", 1)]
    out.extend(rows)
    _write_csv(path, out)


_MP_GOC_HEADERS_PIPE = (
    "GOC_ID", "MEASUREMENT_MODEL", "ANNUAL_COHORT", "AOM_ID",
    "INCEPTION_CURVE_ID", "TIMING_INCEPTION_CURVE",
    "CSM_RELEASE_RATIO_CURVE_ID", "SHARE_TECH_EXP_ENTITY_SHARE",
    "OCI_OPTION", "OCI_OPTION_LRC", "OCI_OPTION_LIC", "GOC_DURATION",
    "GOC_TYPE_IF_NB", "GOC_CURRENCY", "REPORTING_CURRENCY",
    "GOC_TYPE_REINSURANCE", "AGGREG_1_ID", "AGGREG_2_ID", "AGGREG_3_ID",
    "AGGREG_4_ID", "AGGREG_5_ID",
)


def _build_mp_goc_fixture(path: Path, cohort_years: list[int]) -> None:
    rows: list[tuple] = [_MP_GOC_HEADERS_PIPE]
    for y in cohort_years:
        row = [f"IT{y}", "PAA", y, 20] + [None] * (len(_MP_GOC_HEADERS_PIPE) - 4)
        rows.append(tuple(row))
    _write_csv(path, rows)


def _build_curve_id_param_fixture(path: Path, rows: list[tuple]) -> None:
    out: list[tuple] = [("GOC_ID", "VARIABLE_NAME", 1)]
    out.extend(rows)
    _write_csv(path, out)


def test_run_astra_phase1_uses_pairs_from_sunrise(tmp_path: Path) -> None:
    """Astra now receives (GoC, accident_year) pairs from Sunrise instead of
    reading the legacy AAI_P&C_Ceded file."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    sunrise_pairs = [
        {"goc_id": "IT05PABPPLE2024", "goc": "IT05PABPPLE", "year": 2024},
        {"goc_id": "IT05PABPPLE2023", "goc": "IT05PABPPLE", "year": 2023},
        {"goc_id": "IT06ABCDE2024", "goc": "IT06ABCDE", "year": 2024},
    ]
    pp_params = inputs_dir / "1.4_2024.12.31_PROJECTION_PARAMETERS_ENTITY.csv"
    _build_projection_parameters_fixture(pp_params)
    mp_goc_seg = inputs_dir / "1.5_2024.12.31_MP_GOC_SEG.csv"
    _build_mp_goc_seg_fixture(
        mp_goc_seg,
        [
            ("IT05PABPPLE2024_02_P&C", "IT05PABPPLE2024", "02_P&C", 1),
            ("IT05RRIEEBB2024_02_P&C", "IT05RRIEEBB2024", "02_P&C", 1),
        ],
    )
    aom_impact = inputs_dir / "1.6_2024.12.31_ACTUARIAL_AOM_IMPACT.csv"
    _build_aom_impact_fixture(
        aom_impact,
        [("IT05PABPPLE2024", "PVFC_LIC_UNWIND", 0)],  # historical entry
    )
    curve_id_param = inputs_dir / "1.7_2024.12.31_CURVE_ID_PARAM.csv"
    _build_curve_id_param_fixture(
        curve_id_param,
        [
            ("IT05PABPPLE2024", "CLOSING_CURVE_ID", None),
            ("IT05PABPPLE2024", "OPENING_CURVE_ID", None),
            ("IT05PABPPLE2024", "CREDITED_RATE_CURVE_ID", None),
        ],
    )
    mp_goc = inputs_dir / "1.8_2024.12.31_MP_GOC.csv"
    _build_mp_goc_fixture(mp_goc, [2014, 2024])

    outputs = run_astra_phase1(
        input_paths=[pp_params, mp_goc_seg, mp_goc, aom_impact, curve_id_param],
        run_dir=tmp_path,
        year=2024,
        semester=2,
        business_type="Diretto",
        health_perimeter_gocs=["IT05RRIEEBB"],
        actuarial_aom_impact_pairs=[("DA_LIC_OP", 0), ("DA_LIC_CLO", 0)],
        closing_curve_name="pippo",
        opening_curve_name="carlo",
        goc_cohort_pairs=sunrise_pairs,
    )

    assert outputs == [
        tmp_path / "NEW_BUSINESS_PPOS.csv",
        tmp_path / "COVERAGE_UNIT.csv",
        tmp_path / "REINSURANCE.csv",
        tmp_path / "MANDATORY_ACTUALS.csv",
        tmp_path / "PROJECTION_PARAMETERS_ENTITY.csv",
        tmp_path / "MP_GOC_SEG.csv",
        tmp_path / "MP_GOC.csv",
        tmp_path / "ACTUARIAL_AOM_IMPACT.csv",
        tmp_path / "CURVE_ID_PARAM.csv",
        # OCI placeholders disabled (TODO: define population rules)
    ]
    for p in outputs:
        assert p.exists()

    # NEW_BUSINESS_PPOS: one row per pair (3 pairs in fixture)
    nb_rows = _read_astra_csv(outputs[0])
    assert nb_rows == [
        ("GOC_ID", "VARIABLE_NAME", 1),
        ("IT05PABPPLE2024", "CROSS_SUB_FASSCHNG", 0),
        ("IT05PABPPLE2023", "CROSS_SUB_FASSCHNG", 0),
        ("IT06ABCDE2024", "CROSS_SUB_FASSCHNG", 0),
    ]

    cu_rows = _read_astra_csv(outputs[1])
    assert len(cu_rows) == 1 + 3
    assert cu_rows[1] == ("IT05PABPPLE2024", 1) + (0,) * 100
    assert cu_rows[3] == ("IT06ABCDE2024", 1) + (0,) * 100

    rein_rows = _read_astra_csv(outputs[2])
    assert rein_rows == [
        ("GOC_ID", "VARIABLE_NAME", 1, "T"),
        ("IT05PABPPLE2024", "LOSSRECO_IFE_ALLOCATION", 0, 2024),
        ("IT05PABPPLE2023", "LOSSRECO_IFE_ALLOCATION", 0, 2023),
        ("IT05PABPPLE2024", "LOSSRECO_CLOSING", 0, 2024),
        ("IT05PABPPLE2023", "LOSSRECO_CLOSING", 0, 2023),
        ("IT06ABCDE2024", "LOSSRECO_IFE_ALLOCATION", 0, 2024),
        ("IT06ABCDE2024", "LOSSRECO_CLOSING", 0, 2024),
    ]

    ma_rows = _read_astra_csv(outputs[3])
    assert len(ma_rows) == 1 + 3 * 16
    assert ma_rows[1] == ("IT05PABPPLE2024", "ACTUAL_PREMIUM_CF_PAST_SERVICE", 0)
    assert ma_rows[17] == ("IT05PABPPLE2023", "ACTUAL_PREMIUM_CF_PAST_SERVICE", 0)
    assert ma_rows[33] == ("IT06ABCDE2024", "ACTUAL_PREMIUM_CF_PAST_SERVICE", 0)

    pp_rows = _read_astra_csv(outputs[4])
    by_param = {row[0]: row[1] for row in pp_rows[1:]}
    assert by_param["CF_TIMESTEP"] == "SEMESTRIAL"
    assert by_param["REPORTING_MONTH"] == "12_DECEMBER"
    assert by_param["FX_OPENING_DATE"] == "1M24"
    assert by_param["FX_AVERAGE_DATE"] == "HY24"
    assert by_param["FX_CLOSING_DATE"] == "FY24"
    assert by_param["FX_REPORTING_DATE"] == 20241231
    assert by_param["PROJECTED_PERIODS"] == 110

    seg_rows = _read_astra_csv(outputs[5])
    assert seg_rows[1] == ("IT05PABPPLE2024_02_P&C", "IT05PABPPLE2024", "02_P&C", 1)
    assert seg_rows[2] == (
        "IT05RRIEEBB2024_02_HLTH_PC", "IT05RRIEEBB2024", "02_HLTH_PC", 1,
    )

    mp_goc_rows = _read_astra_csv(outputs[6])
    by_cohort = {row[2]: row for row in mp_goc_rows[1:]}
    # E (idx 4), F (idx 5), L (idx 11), P (idx 15)
    # cohort <= 2015 -> col E is the fixed "20211231_ITA_LP100"
    assert by_cohort[2014][4] == "20211231_ITA_LP100"
    assert by_cohort[2024][4] == "20241231_EUR_LP100_FY24_AVG"
    # L uses curve_year = first 4 chars of col E:
    # cohort 2014 -> col E starts "2021" -> max(0, 2024-1-2021)*12 = 24
    assert by_cohort[2014][11] == max(0, 2024 - 1 - 2021) * 12
    assert by_cohort[2024][11] == 0
    for cohort in (2014, 2024):
        assert by_cohort[cohort][15] == "2_RE_ASSUMED"
    # T (idx 19) and U (idx 20) carry the GoC name from col A
    # stripped of its 4-char cohort year suffix. Fixture uses
    # GOC_ID = ``IT{year}`` so stripping leaves ``IT``.
    for cohort in (2014, 2024):
        assert by_cohort[cohort][19] == "IT"
        assert by_cohort[cohort][20] == "IT"

    aom_rows = _read_astra_csv(outputs[7])
    assert aom_rows[0] == ("GOC_ID", "STEP_ID", 1)
    assert aom_rows[1:] == [
        ("IT05PABPPLE2023", "DA_LIC_CLO", 0),
        ("IT05PABPPLE2023", "DA_LIC_OP", 0),
        ("IT05PABPPLE2024", "DA_LIC_CLO", 0),
        ("IT05PABPPLE2024", "DA_LIC_OP", 0),
        ("IT05PABPPLE2024", "PVFC_LIC_UNWIND", 0),
        ("IT06ABCDE2024", "DA_LIC_CLO", 0),
        ("IT06ABCDE2024", "DA_LIC_OP", 0),
    ]

    curve_rows = _read_astra_csv(outputs[8])
    assert curve_rows == [
        ("GOC_ID", "VARIABLE_NAME", 1),
        ("IT05PABPPLE2024", "CLOSING_CURVE_ID", "pippo"),
        ("IT05PABPPLE2024", "OPENING_CURVE_ID", "carlo"),
        ("IT05PABPPLE2024", "CREDITED_RATE_CURVE_ID", "CR_IT05PABPPLE2024"),
    ]


def test_run_astra_phase1_filters_pairs_outside_window(tmp_path: Path) -> None:
    """Pairs outside [year-15, year] are dropped before generation."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    sunrise_pairs = [
        {"goc_id": "IT05PABPPLE2024", "goc": "IT05PABPPLE", "year": 2024},  # kept
        {"goc_id": "IT05PABPPLE2009", "goc": "IT05PABPPLE", "year": 2009},  # kept
        {"goc_id": "IT05PABPPLE2008", "goc": "IT05PABPPLE", "year": 2008},  # dropped
        {"goc_id": "IT05PABPPLE2030", "goc": "IT05PABPPLE", "year": 2030},  # dropped
        {"goc_id": "IT06ABCDE2020", "goc": "IT06ABCDE", "year": 2020},      # kept
    ]
    pp_params = inputs_dir / "1.4_PROJECTION_PARAMETERS_ENTITY.csv"
    _build_projection_parameters_fixture(pp_params)
    mp_goc_seg = inputs_dir / "1.5_MP_GOC_SEG.csv"
    _build_mp_goc_seg_fixture(mp_goc_seg, [])
    aom_impact = inputs_dir / "1.6_ACTUARIAL_AOM_IMPACT.csv"
    _build_aom_impact_fixture(aom_impact, [])
    curve_id_param = inputs_dir / "1.7_CURVE_ID_PARAM.csv"
    _build_curve_id_param_fixture(curve_id_param, [])
    mp_goc = inputs_dir / "1.8_MP_GOC.csv"
    _build_mp_goc_fixture(mp_goc, [])

    outputs = run_astra_phase1(
        input_paths=[pp_params, mp_goc_seg, mp_goc, aom_impact, curve_id_param],
        run_dir=tmp_path,
        year=2024,
        semester=2,
        business_type="Diretto",
        health_perimeter_gocs=[],
        actuarial_aom_impact_pairs=[],
        closing_curve_name="",
        opening_curve_name="",
        goc_cohort_pairs=sunrise_pairs,
    )

    # Only 3 of the 5 pairs survive the filter
    nb_rows = _read_astra_csv(outputs[0])
    assert nb_rows == [
        ("GOC_ID", "VARIABLE_NAME", 1),
        ("IT05PABPPLE2024", "CROSS_SUB_FASSCHNG", 0),
        ("IT05PABPPLE2009", "CROSS_SUB_FASSCHNG", 0),
        ("IT06ABCDE2020", "CROSS_SUB_FASSCHNG", 0),
    ]


def test_run_astra_phase1_missing_required_input(tmp_path: Path) -> None:
    """PROJECTION_PARAMETERS_ENTITY (still mandatory) is missing -> error."""
    other = tmp_path / "irrelevant.xlsx"
    openpyxl.Workbook().save(other)

    with pytest.raises(FileNotFoundError, match="PROJECTION_PARAMETERS_ENTITY"):
        run_astra_phase1(
            input_paths=[other],
            run_dir=tmp_path,
            year=2024,
            semester=2,
            business_type="Diretto",
            health_perimeter_gocs=[],
            actuarial_aom_impact_pairs=[],
            closing_curve_name="",
            opening_curve_name="",
            goc_cohort_pairs=[],
        )


def test_run_phase1_excludes_gocs(tmp_path: Path) -> None:
    """``gocs_to_exclude`` removes every cohort year for the listed GoCs
    from MP_ModelPoint, MP_LoB, MP_ObservationYear, and the (GoC, year)
    pairs passed downstream to Astra."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["Motor", "Property"], year=2025, sinistri=100.0, riserva=50.0,
    )
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Property", "Liability"], year=2024, sinistri=80.0, riserva=40.0,
    )
    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [
            ("GOC", "Aggregation1", "Aggregation2"),
            ("Motor", "Agg1_Motor", "Agg2_Motor"),
            ("Property", "Agg1_Prop", "Agg2_Prop"),
        ],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
        gocs_to_exclude=["Property"],
    )

    # GoC list passed to MP_LoB no longer contains the excluded GoC
    mp_lob_rows = _read_csv(result["outputs"][1])
    gocs_in_lob = {row[0] for row in mp_lob_rows[1:]}
    assert "Property" not in gocs_in_lob
    assert gocs_in_lob == {"Motor", "Liability"}

    # (GoC, year) pairs shared with Astra also drop every cohort of Property
    cohort_gocs = {p["goc"] for p in result["goc_cohort_pairs"]}
    assert "Property" not in cohort_gocs
    assert cohort_gocs == {"Motor", "Liability"}


def test_run_phase1_renames_single_cohort(tmp_path: Path) -> None:
    """``goc_renames`` operates on a specific ``(GoC, cohort)`` pair:
    renaming ``Motor2025`` only moves the 2025 cohort of ``Motor`` to
    the new name; the 2024 cohort of ``Motor`` stays put."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["Motor", "Property"], year=2025, sinistri=100.0, riserva=50.0,
    )
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Property", "Motor"], year=2024, sinistri=80.0, riserva=40.0,
    )
    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [
            ("GOC", "Aggregation1", "Aggregation2"),
            ("Motor", "Agg1_Motor", "Agg2_Motor"),
            ("Property", "Agg1_Prop", "Agg2_Prop"),
        ],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
        goc_renames={"Motor2025": "Auto2025"},
    )
    assert result["warnings"] == []

    # MP_LoB shows both GoCs: Motor still exists (its 2024 cohort survived),
    # Auto exists as the new home of the 2025 cohort of what was Motor.
    mp_lob_rows = _read_csv(result["outputs"][1])
    gocs_in_lob = {row[0] for row in mp_lob_rows[1:]}
    assert "Motor" in gocs_in_lob
    assert "Auto" in gocs_in_lob

    # Cohort pairs handed to Astra reflect the split:
    # - Motor keeps its real 2024 row
    # - Auto owns the real 2025 row (moved from Motor2025)
    # Both GoCs also get zero auto-fill rows for the analysis year and the
    # previous year respectively; the important check is that the real
    # values are on the right sides.
    pairs = {(p["goc"], p["year"]) for p in result["goc_cohort_pairs"]}
    assert ("Motor", 2024) in pairs
    assert ("Auto", 2025) in pairs

    # MP_ModelPoint tells us which side actually carries the real numbers.
    mp_rows = _read_csv(result["outputs"][0])
    motor_2024 = [r for r in mp_rows[1:] if r[1] == "Motor2024" and r[2] == "Motor@Opening"]
    assert len(motor_2024) == 1 and motor_2024[0][9] == -80.0  # real Motor 2024 sinistri
    auto_2025 = [r for r in mp_rows[1:] if r[1] == "Auto2025" and r[2] == "Auto@Closing"]
    assert len(auto_2025) == 1 and auto_2025[0][9] == -100.0  # renamed from Motor2025


def test_run_phase1_renames_all_cohorts(tmp_path: Path) -> None:
    """Renaming both cohorts of a GoC yields the equivalent of the
    old per-GoC rename semantics — the original GoC name disappears."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["Motor"], year=2025, sinistri=100.0, riserva=50.0,
    )
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Motor"], year=2024, sinistri=80.0, riserva=40.0,
    )
    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [("GOC", "Aggregation1", "Aggregation2"), ("Motor", "A", "B")],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
        goc_renames={"Motor2024": "Auto2024", "Motor2025": "Auto2025"},
    )
    assert result["warnings"] == []

    mp_lob_rows = _read_csv(result["outputs"][1])
    gocs_in_lob = {row[0] for row in mp_lob_rows[1:]}
    assert gocs_in_lob == {"Auto"}


def test_run_phase1_rename_collision_raises(tmp_path: Path) -> None:
    """If the target GoC+cohort already exists in the input, the run
    stops with a clear ValueError before writing any output."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["Motor", "Property"], year=2025, sinistri=100.0, riserva=50.0,
    )
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Property", "Motor"], year=2024, sinistri=80.0, riserva=40.0,
    )
    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [("GOC", "Aggregation1", "Aggregation2"), ("Motor", "A", "B"),
         ("Property", "C", "D")],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    with pytest.raises(ValueError, match="already present"):
        run_phase1(
            input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
            run_dir=tmp_path,
            entities=[(6, "AAI")],
            year=2025,
            semester=2,
            goc_renames={"Motor2025": "Property2025"},
        )


def test_run_phase1_rename_invalid_format_warning(tmp_path: Path) -> None:
    """A rename entry that doesn't parse as `<goc><YYYY>` is skipped
    with a warning; the pipeline still completes."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["Motor"], year=2025, sinistri=100.0, riserva=50.0,
    )
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Motor"], year=2024, sinistri=80.0, riserva=40.0,
    )
    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [("GOC", "Aggregation1", "Aggregation2"), ("Motor", "A", "B")],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
        # No trailing year -> unparseable -> skipped with a warning.
        goc_renames={"Motor": "Auto"},
    )
    assert any("invalid format" in w for w in result["warnings"])
    # Motor rows untouched: rename was silently skipped.
    mp_lob_rows = _read_csv(result["outputs"][1])
    gocs_in_lob = {row[0] for row in mp_lob_rows[1:]}
    assert gocs_in_lob == {"Motor"}


def test_run_phase1_rename_then_exclude(tmp_path: Path) -> None:
    """When the same name is both renamed and excluded, rename wins
    (acts on input names) and the post-rename name is what the
    exclusion list compares against."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["Motor", "Property"], year=2025, sinistri=100.0, riserva=50.0,
    )
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Motor"], year=2024, sinistri=80.0, riserva=40.0,
    )
    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [
            ("GOC", "Aggregation1", "Aggregation2"),
            ("Motor", "Agg1_Motor", "Agg2_Motor"),
            ("Property", "Agg1_Prop", "Agg2_Prop"),
        ],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    # Rename both cohorts of Motor into Auto, then exclude Auto entirely.
    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
        goc_renames={"Motor2024": "Auto2024", "Motor2025": "Auto2025"},
        gocs_to_exclude=["Auto"],  # post-rename name (11-char GoC)
    )

    # Auto (formerly Motor's two cohorts) is excluded; only Property survives.
    mp_lob_rows = _read_csv(result["outputs"][1])
    gocs_in_lob = {row[0] for row in mp_lob_rows[1:]}
    assert gocs_in_lob == {"Property"}


def test_run_phase1_autofills_analysis_year(tmp_path: Path) -> None:
    """A non-zero GoC missing the analysis-year row gets a synthetic
    zero row so MP_ModelPoint @Closing and the cohort pairs always
    include the analysis year."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()

    # Current-year (2025) file: only ``Motor`` has a 2025 row; we
    # deliberately leave ``Liability`` out of the 2025 file so the
    # auto-fill kicks in for it.
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["Motor"], year=2025, sinistri=100.0, riserva=50.0,
    )

    # Previous-year file carries ``Liability`` (non-zero in 2024 only).
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Motor", "Liability"],
        year=2024, sinistri=80.0, riserva=40.0,
    )

    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [
            ("GOC", "Aggregation1", "Aggregation2"),
            ("Motor", "Agg1_Motor", "Agg2_Motor"),
            ("Liability", "Agg1_Liab", "Agg2_Liab"),
        ],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
    )

    # Cohort pairs include (Liability, 2025) even though Liability had
    # no 2025 row in the 2025 file.
    pairs_set = {(p["goc"], p["year"]) for p in result["goc_cohort_pairs"]}
    assert ("Liability", 2025) in pairs_set
    assert ("Liability", 2024) in pairs_set

    # MP_ModelPoint has a (Liability, 2025) @Closing row with zeros.
    mp_rows = _read_csv(result["outputs"][0])
    liab_2025 = [
        r for r in mp_rows[1:]
        if r[1] == "Liability2025" and r[2] == "Liability@Closing"
    ]
    assert len(liab_2025) == 1
    # EAXA_Reserve (idx 7) = riserva = 0, Claims_Paid (idx 9) = sinistri = 0
    assert liab_2025[0][7] == 0
    assert liab_2025[0][9] == 0


def test_run_phase1_fills_cohort_year_gaps(tmp_path: Path) -> None:
    """Missing accident years between the oldest and the newest cohort
    inside a single (ANNO_RIFERIMENTO, GoC) group are gap-filled with
    zero rows. Mirrors the user's IT05RRIHVIC case (data at 2021 and
    2025 only, expect 2022/2023/2024 to appear with zeros)."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()

    # Current-year file: GoC X with cohort data at 2021 and 2025 only
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input_Sunrise"
    ws.cell(row=1, column=1, value="GoC")
    ws.cell(row=1, column=2, value="Year")
    ws.cell(row=1, column=4, value="Sinistri")
    ws.cell(row=1, column=5, value="Riserva")
    ws.cell(row=2, column=1, value="X")
    ws.cell(row=2, column=2, value=2021)
    ws.cell(row=2, column=4, value=100.0)
    ws.cell(row=2, column=5, value=50.0)
    ws.cell(row=3, column=1, value="X")
    ws.cell(row=3, column=2, value=2025)
    ws.cell(row=3, column=4, value=200.0)
    ws.cell(row=3, column=5, value=80.0)
    wb.save(ceded_curr)

    # Previous-year file: X with 2024 data (single row, no gaps to fill).
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["X"], year=2024, sinistri=80.0, riserva=40.0,
    )

    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [("GOC", "Aggregation1", "Aggregation2"), ("X", "A", "B")],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
    )

    # Cohort pairs include every year between 2021 and 2025.
    x_pairs = {p["year"] for p in result["goc_cohort_pairs"] if p["goc"] == "X"}
    assert {2021, 2022, 2023, 2024, 2025}.issubset(x_pairs)

    # MP_ModelPoint @Closing has exactly the contiguous 2021..2025 range
    # for X (the 2024 file produces a single @Opening row that's
    # separate from this assertion).
    mp_rows = _read_csv(result["outputs"][0])
    x_closing_years = sorted([
        r[3] for r in mp_rows[1:]
        if r[2] == "X@Closing"
    ])
    assert x_closing_years == [2021, 2022, 2023, 2024, 2025]

    # Gap-filled years (2022/2023/2024) carry zeros; real years keep
    # their negated values.
    by_year = {
        r[3]: r for r in mp_rows[1:]
        if r[2] == "X@Closing"
    }
    for y in (2022, 2023, 2024):
        assert by_year[y][7] == 0
        assert by_year[y][9] == 0
    assert by_year[2021][9] == -100.0
    assert by_year[2021][7] == -50.0
    assert by_year[2025][9] == -200.0
    assert by_year[2025][7] == -80.0


def test_run_phase1_pre_horizon_fold_lands_on_year_minus_14(
    tmp_path: Path,
) -> None:
    """Fold rule differs between the two groups:

    - ``@Closing`` (``anno_rif = year``): 16-year horizon. Fold at
      ``year - 14`` (2011 for YE25). When the fold triggers, a zero
      padding row at ``year - 15`` (2010) is also emitted.
    - ``@Opening`` (``anno_rif = year - 1``): 15-year horizon. Fold at
      ``anno_rif - 14`` = ``year - 15`` (2010 for YE25). NO padding
      row at ``anno_rif - 15`` (2009), even if the fold triggers.
    """
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()

    # Current-year (2025) file: pre-horizon at 2008 (100/50) and 2009 (50/25),
    # plus an in-horizon value at 2020 (200/80).
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input_Sunrise"
    ws.cell(row=1, column=1, value="GoC")
    ws.cell(row=1, column=2, value="Year")
    ws.cell(row=1, column=4, value="Sinistri")
    ws.cell(row=1, column=5, value="Riserva")
    for i, (y, s, r) in enumerate(
        [(2008, 100.0, 50.0), (2009, 50.0, 25.0), (2020, 200.0, 80.0)], start=2,
    ):
        ws.cell(row=i, column=1, value="X")
        ws.cell(row=i, column=2, value=y)
        ws.cell(row=i, column=4, value=s)
        ws.cell(row=i, column=5, value=r)
    wb.save(ceded_curr)

    # Previous-year (2024) file: pre-horizon at 2005 (30/15) plus in-horizon
    # at 2020 (10/5). This exercises the @Opening fold path too.
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input_Sunrise"
    ws.cell(row=1, column=1, value="GoC")
    ws.cell(row=1, column=2, value="Year")
    ws.cell(row=1, column=4, value="Sinistri")
    ws.cell(row=1, column=5, value="Riserva")
    for i, (y, s, r) in enumerate(
        [(2005, 30.0, 15.0), (2020, 10.0, 5.0)], start=2,
    ):
        ws.cell(row=i, column=1, value="X")
        ws.cell(row=i, column=2, value=y)
        ws.cell(row=i, column=4, value=s)
        ws.cell(row=i, column=5, value=r)
    wb.save(ceded_prev)

    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [("GOC", "Aggregation1", "Aggregation2"), ("X", "A", "B")],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
    )

    mp_rows = _read_csv(result["outputs"][0])
    closing = {r[3]: r for r in mp_rows[1:] if r[2] == "X@Closing"}
    opening = {r[3]: r for r in mp_rows[1:] if r[2] == "X@Opening"}

    # @Closing (anno_rif = 2025): fold_year = 2011, padding at 2010.
    # 16-year horizon 2010..2025.
    assert sorted(closing.keys()) == list(range(2010, 2026))
    # Pre-horizon sum (2008 + 2009) → 150/75 into 2011 (negated in output).
    assert closing[2011][9] == -150.0
    assert closing[2011][7] == -75.0
    # 2010 padding row is emitted with zeros.
    assert closing[2010][9] == 0
    assert closing[2010][7] == 0
    # In-horizon real value at 2020 keeps its (negated) value.
    assert closing[2020][9] == -200.0
    assert closing[2020][7] == -80.0

    # @Opening (anno_rif = 2024): fold_year = 2010. 15-year horizon
    # 2010..2024 — the row at anno_rif - 15 = 2009 is NEVER emitted,
    # even though the fold triggers on this side too.
    assert sorted(opening.keys()) == list(range(2010, 2025))
    assert 2009 not in opening
    # Pre-horizon (2005) folded into 2010 → 30/15 (negated in output).
    assert opening[2010][9] == -30.0
    assert opening[2010][7] == -15.0
    # In-horizon real value at 2020 keeps its (negated) value.
    assert opening[2020][9] == -10.0
    assert opening[2020][7] == -5.0


def test_run_phase1_closing_padding_when_only_prev_reaches_year_minus_15(
    tmp_path: Path,
) -> None:
    """@Closing must include a zero padding row at year - 15 whenever
    the GoC has data at or below year - 15 anywhere in the master —
    even if the current-year file itself has nothing pre-horizon.
    Mirrors the user's IT05RRIMTPL case: curr starts at 2020, prev has
    cohorts 2010..2023 → both @Closing and @Opening must start at 2010.
    """
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()

    # curr (2025): X at 2020 only — no accident year below year - 14.
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["X"], year=2020, sinistri=100.0, riserva=50.0,
    )

    # prev (2024): X has cohorts 2010..2023 all non-zero.
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input_Sunrise"
    ws.cell(row=1, column=1, value="GoC")
    ws.cell(row=1, column=2, value="Year")
    ws.cell(row=1, column=4, value="Sinistri")
    ws.cell(row=1, column=5, value="Riserva")
    for i, y in enumerate(range(2010, 2024), start=2):
        ws.cell(row=i, column=1, value="X")
        ws.cell(row=i, column=2, value=y)
        ws.cell(row=i, column=4, value=10.0)
        ws.cell(row=i, column=5, value=5.0)
    wb.save(ceded_prev)

    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [("GOC", "Aggregation1", "Aggregation2"), ("X", "A", "B")],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
    )

    mp_rows = _read_csv(result["outputs"][0])
    closing = {r[3]: r for r in mp_rows[1:] if r[2] == "X@Closing"}
    opening = {r[3]: r for r in mp_rows[1:] if r[2] == "X@Opening"}

    # Both groups start at 2010; @Closing has 16 years, @Opening has 15.
    assert sorted(closing.keys()) == list(range(2010, 2026))
    assert sorted(opening.keys()) == list(range(2010, 2025))

    # @Closing[2010] is a zero padding row (no real 2010 data in curr,
    # no @Closing pre-horizon fold to place there).
    assert closing[2010][7] == 0
    assert closing[2010][9] == 0
    # @Closing[2020] keeps the real (negated) value.
    assert closing[2020][7] == -50.0
    assert closing[2020][9] == -100.0
    # @Opening[2010] carries the real (negated) 2010 value from prev.
    assert opening[2010][7] == -5.0
    assert opening[2010][9] == -10.0


def test_run_phase1_aligns_min_cohort_year_across_groups(
    tmp_path: Path,
) -> None:
    """The oldest cohort year is aligned between @Closing (year file)
    and @Opening (year-1 file): a cohort present only in the year file
    gets a synthetic zero row in the year-1 file too. Mirrors the
    user's IT05RRIPOOL case (2018 present @Closing but absent from
    @Opening)."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()

    # Current-year (2025) file: GoC X has cohorts 2018..2025.
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input_Sunrise"
    ws.cell(row=1, column=1, value="GoC")
    ws.cell(row=1, column=2, value="Year")
    ws.cell(row=1, column=4, value="Sinistri")
    ws.cell(row=1, column=5, value="Riserva")
    for i, y in enumerate(range(2018, 2026), start=2):
        ws.cell(row=i, column=1, value="X")
        ws.cell(row=i, column=2, value=y)
        ws.cell(row=i, column=4, value=10.0)
        ws.cell(row=i, column=5, value=5.0)
    wb.save(ceded_curr)

    # Previous-year (2024) file: X has cohorts 2019..2024 — 2018 is
    # missing on this side.
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input_Sunrise"
    ws.cell(row=1, column=1, value="GoC")
    ws.cell(row=1, column=2, value="Year")
    ws.cell(row=1, column=4, value="Sinistri")
    ws.cell(row=1, column=5, value="Riserva")
    for i, y in enumerate(range(2019, 2025), start=2):
        ws.cell(row=i, column=1, value="X")
        ws.cell(row=i, column=2, value=y)
        ws.cell(row=i, column=4, value=8.0)
        ws.cell(row=i, column=5, value=4.0)
    wb.save(ceded_prev)

    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [("GOC", "Aggregation1", "Aggregation2"), ("X", "A", "B")],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
    )

    mp_rows = _read_csv(result["outputs"][0])
    closing_years = sorted([
        r[3] for r in mp_rows[1:] if r[2] == "X@Closing"
    ])
    opening_years = sorted([
        r[3] for r in mp_rows[1:] if r[2] == "X@Opening"
    ])

    # @Closing: 2018..2025 (real). @Opening: aligned floor at 2018, up to 2024.
    assert closing_years == list(range(2018, 2026))
    assert opening_years == list(range(2018, 2025))

    # The synthetic 2018 @Opening row has zero values.
    row_2018_opening = next(
        r for r in mp_rows[1:]
        if r[2] == "X@Opening" and r[3] == 2018
    )
    assert row_2018_opening[7] == 0
    assert row_2018_opening[9] == 0


def test_run_phase1_autofills_previous_year(tmp_path: Path) -> None:
    """The auto-fill applies to the previous-year file too: a non-zero
    GoC that is present in the current-year file but absent from the
    previous-year file still gets a synthetic ``@Opening`` row with
    Accident_Year=year-1 and zero values."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    # Current-year file: both Motor and Liability with 2025 data.
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["Motor", "Liability"],
        year=2025, sinistri=100.0, riserva=50.0,
    )
    # Previous-year file: only Motor. Liability is absent.
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Motor"], year=2024, sinistri=80.0, riserva=40.0,
    )
    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [
            ("GOC", "Aggregation1", "Aggregation2"),
            ("Motor", "A", "B"),
            ("Liability", "C", "D"),
        ],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
    )

    # Cohort pairs include (Liability, 2024) even though Liability was
    # missing from the 2024 file.
    pairs_set = {(p["goc"], p["year"]) for p in result["goc_cohort_pairs"]}
    assert ("Liability", 2024) in pairs_set
    assert ("Liability", 2025) in pairs_set

    # MP_ModelPoint @Opening row for Liability with Accident_Year=2024
    # and zero claim values.
    mp_rows = _read_csv(result["outputs"][0])
    liab_opening = [
        r for r in mp_rows[1:]
        if r[1] == "Liability2024" and r[2] == "Liability@Opening"
    ]
    assert len(liab_opening) == 1
    # Sign-flipped, but zero stays zero
    assert liab_opening[0][7] == 0  # EAXA_Reserve
    assert liab_opening[0][9] == 0  # Claims_Paid


def test_run_phase1_no_autofill_for_all_zero_gocs(tmp_path: Path) -> None:
    """The all-zero exclusion rule still wins: a GoC that has only
    zero rows everywhere is dropped and the analysis-year row is NOT
    auto-filled for it."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["Motor", "Empty"], year=2025,
        sinistri=0.0, riserva=0.0,  # both zero in 2025 file
    )
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    # ``Motor`` is non-zero in 2024 so it survives; ``Empty`` is zero
    # in 2024 too.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input_Sunrise"
    ws.cell(row=1, column=1, value="GoC")
    ws.cell(row=1, column=2, value="Year")
    ws.cell(row=1, column=4, value="Sinistri")
    ws.cell(row=1, column=5, value="Riserva")
    ws.cell(row=2, column=1, value="Motor")
    ws.cell(row=2, column=2, value=2024)
    ws.cell(row=2, column=4, value=80.0)
    ws.cell(row=2, column=5, value=40.0)
    ws.cell(row=3, column=1, value="Empty")
    ws.cell(row=3, column=2, value=2024)
    ws.cell(row=3, column=4, value=0.0)
    ws.cell(row=3, column=5, value=0.0)
    wb.save(ceded_prev)

    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [("GOC", "Aggregation1", "Aggregation2"), ("Motor", "A", "B")],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
    )

    cohort_gocs = {p["goc"] for p in result["goc_cohort_pairs"]}
    # ``Empty`` is excluded entirely by the all-zero filter — no auto-fill.
    assert "Empty" not in cohort_gocs
    # ``Motor`` survives and has an auto-filled (Motor, 2025).
    assert ("Motor", 2025) in {(p["goc"], p["year"]) for p in result["goc_cohort_pairs"]}


def test_run_phase1_rename_unknown_old_warning(tmp_path: Path) -> None:
    """Rename sources not in the input list surface as a warning but
    the run still completes."""
    inputs_dir = tmp_path / "inputs"
    inputs_dir.mkdir()
    ceded_curr = inputs_dir / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_curr, ["Motor"], year=2025, sinistri=100.0, riserva=50.0,
    )
    ceded_prev = inputs_dir / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_input_sunrise_workbook(
        ceded_prev, ["Motor"], year=2024, sinistri=80.0, riserva=40.0,
    )
    transcodifica = inputs_dir / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _write_csv(
        transcodifica,
        [("GOC", "Aggregation1", "Aggregation2"), ("Motor", "Ag1", "Ag2")],
    )
    payments = inputs_dir / "1.4_2025.12.31_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns_fixture(payments)

    result = run_phase1(
        input_paths=[ceded_curr, ceded_prev, transcodifica, payments],
        run_dir=tmp_path,
        entities=[(6, "AAI")],
        year=2025,
        semester=2,
        # Well-formed rename entry that doesn't match anything in the input.
        goc_renames={"Ghost1900": "Phantom1900"},
    )

    assert result["warnings"] != []
    assert any(
        "Ghost1900" in w and "not in the input list" in w
        for w in result["warnings"]
    )
    # Real GoCs untouched.
    mp_lob_rows = _read_csv(result["outputs"][1])
    gocs_in_lob = {row[0] for row in mp_lob_rows[1:]}
    assert gocs_in_lob == {"Motor"}

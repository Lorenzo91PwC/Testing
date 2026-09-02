"""Tests for excel_pipeline.input_validation."""
from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl

from excel_pipeline.input_validation import (
    ValidationIssue,
    ValidationReport,
    validate_actuarial_aom_impact_file,
    validate_astra_inputs,
    validate_ceded_assumed_file,
    validate_curve_id_param_file,
    validate_mp_goc_file,
    validate_mp_goc_seg_file,
    validate_payment_patterns_file,
    validate_projection_parameters_entity_file,
    validate_sunrise_inputs,
    validate_transcodifica_file,
)


# ---------------------------------------------------------------------------
# Shared fixture builders
# ---------------------------------------------------------------------------
def _write_csv(path: Path, rows: list[tuple]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        for row in rows:
            writer.writerow(["" if v is None else v for v in row])


def _build_ceded_workbook(
    path: Path,
    rows: list[tuple],
    *,
    sheet_name: str = "Input_Sunrise",
    headers: tuple = ("GOC", "ANNO", "PERIMETRO", "SINISTRI", "RISERVA_SINISTRI"),
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)


def _build_transcodifica_csv(path: Path, rows: list[tuple]) -> None:
    out: list[tuple] = [("GOC_ID", "Aggregation1", "Aggregation2", "H-NH")]
    out.extend(rows)
    _write_csv(path, out)


def _build_payment_patterns(
    path: Path,
    *,
    include_ra: bool = True,
    include_pp: bool = True,
    ra_year_headers: list[str] | None = None,
    pp_with_23_cols: bool = True,
) -> None:
    wb = openpyxl.Workbook()
    default = wb.active
    if include_ra:
        default.title = "ra_AAI_REINS"
        default.cell(row=1, column=7, value="GoC")
        headers = ra_year_headers or ["HY_2024", "FY_2024", "HY_2025", "FY_2025"]
        for i, h in enumerate(headers, start=8):
            default.cell(row=1, column=i, value=h)
    else:
        default.title = "OTHER"
    if include_pp:
        pp = wb.create_sheet("pp_AAI_REINS")
        pp.cell(row=1, column=3, value="GoC")
        pp.cell(row=1, column=4, value="Year")
        if pp_with_23_cols:
            for i in range(23):
                pp.cell(row=1, column=5 + i, value=str(i))
    wb.save(path)


# ---------------------------------------------------------------------------
# ValidationReport
# ---------------------------------------------------------------------------
def test_validation_report_blocking_and_persistence(tmp_path: Path) -> None:
    report = ValidationReport()
    report.add(ValidationIssue(
        file="a.xlsx", severity="warning", location="data",
        code="SOFT", message="just a warning",
    ))
    assert not report.is_blocking
    report.add(ValidationIssue(
        file="b.xlsx", severity="error", location="header",
        code="HARD", message="something is broken",
    ))
    assert report.is_blocking
    assert len(report.errors) == 1
    assert len(report.warnings) == 1

    out = tmp_path / "validation.json"
    report.save(out)
    data = json.loads(out.read_text(encoding="utf-8"))
    assert {i["code"] for i in data["errors"]} == {"HARD"}
    assert {i["code"] for i in data["warnings"]} == {"SOFT"}


# ---------------------------------------------------------------------------
# Per-file validators — Sunrise
# ---------------------------------------------------------------------------
def test_validate_ceded_assumed_ok(tmp_path: Path) -> None:
    path = tmp_path / "2025.12.31_AAI_Ceded.xlsx"
    _build_ceded_workbook(
        path, [("IT05PABPPLE", 2024, "Direct", 100.0, 50.0)],
    )
    issues = validate_ceded_assumed_file(path)
    assert issues == []


def test_validate_ceded_assumed_missing_sheet(tmp_path: Path) -> None:
    path = tmp_path / "2025.12.31_AAI_Ceded.xlsx"
    _build_ceded_workbook(
        path, [], sheet_name="WrongSheet",
    )
    issues = validate_ceded_assumed_file(path)
    assert any(i.code == "MISSING_SHEET" for i in issues)


def test_validate_ceded_assumed_invalid_anno(tmp_path: Path) -> None:
    path = tmp_path / "2025.12.31_AAI_Ceded.xlsx"
    _build_ceded_workbook(
        path, [("IT05PABPPLE", "NOT_A_YEAR", "Direct", 100.0, 50.0)],
    )
    issues = validate_ceded_assumed_file(path)
    assert any(i.code == "INVALID_ANNO" for i in issues)


def test_validate_ceded_assumed_all_zero_warning_lists_gocs(tmp_path: Path) -> None:
    """A GoC whose SINISTRI + RISERVA are zero on every row surfaces as a
    warning that names the GoC. Multiple all-zero GoCs are listed in the
    order they first appear in the file.
    """
    path = tmp_path / "2025.12.31_AAI_Ceded.xlsx"
    _build_ceded_workbook(
        path, [
            ("IT05PABPPLE", 2024, "Direct", 0.0, 0.0),
            ("IT06ABCDE", 2024, "Direct", 0, 0),
        ],
    )
    issues = validate_ceded_assumed_file(path)
    warnings = [
        i for i in issues
        if i.code == "ALL_ZERO_VALUES" and i.severity == "warning"
    ]
    assert len(warnings) == 1
    msg = warnings[0].message
    assert "IT05PABPPLE" in msg
    assert "IT06ABCDE" in msg


def test_validate_ceded_assumed_all_zero_warning_only_lists_zero_gocs(
    tmp_path: Path,
) -> None:
    """When some GoCs are all-zero and others have real values, the warning
    surfaces and names *only* the all-zero ones — not the healthy ones.
    """
    path = tmp_path / "2025.12.31_AAI_Ceded.xlsx"
    _build_ceded_workbook(
        path, [
            ("IT_KEEP", 2024, "Direct", 100.0, 50.0),
            ("IT_DROP", 2024, "Direct", 0.0, 0.0),
            ("IT_DROP", 2023, "Direct", 0.0, 0.0),
            ("IT_KEEP", 2023, "Direct", 0.0, 0.0),  # keeps: has a non-zero elsewhere
        ],
    )
    issues = validate_ceded_assumed_file(path)
    warnings = [i for i in issues if i.code == "ALL_ZERO_VALUES"]
    assert len(warnings) == 1
    msg = warnings[0].message
    assert "IT_DROP" in msg
    assert "IT_KEEP" not in msg


def test_validate_ceded_assumed_no_all_zero_warning_when_all_healthy(
    tmp_path: Path,
) -> None:
    """No warning when every GoC has at least one non-zero SINISTRI or
    RISERVA_SINISTRI value.
    """
    path = tmp_path / "2025.12.31_AAI_Ceded.xlsx"
    _build_ceded_workbook(
        path, [
            ("IT05PABPPLE", 2024, "Direct", 100.0, 50.0),
            ("IT06ABCDE", 2024, "Direct", 0.0, 25.0),
        ],
    )
    issues = validate_ceded_assumed_file(path)
    assert not any(i.code == "ALL_ZERO_VALUES" for i in issues)


def test_validate_transcodifica_ok(tmp_path: Path) -> None:
    path = tmp_path / "Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(
        path, [("IT05PABPPLE2", "A", "B", "H")],
    )
    issues = validate_transcodifica_file(path)
    # H present, single row, length is 12 (warning expected for GOC length)
    error_codes = {i.code for i in issues if i.severity == "error"}
    assert error_codes == set()


def test_validate_transcodifica_invalid_hnh(tmp_path: Path) -> None:
    path = tmp_path / "Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(
        path, [("IT05PABPPLE", "A", "B", "WAT")],
    )
    issues = validate_transcodifica_file(path)
    assert any(i.code == "INVALID_HNH" for i in issues)


def test_validate_transcodifica_no_h(tmp_path: Path) -> None:
    path = tmp_path / "Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(
        path, [("IT05PABPPLE", "A", "B", "NH")],
    )
    issues = validate_transcodifica_file(path)
    assert any(i.code == "NO_H_FLAGGED_GOCS" for i in issues)


def test_validate_transcodifica_duplicate_goc(tmp_path: Path) -> None:
    path = tmp_path / "Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(
        path,
        [
            ("IT05PABPPLE", "A", "B", "H"),
            ("IT05PABPPLE", "A", "B", "H"),
        ],
    )
    issues = validate_transcodifica_file(path)
    assert any(i.code == "DUPLICATE_GOC" for i in issues)


def test_validate_payment_patterns_ok(tmp_path: Path) -> None:
    path = tmp_path / "Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns(path)
    issues = validate_payment_patterns_file(path)
    assert issues == []


def test_validate_payment_patterns_missing_sheets(tmp_path: Path) -> None:
    path = tmp_path / "Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns(path, include_pp=False)
    issues = validate_payment_patterns_file(path)
    assert any(
        i.code == "MISSING_SHEET" and "pp_AAI_REINS" in i.message
        for i in issues
    )


def test_validate_payment_patterns_missing_year_cols(tmp_path: Path) -> None:
    path = tmp_path / "Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns(path, ra_year_headers=["foo", "bar"])
    issues = validate_payment_patterns_file(path)
    assert any(i.code == "MISSING_YEAR_COLUMNS" for i in issues)


# ---------------------------------------------------------------------------
# Per-file validators — Astra
# ---------------------------------------------------------------------------
def test_validate_projection_parameters_missing_target_param(
    tmp_path: Path,
) -> None:
    path = tmp_path / "PROJECTION_PARAMETERS_ENTITY.csv"
    _write_csv(
        path,
        [
            ("PARAMETER", "VALUE"),
            ("CF_TIMESTEP", "YEARLY"),
            # REPORTING_MONTH and FX_* parameters absent
        ],
    )
    issues = validate_projection_parameters_entity_file(path)
    codes = {i.code for i in issues}
    assert "MISSING_PARAMETER" in codes
    assert all(i.severity == "warning" for i in issues)


def test_validate_projection_parameters_bad_header(tmp_path: Path) -> None:
    path = tmp_path / "PROJECTION_PARAMETERS_ENTITY.csv"
    _write_csv(
        path,
        [
            ("WRONG", "VALUE"),
            ("CF_TIMESTEP", "YEARLY"),
        ],
    )
    issues = validate_projection_parameters_entity_file(path)
    assert any(
        i.code == "HEADER_MISMATCH" and i.severity == "error"
        for i in issues
    )


def test_validate_mp_goc_seg_short_goc_warning(tmp_path: Path) -> None:
    path = tmp_path / "MP_GOC_SEG.csv"
    _write_csv(
        path,
        [
            ("GOC_SEG_ID", "GOC_ID", "SEG_ID", "ALLOCATION_RATIO"),
            ("IT01_02", "IT01", "02_P&C", 1),
        ],
    )
    issues = validate_mp_goc_seg_file(path)
    assert any(i.code == "GOC_ID_TOO_SHORT" for i in issues)


def test_validate_mp_goc_not_enough_columns(tmp_path: Path) -> None:
    path = tmp_path / "MP_GOC.csv"
    _write_csv(
        path,
        [
            ("GOC_ID", "A", "ANNUAL_COHORT"),  # only 3 cols
            ("IT01PABPPLE", "M", 2024),
        ],
    )
    issues = validate_mp_goc_file(path)
    assert any(i.code == "NOT_ENOUGH_COLUMNS" for i in issues)


def test_validate_mp_goc_ok(tmp_path: Path) -> None:
    path = tmp_path / "MP_GOC.csv"
    header = (
        "GOC_ID", "MEASUREMENT_MODEL", "ANNUAL_COHORT", "AOM_ID",
        "INCEPTION_CURVE_ID", "TIMING_INCEPTION_CURVE",
        "CSM_RELEASE_RATIO_CURVE_ID", "SHARE_TECH_EXP_ENTITY_SHARE",
        "OCI_OPTION", "OCI_OPTION_LRC", "OCI_OPTION_LIC", "GOC_DURATION",
        "GOC_TYPE_IF_NB", "GOC_CURRENCY", "REPORTING_CURRENCY",
        "GOC_TYPE_REINSURANCE", "AGGREG_1_ID", "AGGREG_2_ID",
    )
    _write_csv(
        path,
        [
            header,
            ("IT01PABPPLE", "PAA", 2024, 20) + (None,) * 14,
        ],
    )
    issues = validate_mp_goc_file(path)
    assert issues == []


def test_validate_actuarial_aom_impact_ok(tmp_path: Path) -> None:
    path = tmp_path / "ACTUARIAL_AOM_IMPACT.csv"
    _write_csv(
        path,
        [
            ("GOC_ID", "STEP_ID", 1),
            ("IT01PABPPLE", "DA_LIC_OP", 0),
        ],
    )
    issues = validate_actuarial_aom_impact_file(path)
    assert issues == []


def test_validate_curve_id_param_missing_variable_warning(tmp_path: Path) -> None:
    path = tmp_path / "CURVE_ID_PARAM.csv"
    _write_csv(
        path,
        [
            ("GOC_ID", "VARIABLE_NAME", 1),
            ("IT01PABPPLE", "CLOSING_CURVE_ID", None),
            # OPENING_CURVE_ID and CREDITED_RATE_CURVE_ID absent
        ],
    )
    issues = validate_curve_id_param_file(path)
    missing = {i.message for i in issues if i.code == "MISSING_VARIABLE_NAME"}
    assert any("OPENING_CURVE_ID" in m for m in missing)
    assert any("CREDITED_RATE_CURVE_ID" in m for m in missing)


# ---------------------------------------------------------------------------
# Aggregate validators
# ---------------------------------------------------------------------------
def test_validate_sunrise_inputs_blocks_when_required_files_missing(
    tmp_path: Path,
) -> None:
    ceded_curr = tmp_path / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_ceded_workbook(ceded_curr, [("IT01PABPPLE", 2024, "D", 100.0, 50.0)])

    report = validate_sunrise_inputs(
        [ceded_curr], year=2025, semester=2,
    )
    codes = {i.code for i in report.errors}
    assert "MISSING_TRANSCODIFICA" in codes
    assert "MISSING_PAYMENT_PATTERNS" in codes
    assert "MISSING_CEDED_FOR_DATE" in codes
    assert report.is_blocking


def test_validate_sunrise_inputs_all_present_no_errors(tmp_path: Path) -> None:
    ceded_curr = tmp_path / "1.1_2025.12.31_AAI_Ceded.xlsx"
    _build_ceded_workbook(ceded_curr, [("IT01PABPPLE", 2024, "D", 100.0, 50.0)])
    ceded_prev = tmp_path / "1.2_2024.12.31_AAI_Ceded.xlsx"
    _build_ceded_workbook(ceded_prev, [("IT01PABPPLE", 2023, "D", 80.0, 40.0)])
    transcodifica = tmp_path / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(
        transcodifica, [("IT01PABPPLE", "A", "B", "H")],
    )
    payments = tmp_path / "1.4_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns(payments)

    report = validate_sunrise_inputs(
        [ceded_curr, ceded_prev, transcodifica, payments],
        year=2025, semester=2,
    )
    assert report.errors == []
    assert not report.is_blocking


def test_validate_sunrise_inputs_hy_uses_june_30(tmp_path: Path) -> None:
    ceded_curr = tmp_path / "1.1_2025.06.30_AAI_Ceded.xlsx"
    _build_ceded_workbook(ceded_curr, [("IT01PABPPLE", 2024, "D", 100.0, 50.0)])
    ceded_prev = tmp_path / "1.2_2024.06.30_AAI_Ceded.xlsx"
    _build_ceded_workbook(ceded_prev, [("IT01PABPPLE", 2023, "D", 80.0, 40.0)])
    transcodifica = tmp_path / "1.3_Transcodifica_aggregazione_GOC_H_NH.csv"
    _build_transcodifica_csv(
        transcodifica, [("IT01PABPPLE", "A", "B", "H")],
    )
    payments = tmp_path / "1.4_Payment_Patterns_&_Risk_Adjustments.xlsx"
    _build_payment_patterns(payments)

    report = validate_sunrise_inputs(
        [ceded_curr, ceded_prev, transcodifica, payments],
        year=2025, semester=1,
    )
    assert report.errors == []


def test_validate_astra_inputs_blocks_when_required_missing(
    tmp_path: Path,
) -> None:
    only_one = tmp_path / "PROJECTION_PARAMETERS_ENTITY.csv"
    _write_csv(
        only_one,
        [
            ("PARAMETER", "VALUE"),
            ("CF_TIMESTEP", "YEARLY"),
            ("REPORTING_MONTH", "12_DECEMBER"),
            ("FX_OPENING_DATE", "x"),
            ("FX_AVERAGE_DATE", "x"),
            ("FX_CLOSING_DATE", "x"),
            ("FX_REPORTING_DATE", "x"),
        ],
    )
    report = validate_astra_inputs([only_one])
    codes = {i.code for i in report.errors}
    # 4 of the 5 expected suffixes are missing
    assert "MISSING_MP_GOC_SEG" in codes
    assert "MISSING_MP_GOC" in codes
    assert "MISSING_ACTUARIAL_AOM_IMPACT" in codes
    assert "MISSING_CURVE_ID_PARAM" in codes
    assert report.is_blocking


def test_validate_astra_inputs_all_present(tmp_path: Path) -> None:
    pp = tmp_path / "PROJECTION_PARAMETERS_ENTITY.csv"
    _write_csv(
        pp,
        [
            ("PARAMETER", "VALUE"),
            ("CF_TIMESTEP", "YEARLY"),
            ("REPORTING_MONTH", "12_DECEMBER"),
            ("FX_OPENING_DATE", "x"),
            ("FX_AVERAGE_DATE", "x"),
            ("FX_CLOSING_DATE", "x"),
            ("FX_REPORTING_DATE", "x"),
        ],
    )
    seg = tmp_path / "MP_GOC_SEG.csv"
    _write_csv(
        seg,
        [
            ("GOC_SEG_ID", "GOC_ID", "SEG_ID", "ALLOCATION_RATIO"),
            ("IT01PABPPLE_02", "IT01PABPPLE", "02_P&C", 1),
        ],
    )
    mp_goc = tmp_path / "MP_GOC.csv"
    header = (
        "GOC_ID", "MEASUREMENT_MODEL", "ANNUAL_COHORT", "AOM_ID",
        "INCEPTION_CURVE_ID", "TIMING_INCEPTION_CURVE",
        "CSM_RELEASE_RATIO_CURVE_ID", "SHARE_TECH_EXP_ENTITY_SHARE",
        "OCI_OPTION", "OCI_OPTION_LRC", "OCI_OPTION_LIC", "GOC_DURATION",
        "GOC_TYPE_IF_NB", "GOC_CURRENCY", "REPORTING_CURRENCY",
        "GOC_TYPE_REINSURANCE", "AGGREG_1_ID", "AGGREG_2_ID",
    )
    _write_csv(
        mp_goc,
        [header, ("IT01PABPPLE", "PAA", 2024, 20) + (None,) * 14],
    )
    aom = tmp_path / "ACTUARIAL_AOM_IMPACT.csv"
    _write_csv(
        aom,
        [
            ("GOC_ID", "STEP_ID", 1),
            ("IT01PABPPLE", "DA_LIC_OP", 0),
        ],
    )
    curve = tmp_path / "CURVE_ID_PARAM.csv"
    _write_csv(
        curve,
        [
            ("GOC_ID", "VARIABLE_NAME", 1),
            ("IT01PABPPLE", "CLOSING_CURVE_ID", None),
            ("IT01PABPPLE", "OPENING_CURVE_ID", None),
            ("IT01PABPPLE", "CREDITED_RATE_CURVE_ID", None),
        ],
    )

    report = validate_astra_inputs([pp, seg, mp_goc, aom, curve])
    assert report.errors == []
    assert not report.is_blocking


def test_validate_astra_inputs_duplicate_upload_is_error(tmp_path: Path) -> None:
    pp_a = tmp_path / "a_PROJECTION_PARAMETERS_ENTITY.csv"
    pp_b = tmp_path / "b_PROJECTION_PARAMETERS_ENTITY.csv"
    for p in (pp_a, pp_b):
        _write_csv(
            p,
            [
                ("PARAMETER", "VALUE"),
                ("CF_TIMESTEP", "YEARLY"),
                ("REPORTING_MONTH", "12_DECEMBER"),
                ("FX_OPENING_DATE", "x"),
                ("FX_AVERAGE_DATE", "x"),
                ("FX_CLOSING_DATE", "x"),
                ("FX_REPORTING_DATE", "x"),
            ],
        )
    report = validate_astra_inputs([pp_a, pp_b])
    assert any(
        i.code == "DUPLICATE_PROJECTION_PARAMETERS_ENTITY"
        for i in report.errors
    )
